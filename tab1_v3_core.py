"""Core logic for Tab 1 v3 project-quality appraisal.

This module is intentionally Streamlit-free so the same v3 rubric logic can be
used by the Streamlit experimental tab and by the Custom GPT action backend.
"""

from __future__ import annotations

import io
import json
import os
import re
import tempfile
import time
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Any, Callable

import pandas as pd

RUBRICA_V3_PATH = "./Rubrica_Tab1_Detallada_Full_v3.xlsx"
SHEET_NAME = "Rúbrica Tab 1"
HEADER_ROW_INDEX = 1
MODEL = "gpt-5-mini"
MAX_WORKERS = 48
MAX_COMPLETION_TOKENS = 4000
STABILITY_REPEATS = 10
STABILITY_THRESHOLD_PCT = 80.0
MAX_REPEAT_CALL_RETRIES = 4
RETRY_BACKOFF_SECONDS = 1.5
ANSWER_ORDER = ["Yes", "Partial", "No", "Not Found", "N/A", "Error"]

RESULT_COLUMNS = [
    "ID",
    "Subsección",
    "Pregunta orientadora (no evaluada)",
    "Criterio",
    "Subjetividad",
    "Transversales",
    "Respuesta",
    "N corridas",
    "Conteo modal",
    "Estabilidad (%)",
    "Estable (>=80%)",
    "Resultado Alternativo",
    "Distribución de respuestas",
    "Corridas con error",
    "Razonamiento",
    "Evidencia",
    "Status",
    "Revisión humana recomendada",
]

# ════════════════════════════════════════════════════════════════════════
# Idioma de salida. La rúbrica autoritativa sigue siendo la española: el
# idioma sólo cambia la PRESENTACIÓN, nunca la evaluación. Un mismo PRODOC
# produce los mismos veredictos en "es" y en "en".
# ════════════════════════════════════════════════════════════════════════
DEFAULT_LANG = "es"
LANGS = ("es", "en")


def _lang(lang: str | None) -> str:
    """Normalise to a supported language, defaulting to Spanish."""
    code = str(lang or DEFAULT_LANG).strip().lower()[:2]
    return code if code in LANGS else DEFAULT_LANG


# Etiquetas de presentación. Las claves internas del resultado NO cambian
# (siguen en español): sólo se traducen al escribir el Excel.
LABELS: dict[str, dict[str, str]] = {
    "es": {
        "sheet_result": "Resultado Diagnostico",
        "sheet_priority": "Revisión prioritaria",
        "sheet_rubric": "Rubrica aplicada",
        "why": "POR QUÉ",
        "why_generic": "POR QUÉ ESTE RESULTADO",
        "checks_met": "Se cumplen {ok} de {total} chequeos.",
        "missing": "Falta",
        "and_more": "y {n} más",
        "verification": "VERIFICACIÓN",
        "stability": "ESTABILIDAD",
        "runs_agreed": "{modal} de {total} corridas coincidieron",
        "alt_result": "Resultado alternativo",
        "rule": "REGLA",
        "rule_all": "se requieren los {n} chequeos",
        "rule_atleast": "se requieren al menos {k} de {n} chequeos",
        "rule_mixed": "regla compuesta; cumplidos {ok} de {n}",
        "prior_unstable": "Inestable ({pct} < {thr})",
        "prior_no": "Resultado No",
        "prior_none": ("Ningún criterio exige revisión obligatoria: no hubo "
                       "resultados inestables ni veredictos «No»."),
        "yes": "Sí",
        "no": "No",
        "review_yes": "Sí",
        "review_no": "No - revisión muestral",
        "review_subj": "subjetividad alta",
        "review_stab": "estabilidad <{thr} ({pct})",
        "review_verdict": "resultado requiere revisión",
        "modal_prefix": ("Resultado modal tras {total} corridas independientes: "
                         "{ans} ({modal}/{total}; {pct} de estabilidad). "
                         "Distribución: {dist}."),
        "unstable_note": " Resultado inestable (<{thr}).",
        "alt_note": " Resultado alternativo: {drift}.",
        "representative": "Razonamiento representativo de una corrida con el resultado modal:",
        "no_runs": "No se recibieron corridas para este criterio.",
        "tie": "empate: ",
        "remaining": "{n}/{rest} restantes; {pct} total",
        "prior_placeholder": "—",
    },
    "en": {
        "sheet_result": "Appraisal Result",
        "sheet_priority": "Priority Review",
        "sheet_rubric": "Rubric Applied",
        "why": "WHY",
        "why_generic": "WHY THIS RESULT",
        "checks_met": "{ok} of {total} checks are met.",
        "missing": "Missing",
        "and_more": "and {n} more",
        "verification": "VERIFICATION",
        "stability": "STABILITY",
        "runs_agreed": "{modal} of {total} runs agreed",
        "alt_result": "Alternative result",
        "rule": "RULE",
        "rule_all": "all {n} checks are required",
        "rule_atleast": "at least {k} of {n} checks are required",
        "rule_mixed": "composite rule; {ok} of {n} met",
        "prior_unstable": "Unstable ({pct} < {thr})",
        "prior_no": "Verdict No",
        "prior_none": ("No criterion requires mandatory review: there were no "
                       "unstable results and no \u201cNo\u201d verdicts."),
        "yes": "Yes",
        "no": "No",
        "review_yes": "Yes",
        "review_no": "No - sample check",
        "review_subj": "high subjectivity",
        "review_stab": "stability <{thr} ({pct})",
        "review_verdict": "verdict requires review",
        "modal_prefix": ("Modal result after {total} independent runs: {ans} "
                         "({modal}/{total}; {pct} stability). "
                         "Distribution: {dist}."),
        "unstable_note": " Unstable result (<{thr}).",
        "alt_note": " Alternative result: {drift}.",
        "representative": "Representative reasoning from a run with the modal result:",
        "no_runs": "No runs were received for this criterion.",
        "tie": "tie: ",
        "remaining": "{n}/{rest} remaining; {pct} total",
        "prior_placeholder": "—",
    },
}

# Cabeceras de columna: clave interna (es) -> encabezado mostrado (en).
COLUMN_LABELS_EN = {
    "Subsección": "Subsection",
    "Pregunta orientadora (no evaluada)": "Guiding question (not evaluated)",
    "Criterio": "Criterion",
    "Subjetividad": "Subjectivity",
    "Transversales": "Cross-cutting",
    "Respuesta": "Verdict",
    "N corridas": "Runs",
    "Conteo modal": "Modal count",
    "Estabilidad (%)": "Stability (%)",
    "Estable (>=80%)": "Stable (>=80%)",
    "Resultado Alternativo": "Alternative result",
    "Distribución de respuestas": "Verdict distribution",
    "Corridas con error": "Runs with error",
    "Razonamiento": "Reasoning",
    "Evidencia": "Evidence",
    "Revisión humana recomendada": "Human review recommended",
    "Motivo de prioridad": "Reason for priority",
    "Sección": "Section",
    "Criterio a evaluar": "Criterion",
    "Tipo de criterio": "Criterion type",
    "Aplicabilidad": "Applicability",
    "Aspectos transversales": "Cross-cutting aspects",
    "Elementos a verificar": "Elements to verify",
    "Rúbrica — Sí": "Rubric — Yes",
    "Rúbrica — Parcial": "Rubric — Partial",
    "Rúbrica — No": "Rubric — No",
    "Rúbrica — No aplica": "Rubric — N/A",
    "Anclas verificables (v3)": "Verifiable anchors (v3)",
    "Subjetividad residual (v3)": "Residual subjectivity (v3)",
}


SHEET_RESULT = "Resultado Diagnostico"
SHEET_PRIORITY = "Revisión prioritaria"
SHEET_RUBRIC = "Rubrica aplicada"

# Revisión obligatoria: inestables (la repetición no convergió) y "No"
# (incumplimiento declarado). La subjetividad alta queda a criterio del usuario.
PRIORITY_COLUMNS = [
    "ID",
    "Subsección",
    "Criterio",
    "Motivo de prioridad",
    "Respuesta",
    "Estabilidad (%)",
    "Resultado Alternativo",
    "Distribución de respuestas",
    "Subjetividad",
    "Razonamiento",
    "Evidencia",
]

# Definición del criterio que acompaña a cada veredicto, para que el Excel
# sea autocontenido: quien audita ve la regla junto al resultado.
RUBRIC_SHEET_COLUMNS = [
    "ID",
    "Sección",
    "Subsección",
    "Criterio a evaluar",
    "Tipo de criterio",
    "Aplicabilidad",
    "Aspectos transversales",
    "Elementos a verificar",
    "Rúbrica — Sí",
    "Rúbrica — Parcial",
    "Rúbrica — No",
    "Rúbrica — No aplica",
    "Anclas verificables (v3)",
    "Subjetividad residual (v3)",
]

# Clave interna que produce stability/aggregation, renombrada en la salida.
DRIFT_SOURCE_COLUMN = "Deriva principal (si inestable)"

PUBLIC_RESULT_COLUMNS = [
    "ID",
    "Subsección",
    "Criterio",
    "Subjetividad",
    "Transversales",
    "Resultado de valoración",
    "Estabilidad (%)",
    "Resultado Alternativo",
    "Lectura rápida",
    "Principal oportunidad de mejora",
    "Evidencia clave",
    "Revisión humana recomendada",
]

V3_SYSTEM_PROMPT = """Eres un analista experto en evaluación de documentos de proyecto (PRODOC) de la OIT.

Tu tarea: evaluar UN criterio específico contra el documento provisto, aplicando una rúbrica mecanizada.

ESTRUCTURA DE LA RÚBRICA QUE RECIBES:
- CRITERIO: el enunciado a evaluar.
- PREGUNTA ORIENTADORA: contexto general (NO se evalúa, solo enmarca).
- TIPO: forma del criterio (binario, lista, transversal, etc.).
- APLICABILIDAD: si el criterio aplica siempre o bajo condición.
- ASPECTOS TRANSVERSALES: indica si aplica filtro DEDICADO vs MARCO.
- ELEMENTOS A VERIFICAR: componentes atómicos del criterio.
- RÚBRICA Sí/Parcial/No/No aplica: TESTS atómicos (T1, T2, ...) y DECISIÓN booleana.
- ANCLAS VERIFICABLES: patrones de texto, códigos, nombres concretos a buscar.

PROCESO OBLIGATORIO (en este orden):
1. Si APLICABILIDAD es condicional: primero determina si la condición se cumple. Si no, resultado = "N/A".
2. Localiza en el DOCUMENTO la(s) sección(es) que tratan del criterio. Usa las ANCLAS como guía de búsqueda.
3. Para cada TEST listado en la rúbrica Sí (T1, T2, T3, ...): evalúa explícitamente si se cumple (verdadero/falso) con base en el documento.
4. Aplica la DECISIÓN de Sí primero. Si no se cumple, aplica la DECISIÓN de Parcial. Si tampoco, aplica No.
5. Resultado final: Yes / Partial / No / Not Found / N/A.

DEFINICIÓN CANÓNICA DE «DEDICADO vs MARCO» (idéntica a la usada en Tab 1):

Una mención del sujeto (género, discapacidad, pueblos indígenas, etc.) cuenta como
MARCO (NO cuenta para cumplimiento) cuando es cualquiera de:
  - Mención en el objetivo general / declaración de impacto que enumera varios grupos
  - Listas de partes interesadas / consulta / participantes de investigación
  - Enumeraciones de alcance de seguimiento («…entre otros», «…incluyendo X, Y, Z»)
  - Lenguaje boilerplate de inclusión
  - Cualquier pasaje donde el sujeto aparece en una lista de ≥3 grupos sin seguimiento dedicado

Una mención cuenta como DEDICADO si es cualquiera de:
  A. Sub-objetivo, resultado o producto cuyo título/propósito nombra al sujeto
  B. Indicador desagregado por el sujeto o que lo mide específicamente
  C. Actividad cuyo propósito principal aborda al sujeto
  D. Partida presupuestaria o asignación de recursos para el sujeto
  E. Meta cuantificable relativa al sujeto

Si toda la evidencia citable es MARCO, el resultado DEBE ser "No" o "Not Found",
sin importar cuántas veces se nombre al sujeto.

REGLAS CRÍTICAS:
- NO inventes elementos. Si la rúbrica lista T1/T2/T3, evalúa exactamente esos — no añadas T4 propio.
- El filtro DEDICADO vs MARCO definido arriba aplica SOLO cuando la rúbrica del criterio lo invoque explícitamente (criterios marcados con ASPECTOS TRANSVERSALES ≠ «Ninguno»). Para criterios sin transversales, ignóralo.
- Cita evidencia textual entre comillas. Si la evidencia es ausencia, dilo: «No se encontró sección X».
- Si el documento carece de información para evaluar el criterio, resultado = "Not Found".
- "N/A" solo cuando la APLICABILIDAD condicional no se satisface.
- El Razonamiento DEBE enumerar el resultado de cada TEST seguido de la DECISIÓN aplicada.
- No presentes el resultado como una determinación oficial de la OIT; es una valoración asistida que requiere validación experta.

FORMATO DEL Razonamiento (estricto):
  T1: <verdadero/falso> — <una línea de justificación>
  T2: <verdadero/falso> — <una línea de justificación>
  ...
  DECISIÓN: <regla booleana evaluada con los resultados, p.ej. T1 ∧ T2 ∧ ¬T3 = falso → Parcial>
  RESULTADO: <Yes/No/Partial/Not Found/N/A>

Devuelve SIEMPRE JSON con: {"Respuesta", "Razonamiento", "Evidencia"}. Idioma: español."""


# Frase del prompt base que fija el idioma. En inglés hay que SUSTITUIRLA:
# dejarla y añadir después «escribe en inglés» produce un prompt que se
# contradice, y el modelo sigue la orden en español (toda la rúbrica y el
# prompt de usuario están en español, lo que refuerza ese sesgo).
_ES_LANG_DIRECTIVE = "Idioma: español."
_EN_LANG_DIRECTIVE = (
    "IDIOMA DE SALIDA: INGLÉS. Escribe «Razonamiento» y «Evidencia» en inglés "
    "(ver la sección IDIOMA DE SALIDA al final para las etiquetas que NO se "
    "traducen)."
)

# El andamiaje estructural (verdadero/falso, DECISIÓN:, RESULTADO:) se mantiene
# SIEMPRE en español porque el renderizador lo parsea; el lector nunca lo ve
# (se convierte en ✓/✗). Sólo la prosa cambia de idioma.
V3_SYSTEM_PROMPT_EN_SUFFIX = """

═════ IDIOMA DE SALIDA ═════
Escribe la JUSTIFICACIÓN de cada test y el campo "Evidencia" en INGLÉS.
Las citas textuales del documento van tal cual aparecen (no las traduzcas).

CRÍTICO — no cambies estas etiquetas estructurales, deben seguir en español
aunque el resto esté en inglés:
  - «verdadero» / «falso» tras cada Tn (NO uses true/false)
  - «DECISIÓN:» (NO uses DECISION:)
  - «RESULTADO:» (NO uses RESULT:)
Ejemplo correcto:
  T1: verdadero — The document distinguishes the two approaches in section 3.2.
  T2: falso — No mechanism is articulated.
  DECISIÓN: T1 ∧ T2 = falso → Parcial
  RESULTADO: Partial"""


def system_prompt(lang: str = DEFAULT_LANG) -> str:
    """System prompt for the requested output language."""
    if _lang(lang) != "en":
        return V3_SYSTEM_PROMPT
    assert _ES_LANG_DIRECTIVE in V3_SYSTEM_PROMPT, "cambió la frase de idioma"
    base = V3_SYSTEM_PROMPT.replace(_ES_LANG_DIRECTIVE, _EN_LANG_DIRECTIVE)
    return base + V3_SYSTEM_PROMPT_EN_SUFFIX


V3_USER_PROMPT_TEMPLATE = """═════ CRITERIO A EVALUAR ═════
ID del criterio: {id}
CRITERIO: {criterio}

PREGUNTA ORIENTADORA (solo contexto, NO evaluar): {head}

═════ METADATA ═════
TIPO: {tipo}
APLICABILIDAD: {aplicabilidad}
ASPECTOS TRANSVERSALES: {transv}

═════ ELEMENTOS A VERIFICAR ═════
{elementos}

═════ RÚBRICA — Sí ═════
{si}

═════ RÚBRICA — Parcial ═════
{par}

═════ RÚBRICA — No ═════
{no}

═════ RÚBRICA — No aplica ═════
{na}

═════ ANCLAS VERIFICABLES ═════
{anclas}

═════ DOCUMENTO (PRODOC) ═════
{document_text}

═════ TU TAREA ═════
1. Verifica APLICABILIDAD. Si no aplica → Respuesta = "N/A".
2. Evalúa cada TEST de la rúbrica de Sí con base en el DOCUMENTO.
3. Aplica DECISIÓN Sí → Parcial → No en orden.
4. Devuelve JSON con Respuesta + Razonamiento (con todos los TESTS enumerados + DECISIÓN + RESULTADO) + Evidencia (citas textuales)."""


V3_RESPONSE_SCHEMA = {
    "name": "rubric_eval_v3",
    "schema": {
        "type": "object",
        "properties": {
            "Respuesta": {
                "type": "string",
                "enum": ["Yes", "No", "Partial", "Not Found", "N/A"],
            },
            "Razonamiento": {"type": "string"},
            "Evidencia": {"type": "string"},
        },
        "required": ["Respuesta", "Razonamiento", "Evidencia"],
        "additionalProperties": False,
    },
    "strict": True,
}


def load_rubrica_v3(rubric_path: str = RUBRICA_V3_PATH) -> pd.DataFrame:
    """Load the v3 rubric xlsx and return one row per criterion."""
    df = pd.read_excel(rubric_path, sheet_name=SHEET_NAME, header=HEADER_ROW_INDEX)
    df = df.dropna(subset=["ID"]).reset_index(drop=True)
    df["ID"] = df["ID"].astype(str)
    return df


def _id_sort_key(id_str: str) -> tuple[int, ...]:
    return tuple(int(p) for p in str(id_str).split(".") if p.isdigit())


def _extract_section(id_str: str) -> int:
    m = re.match(r"(\d+)", str(id_str))
    return int(m.group(1)) if m else 0


def _extract_subsection(id_str: str) -> str:
    m = re.match(r"(\d+\.\d+)", str(id_str))
    return m.group(1) if m else ""


def extract_docx_text_from_path(docx_path: str) -> str:
    """Extract full text from a DOCX path."""
    try:
        from docx2python import docx2python
    except ImportError as exc:
        raise RuntimeError(
            "docx2python is required for DOCX extraction. Install project requirements before running evaluations."
        ) from exc

    result = docx2python(docx_path)
    return result.text or ""


def extract_docx_text_from_bytes(docx_bytes: bytes) -> str:
    """Extract full text from DOCX bytes."""
    with tempfile.NamedTemporaryFile(delete=False, suffix=".docx") as tmp:
        tmp.write(docx_bytes)
        tmp_path = tmp.name
    try:
        return extract_docx_text_from_path(tmp_path)
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass


def filter_rubric(
    df_rubric: pd.DataFrame,
    sections: list[int] | None = None,
    subsections: list[str] | None = None,
) -> pd.DataFrame:
    """Filter rubric rows by section and/or subsection."""
    df = df_rubric.copy()
    df["_section"] = df["ID"].apply(_extract_section)
    df["_subsection"] = df["ID"].apply(_extract_subsection)
    if sections:
        df = df[df["_section"].isin([int(s) for s in sections])]
    if subsections:
        df = df[df["_subsection"].isin([str(s) for s in subsections])]
    return df.reset_index(drop=True)


_EN_USER_REMINDER = """

═════ RECORDATORIO DE IDIOMA ═════
La rúbrica de arriba está en español, pero la SALIDA va en INGLÉS:
«Razonamiento» y «Evidencia» en inglés. Mantén en español sólo las etiquetas
estructurales: verdadero / falso, DECISIÓN:, RESULTADO:."""


def build_user_prompt(
    row: pd.Series, document_text: str, lang: str = DEFAULT_LANG
) -> str:
    """Inject one criterion's rubric into the user prompt template."""

    def get(col: str, default: str = "") -> str:
        v = row.get(col, default)
        if pd.isna(v) or v is None:
            return default
        s = str(v).strip()
        return s if s else default

    na_text = get("Rúbrica — No aplica", "")
    if not na_text:
        na_text = "(esta rúbrica no contempla la categoría «No aplica»)"

    return V3_USER_PROMPT_TEMPLATE.format(
        id=get("ID"),
        criterio=get("Criterio a evaluar"),
        head=get("Pregunta orientadora (CONTEXTO — no evaluar)"),
        tipo=get("Tipo de criterio"),
        aplicabilidad=get("Aplicabilidad"),
        transv=get("Aspectos transversales", "Ninguno"),
        elementos=get("Elementos a verificar"),
        si=get("Rúbrica — Sí"),
        par=get("Rúbrica — Parcial"),
        no=get("Rúbrica — No"),
        na=na_text,
        anclas=get("Anclas verificables (v3)"),
        document_text=document_text,
    ) + (_EN_USER_REMINDER if _lang(lang) == "en" else "")


def _extract_usage(resp: Any) -> dict[str, int]:
    """Pull token-usage fields from a chat.completions response (all default to 0)."""
    usage = getattr(resp, "usage", None)
    if usage is None:
        return {
            "usage_prompt_tokens": 0,
            "usage_cached_tokens": 0,
            "usage_completion_tokens": 0,
            "usage_reasoning_tokens": 0,
        }
    prompt_details = getattr(usage, "prompt_tokens_details", None)
    completion_details = getattr(usage, "completion_tokens_details", None)
    return {
        "usage_prompt_tokens": getattr(usage, "prompt_tokens", 0) or 0,
        "usage_cached_tokens": getattr(prompt_details, "cached_tokens", 0) or 0,
        "usage_completion_tokens": getattr(usage, "completion_tokens", 0) or 0,
        "usage_reasoning_tokens": getattr(completion_details, "reasoning_tokens", 0) or 0,
    }


def evaluate_criterion(
    client: Any, row: pd.Series, document_text: str, lang: str = DEFAULT_LANG
) -> dict[str, Any]:
    """Run a single criterion through the v3 prompt and return a result dict."""
    crit_id = str(row.get("ID", ""))
    subj = str(row.get("Subjetividad residual (v3)", "Media")).strip()
    effort = "medium" if subj == "Alta" else "minimal"

    base = {
        "ID": crit_id,
        "Subsección": _extract_subsection(crit_id),
        "Pregunta orientadora (no evaluada)": str(
            row.get("Pregunta orientadora (CONTEXTO — no evaluar)", "")
        ),
        "Criterio": str(row.get("Criterio a evaluar", "")),
        "Tipo": str(row.get("Tipo de criterio", "")),
        "Subjetividad": subj,
        "Transversales": str(row.get("Aspectos transversales", "Ninguno")),
        "reasoning_effort": effort,
        "max_completion_tokens": MAX_COMPLETION_TOKENS,
        "usage_prompt_tokens": 0,
        "usage_cached_tokens": 0,
        "usage_completion_tokens": 0,
        "usage_reasoning_tokens": 0,
    }

    if not document_text or not document_text.strip():
        return {
            **base,
            "Respuesta": "Not Found",
            "Razonamiento": "Documento vacío.",
            "Evidencia": "",
            "Status": "Success",
        }

    try:
        resp = client.chat.completions.create(
            model=MODEL,
            messages=[
                {"role": "system", "content": system_prompt(lang)},
                {"role": "user", "content": build_user_prompt(row, document_text, lang)},
            ],
            max_completion_tokens=MAX_COMPLETION_TOKENS,
            reasoning_effort=effort,
            response_format={"type": "json_schema", "json_schema": V3_RESPONSE_SCHEMA},
        )
        content = (resp.choices[0].message.content or "").strip()
        if not content:
            return {
                **base,
                **_extract_usage(resp),
                "Respuesta": "Error",
                "Razonamiento": "Respuesta vacía del modelo.",
                "Evidencia": "",
                "Status": "Error",
            }
        result = json.loads(content)
        return {
            **base,
            **_extract_usage(resp),
            "Respuesta": result.get("Respuesta", "Not Found"),
            "Razonamiento": result.get("Razonamiento", ""),
            "Evidencia": result.get("Evidencia", ""),
            "Status": "Success",
        }
    except Exception as e:  # noqa: BLE001 - contain failures per criterion
        return {
            **base,
            "Respuesta": "Error",
            "Razonamiento": f"Error en evaluación v3: {e}",
            "Evidencia": "",
            "Status": "Error",
        }


def _answer_order_index(answer: str) -> int:
    try:
        return ANSWER_ORDER.index(str(answer))
    except ValueError:
        return len(ANSWER_ORDER)


def _format_pct(value: float) -> str:
    return f"{value:.0f}%" if float(value).is_integer() else f"{value:.1f}%"


def _distribution_text(counts: Counter[str]) -> str:
    parts = []
    for answer in ANSWER_ORDER:
        count = counts.get(answer, 0)
        if count:
            parts.append(f"{answer}={count}")
    for answer in sorted(set(counts) - set(ANSWER_ORDER)):
        parts.append(f"{answer}={counts[answer]}")
    return "; ".join(parts)


def _drift_text(
    counts: Counter[str],
    modal_answer: str,
    total: int,
    stable: bool,
    lang: str = DEFAULT_LANG,
) -> str:
    if stable:
        return ""

    remaining = total - counts.get(modal_answer, 0)
    if remaining <= 0:
        return ""

    non_modal = {
        answer: count
        for answer, count in counts.items()
        if answer != modal_answer and count > 0
    }
    if not non_modal:
        return ""

    max_count = max(non_modal.values())
    top_answers = sorted(
        [answer for answer, count in non_modal.items() if count == max_count],
        key=_answer_order_index,
    )
    L = LABELS[_lang(lang)]
    label = " / ".join(top_answers)
    tie_prefix = L["tie"] if len(top_answers) > 1 else ""
    total_pct = 100 * max_count / total
    rest = L["remaining"].format(
        n=max_count, rest=remaining, pct=_format_pct(total_pct)
    )
    return f"{label} ({tie_prefix}{rest})"


def _representative_modal_result(
    repeated_results: list[dict[str, Any]],
    modal_answer: str,
) -> dict[str, Any]:
    candidates = [
        result
        for result in repeated_results
        if str(result.get("Respuesta", "Error")) == modal_answer
    ] or repeated_results

    return max(
        candidates,
        key=lambda result: (
            result.get("Status") == "Success",
            len(str(result.get("Evidencia", ""))),
            -int(result.get("repeat", 0) or 0),
        ),
    )


def aggregate_repeated_criterion_results(
    repeated_results: list[dict[str, Any]],
    stability_threshold_pct: float = STABILITY_THRESHOLD_PCT,
    lang: str = DEFAULT_LANG,
) -> dict[str, Any]:
    """Collapse repeated criterion evaluations into one modal result plus stability."""
    L = LABELS[_lang(lang)]
    if not repeated_results:
        return {
            "Respuesta": "Error",
            "N corridas": 0,
            "Conteo modal": 0,
            "Estabilidad (%)": 0.0,
            "Estable (>=80%)": "No",
            "Deriva principal (si inestable)": "",
            "Distribución de respuestas": "",
            "Corridas con error": 0,
            "Razonamiento": L["no_runs"],
            "Evidencia": "",
            "Status": "Error",
        }

    repeated_results = sorted(
        repeated_results,
        key=lambda result: int(result.get("repeat", 0) or 0),
    )
    counts: Counter[str] = Counter(
        str(result.get("Respuesta", "Error")) for result in repeated_results
    )
    total = len(repeated_results)
    modal_answer = min(
        counts,
        key=lambda answer: (-counts[answer], _answer_order_index(answer)),
    )
    modal_count = counts[modal_answer]
    stability_pct = round(100 * modal_count / total, 1)
    stable = stability_pct >= stability_threshold_pct
    drift = _drift_text(counts, modal_answer, total, stable, lang)
    distribution = _distribution_text(counts)
    representative = _representative_modal_result(repeated_results, modal_answer)
    error_count = sum(
        1
        for result in repeated_results
        if result.get("Status") == "Error" or str(result.get("Respuesta")) == "Error"
    )

    reasoning_prefix = L["modal_prefix"].format(
        total=total, ans=modal_answer, modal=modal_count,
        pct=_format_pct(stability_pct), dist=distribution,
    )
    if not stable:
        reasoning_prefix += L["unstable_note"].format(
            thr=_format_pct(stability_threshold_pct)
        )
        if drift:
            reasoning_prefix += L["alt_note"].format(drift=drift)

    return {
        **representative,
        "Respuesta": modal_answer,
        "N corridas": total,
        "Conteo modal": modal_count,
        "Estabilidad (%)": stability_pct,
        "Estable (>=80%)": L["yes"] if stable else L["no"],
        "Deriva principal (si inestable)": drift,
        "Distribución de respuestas": distribution,
        "Corridas con error": error_count,
        "Razonamiento": (
            f"{reasoning_prefix}\n\n"
            f"{L['representative']}\n"
            f"{representative.get('Razonamiento', '')}"
        ).strip(),
        "Status": "Error" if modal_answer == "Error" else "Success",
    }


def evaluate_criterion_with_retries(
    client: Any,
    row: pd.Series,
    document_text: str,
    max_retries: int = MAX_REPEAT_CALL_RETRIES,
    lang: str = DEFAULT_LANG,
) -> dict[str, Any]:
    """Run one criterion call, retrying only model/API failures reported as Error."""
    result = evaluate_criterion(client, row, document_text, lang)
    for attempt in range(1, max_retries + 1):
        if result.get("Status") != "Error":
            break
        time.sleep(RETRY_BACKOFF_SECONDS * attempt)
        result = evaluate_criterion(client, row, document_text, lang)
    return result


def evaluate_criteria(
    client: Any,
    df_criteria: pd.DataFrame,
    document_text: str,
    max_workers: int = MAX_WORKERS,
    progress_callback: Callable[[int, int], None] | None = None,
    lang: str = DEFAULT_LANG,
) -> list[dict[str, Any]]:
    """Evaluate all rows using the standard 10-repeat stability scheme."""
    lang = _lang(lang)
    total_criteria = len(df_criteria)
    results: list[dict[str, Any]] = []
    if total_criteria == 0:
        return results

    rows = [(str(row["ID"]), row.copy()) for _, row in df_criteria.iterrows()]
    total_calls = total_criteria * STABILITY_REPEATS
    repeated_by_id: dict[str, list[dict[str, Any]]] = {
        crit_id: [] for crit_id, _ in rows
    }
    workers = max(1, min(int(max_workers), total_calls))
    with ThreadPoolExecutor(max_workers=workers) as ex:
        futures = {}
        for crit_id, row in rows:
            for repeat in range(1, STABILITY_REPEATS + 1):
                futures[
                    ex.submit(
                        evaluate_criterion_with_retries,
                        client, row, document_text, MAX_REPEAT_CALL_RETRIES, lang,
                    )
                ] = (crit_id, repeat)

        done = 0
        for fut in as_completed(futures):
            crit_id, repeat = futures[fut]
            result = fut.result()
            result["repeat"] = repeat
            repeated_by_id[crit_id].append(result)
            done += 1
            if progress_callback:
                progress_callback(done, total_calls)

    results = [
        _with_readable_reasoning(
            aggregate_repeated_criterion_results(
                repeated_by_id[crit_id], STABILITY_THRESHOLD_PCT, lang
            ),
            row,
            lang,
        )
        for crit_id, row in rows
    ]
    results.sort(key=lambda r: _id_sort_key(r["ID"]))
    return results


# ════════════════════════════════════════════════════════════════════════
# Lectura del Razonamiento: sustituye las etiquetas T1/T2/... por el texto
# del chequeo que representan. No cambia la lógica ni la rúbrica: sólo el
# formato de salida. Si el modelo se aparta del formato esperado, se
# devuelve el razonamiento original sin tocar.
# ════════════════════════════════════════════════════════════════════════

_TEST_LABEL = r"(T\d+|[SMART])"
# Definición en la rúbrica:  "T1: ¿Distingue...? (sí/no)"
_TEST_DEF_RE = re.compile(rf"^\s*{_TEST_LABEL}\s*(?:\([^)]*\))?\s*[:—-]\s*(.+?)\s*$", re.M)
# Veredicto del modelo:      "T1: verdadero — justificación"
# Tolerante a viñetas, negritas markdown y separadores ":" / "—" / "-",
# porque el modelo no siempre respeta el formato al pie de la letra.
_TEST_OBS_RE = re.compile(
    rf"^\s*(?:[-*•]\s+)?\*{{0,2}}_{{0,2}}{_TEST_LABEL}_{{0,2}}\*{{0,2}}\s*[:—–-]\s*"
    rf"(verdadero|falso|true|false|s[íi]|yes|no|n/?a|no aplica)\b[\s.,;—–:-]*(.*)$",
    re.M | re.I,
)
_DECISION_RE = re.compile(r"DECISI[ÓO]N\s*:\s*(.+)")
_TRUE_WORDS = {"verdadero", "sí", "si", "true", "yes"}
_NA_WORDS = {"n/a", "na", "no aplica"}


def _clean_test_text(text: str) -> str:
    """Trim answer-format hints and resolve T-references inside a statement.

    A few rubric statements refer to sibling checks by label ("vincula la
    intervención a T1/T2/T3"). Left as-is they reintroduce exactly the lookup
    the rendering removes, so they are turned into plain back-references. The
    rubric itself is untouched: the model still sees the labels in the prompt.
    """
    text = re.sub(r"\s*\((?:s[íi]\s*/\s*no|verdadero\s*/\s*falso)[^)]*\)\s*$", "", text, flags=re.I)
    text = re.sub(r"\bT\d(?:\s*/\s*T\d)+\b", "los chequeos anteriores", text)
    text = re.sub(r"\bT(\d)\b", r"el chequeo \1", text)
    text = re.sub(r"\bde el chequeo\b", "del chequeo", text)
    text = re.sub(r"\ba el chequeo\b", "al chequeo", text)
    return text.strip().rstrip(".").strip()


def _tidy_quote(text: str) -> str:
    """Collapse escaped newlines the model emits inside quoted document text."""
    text = text.replace("\\n", " ").replace("\r", " ")
    return re.sub(r"[ \t]{2,}", " ", text).strip()


def _as_statement(text: str) -> str:
    """Drop the interrogative wrapper so the text reads inside a sentence."""
    return text.strip().lstrip("¿").rstrip("?").strip()


def parse_rubric_tests(rubrica_si: str) -> dict[str, str]:
    """Map each test label to its statement, as written in the rubric."""
    tests: dict[str, str] = {}
    body = rubrica_si.split("DECISI")[0]
    for label, text in _TEST_DEF_RE.findall(body):
        cleaned = _clean_test_text(text)
        if cleaned and label not in tests:
            tests[label] = cleaned
    return tests


def parse_observed_tests(reasoning: str) -> list[tuple[str, str | None, str]]:
    """Extract (label, verdict, justification) from the model's reasoning."""
    observed: list[tuple[str, str | None, str]] = []
    seen: set[str] = set()
    for label, verdict, justification in _TEST_OBS_RE.findall(reasoning):
        if label in seen:
            continue
        seen.add(label)
        v = verdict.strip().lower()
        state = None if v in _NA_WORDS else (v in _TRUE_WORDS)
        observed.append((label, state, justification.strip()))
    return observed


def _rule_summary(decision: str, n_ok: int, n_total: int, lang: str = DEFAULT_LANG) -> str:
    """One plain-language line describing what the rule required."""
    L = LABELS[_lang(lang)]
    d = decision.strip()
    counting = re.search(r"#\s*verdaderos[^≥>]*[≥>]=?\s*(\d+)", d, re.I)
    if counting:
        return L["rule_atleast"].format(k=counting.group(1), n=n_total)
    if "∨" not in d and "#" not in d and d.count("∧") == max(n_total - 1, 0) and n_total > 1:
        return L["rule_all"].format(n=n_total)
    return L["rule_mixed"].format(ok=n_ok, n=n_total)


def render_reasoning(
    result: dict[str, Any], rubrica_si: str, lang: str = DEFAULT_LANG
) -> str:
    """Rewrite the reasoning so every check reads on its own, without T-labels."""
    L = LABELS[_lang(lang)]
    raw = str(result.get("Razonamiento", "")).strip()
    tests = parse_rubric_tests(rubrica_si)
    observed = parse_observed_tests(raw)
    if not tests or not observed:
        return raw  # formato inesperado: no tocar

    answer = str(result.get("Respuesta", "")).strip()
    n_total = len(observed)
    n_ok = sum(1 for _, state, _ in observed if state is True)
    faltan = [
        _as_statement(tests.get(label, label))
        for label, state, _ in observed
        if state is False
    ]

    lineas: list[str] = []
    cabecera = f"{L['why']} {answer.upper()}" if answer else L["why_generic"]
    resumen = L["checks_met"].format(ok=n_ok, total=n_total)
    if faltan and answer != "Yes":
        falta_txt = "; ".join(faltan[:3])
        if len(faltan) > 3:
            falta_txt += "; " + L["and_more"].format(n=len(faltan) - 3)
        resumen += f" {L['missing']}: {falta_txt}."
    lineas.append(f"{cabecera} · {resumen}")
    lineas.append("")
    lineas.append(L["verification"])
    for label, state, justification in observed:
        marca = "✓" if state is True else ("✗" if state is False else "–")
        enunciado = tests.get(label, f"(chequeo {label})")
        lineas.append(f" {marca} {enunciado}")
        justification = _tidy_quote(justification)
        if justification:
            lineas.append(f"     {justification}")

    n_corridas = result.get("N corridas")
    n_modal = result.get("Conteo modal")
    if n_corridas and n_modal:
        agreed = L["runs_agreed"].format(modal=n_modal, total=n_corridas)
        est = f"{L['stability']} · {agreed}"
        drift = str(result.get(DRIFT_SOURCE_COLUMN, "") or "").strip()
        est += f". {L['alt_result']}: {drift}." if drift else "."
        lineas.append("")
        lineas.append(est)

    decision = _DECISION_RE.search(rubrica_si)
    if decision:
        regla = decision.group(1).strip().splitlines()[0].strip()
        resumen_regla = _rule_summary(regla, n_ok, n_total, lang)
        lineas.append(f"{L['rule']} · {resumen_regla}  ({regla})")

    return "\n".join(lineas)


def _with_readable_reasoning(
    result: dict[str, Any], row: pd.Series, lang: str = DEFAULT_LANG
) -> dict[str, Any]:
    """Apply render_reasoning, falling back to the original text on any error."""
    try:
        rendered = render_reasoning(result, str(row.get("Rúbrica — Sí", "")), lang)
    except Exception:
        return result
    return {**result, "Razonamiento": rendered} if rendered else result


def results_to_dataframe(
    results: list[dict[str, Any]], lang: str = DEFAULT_LANG
) -> pd.DataFrame:
    """Return a sorted dataframe with the public result columns."""
    if not results:
        return pd.DataFrame(columns=RESULT_COLUMNS)
    df = pd.DataFrame(results)
    if DRIFT_SOURCE_COLUMN in df.columns:
        df = df.rename(columns={DRIFT_SOURCE_COLUMN: "Resultado Alternativo"})
    df["Revisión humana recomendada"] = df.apply(_review_flag, axis=1, lang=lang)
    available = [c for c in RESULT_COLUMNS if c in df.columns]
    df_sorted = (
        df.assign(_sk=df["ID"].map(_id_sort_key))
        .sort_values("_sk")
        .drop(columns="_sk")
        .reset_index(drop=True)
    )
    return df_sorted[available]


def _compact_text(value: Any, max_chars: int = 700) -> str:
    """Return readable one-cell text without technical decision scaffolding."""
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    lines = []
    for line in text.splitlines():
        clean = line.strip()
        if not clean:
            continue
        if re.match(r"^T\d+\s*:", clean):
            continue
        if clean.upper().startswith(("DECISIÓN:", "DECISION:", "RESULTADO:", "VEREDICTO:")):
            continue
        lines.append(clean)
    compact = " ".join(lines) if lines else text.replace("\n", " ")
    compact = re.sub(r"\s+", " ", compact).strip()
    if len(compact) <= max_chars:
        return compact
    return compact[: max_chars - 1].rstrip() + "…"


def _quick_reading(answer: str) -> str:
    mapping = {
        "Yes": "Cumple el criterio según la evidencia encontrada.",
        "Partial": "Cumple parcialmente; hay elementos específicos por reforzar.",
        "No": "No cumple el criterio con la evidencia disponible.",
        "Not Found": "No se encontró información suficiente para valorar el criterio.",
        "N/A": "El criterio no aplica según las condiciones de la rúbrica.",
        "Error": "Hubo una falla técnica; revisar o reintentar la evaluación.",
    }
    return mapping.get(str(answer), "Revisar el detalle técnico.")


def _improvement_hint(answer: str) -> str:
    mapping = {
        "Yes": "Mantener la evidencia y verificar que las citas sean localizables en el PRODOC.",
        "Partial": "Revisar los chequeos marcados como falso en la auditoría técnica y completar esos elementos.",
        "No": "Incorporar evidencia específica para los elementos exigidos por la rúbrica.",
        "Not Found": "Añadir una sección, anexo o referencia explícita que permita valorar este criterio.",
        "N/A": "Sin acción, salvo que cambien las condiciones de aplicabilidad del proyecto.",
        "Error": "Reintentar la evaluación o revisar si el documento/rúbrica contiene texto problemático.",
    }
    return mapping.get(str(answer), "Revisar la auditoría técnica.")


def _review_flag(row: pd.Series, lang: str = DEFAULT_LANG) -> str:
    L = LABELS[_lang(lang)]
    reasons = []
    if str(row.get("Subjetividad", "")).strip() == "Alta":
        reasons.append(L["review_subj"])
    stability = row.get("Estabilidad (%)")
    if stability is not None and not pd.isna(stability):
        try:
            if float(stability) < STABILITY_THRESHOLD_PCT:
                reasons.append(L["review_stab"].format(
                    thr=_format_pct(STABILITY_THRESHOLD_PCT),
                    pct=_format_pct(float(stability)),
                ))
        except (TypeError, ValueError):
            pass
    if str(row.get("Respuesta", "")).strip() in {"Partial", "No", "Not Found", "Error"}:
        reasons.append(L["review_verdict"])
    if not reasons:
        return L["review_no"]
    return f"{L['review_yes']} - " + "; ".join(reasons)


def results_to_public_dataframe(results: list[dict[str, Any]]) -> pd.DataFrame:
    """Return a user-facing results table without technical test scaffolding."""
    technical = results_to_dataframe(results)
    if technical.empty:
        return pd.DataFrame(columns=PUBLIC_RESULT_COLUMNS)

    public = technical.copy()
    public["Resultado de valoración"] = public["Respuesta"]
    public["Lectura rápida"] = public["Respuesta"].map(_quick_reading)
    public["Principal oportunidad de mejora"] = public["Respuesta"].map(_improvement_hint)
    public["Evidencia clave"] = public["Evidencia"].map(_compact_text)
    available = [c for c in PUBLIC_RESULT_COLUMNS if c in public.columns]
    return public[available]


def _is_unstable(row: pd.Series) -> bool:
    stability = row.get("Estabilidad (%)")
    if stability is None or pd.isna(stability):
        return False
    try:
        return float(stability) < STABILITY_THRESHOLD_PCT
    except (TypeError, ValueError):
        return False


def _priority_reason(row: pd.Series, lang: str = DEFAULT_LANG) -> str:
    """Why this criterion must be reviewed. Empty string = not priority."""
    L = LABELS[_lang(lang)]
    motivos = []
    if _is_unstable(row):
        motivos.append(L["prior_unstable"].format(
            pct=_format_pct(float(row["Estabilidad (%)"])),
            thr=_format_pct(STABILITY_THRESHOLD_PCT),
        ))
    if str(row.get("Respuesta", "")).strip() == "No":
        motivos.append(L["prior_no"])
    return " + ".join(motivos)


def priority_to_dataframe(
    results: list[dict[str, Any]], lang: str = DEFAULT_LANG
) -> pd.DataFrame:
    """Criteria requiring mandatory human review: unstable and/or answered No."""
    df = results_to_dataframe(results, lang)
    if df.empty:
        return pd.DataFrame(columns=PRIORITY_COLUMNS)
    df = df.copy()
    df["Motivo de prioridad"] = df.apply(_priority_reason, axis=1, lang=lang)
    df = df[df["Motivo de prioridad"] != ""]
    if df.empty:
        return pd.DataFrame(columns=PRIORITY_COLUMNS)
    # Los dos motivos juntos primero, luego inestables, luego "No".
    orden = {2: 0, 1: 1}
    df = df.assign(
        _rank=df["Motivo de prioridad"].map(lambda m: orden[m.count("+") + 1]),
        _sk=df["ID"].map(_id_sort_key),
    ).sort_values(["_rank", "_sk"]).drop(columns=["_rank", "_sk"]).reset_index(drop=True)
    available = [c for c in PRIORITY_COLUMNS if c in df.columns]
    return df[available]


def rubric_to_dataframe(
    criteria: pd.DataFrame, results: list[dict[str, Any]]
) -> pd.DataFrame:
    """Rubric rows for the criteria actually evaluated, in result order."""
    if criteria is None or criteria.empty:
        return pd.DataFrame(columns=RUBRIC_SHEET_COLUMNS)
    evaluated = [str(r.get("ID", "")) for r in results]
    df = criteria.copy()
    df["ID"] = df["ID"].astype(str)
    df = df[df["ID"].isin(evaluated)]
    available = [c for c in RUBRIC_SHEET_COLUMNS if c in df.columns]
    return (
        df.assign(_sk=df["ID"].map(_id_sort_key))
        .sort_values("_sk")
        .drop(columns="_sk")
        .reset_index(drop=True)[available]
    )


def results_to_xlsx_bytes(
    results: list[dict[str, Any]],
    criteria: pd.DataFrame | None = None,
    lang: str = DEFAULT_LANG,
) -> bytes:
    """Serialize v3 results to an XLSX workbook.

    Sheet 1, "Resultado Diagnostico": one row per criterion with the verdict,
    the stability figures, the check-by-check audit trail and the evidence.
    Sheet 2, "Revisión prioritaria": the subset that must be reviewed by a
    person — unstable results and every "No".
    Sheet 3, "Rubrica aplicada" (when `criteria` is given): the definition of
    each criterion evaluated, so the file can be audited on its own.
    """
    lang = _lang(lang)
    L = LABELS[lang]
    df = results_to_dataframe(results, lang)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
        workbook = writer.book
        header_fmt = workbook.add_format(
            {"bold": True, "text_wrap": True, "valign": "top", "fg_color": "#D9EAF7", "border": 1}
        )
        wrap_fmt = workbook.add_format({"text_wrap": True, "valign": "top"})

        def write_sheet(frame: pd.DataFrame, name: str, widths: dict[str, int]) -> None:
            # Las claves internas siguen en español; sólo se traduce al mostrar.
            if lang != DEFAULT_LANG:
                frame = frame.rename(columns=COLUMN_LABELS_EN)
            frame.to_excel(writer, index=False, sheet_name=name)
            ws = writer.sheets[name]
            ws.freeze_panes(1, 0)
            ws.autofilter(0, 0, max(len(frame), 1), max(len(frame.columns) - 1, 0))
            for col_num, value in enumerate(frame.columns):
                ws.write(0, col_num, value, header_fmt)
            for col_range, width in widths.items():
                ws.set_column(col_range, width, wrap_fmt)

        write_sheet(df, L["sheet_result"], {
            "A:B": 10,    # ID, Subseccion
            "C:D": 52,    # Pregunta orientadora, Criterio
            "E:F": 18,    # Subjetividad, Transversales
            "G:G": 16,    # Respuesta
            "H:I": 14,    # N corridas, Conteo modal
            "J:K": 16,    # Estabilidad (%), Estable (>=80%)
            "L:L": 20,    # Resultado Alternativo
            "M:N": 26,    # Distribucion de respuestas, Corridas con error
            "O:P": 60,    # Razonamiento, Evidencia
            "Q:Q": 14,    # Status
            "R:R": 34,    # Revision humana recomendada
        })

        # La hoja se escribe SIEMPRE: su ausencia era indistinguible de un
        # despliegue viejo. Vacía, dice explícitamente que no hay pendientes.
        priority_df = priority_to_dataframe(results, lang)
        if priority_df.empty:
            write_sheet(
                pd.DataFrame(
                    [{PRIORITY_COLUMNS[0]: L["prior_placeholder"],
                      "Motivo de prioridad": L["prior_none"]}],
                    columns=PRIORITY_COLUMNS,
                ).fillna(""),
                L["sheet_priority"],
                {"A:B": 10, "C:C": 52, "D:D": 76},
            )
        else:
            write_sheet(priority_df, L["sheet_priority"], {
                "A:B": 10,    # ID, Subseccion
                "C:C": 52,    # Criterio
                "D:D": 30,    # Motivo de prioridad
                "E:E": 16,    # Respuesta
                "F:G": 18,    # Estabilidad (%), Resultado Alternativo
                "H:H": 26,    # Distribucion de respuestas
                "I:I": 16,    # Subjetividad
                "J:K": 60,    # Razonamiento, Evidencia
            })


        rubric_df = rubric_to_dataframe(criteria, results)
        if not rubric_df.empty:
            write_sheet(rubric_df, L["sheet_rubric"], {
                "A:A": 10, "B:C": 16, "D:D": 54, "E:G": 18,
                "H:H": 46, "I:L": 52, "M:M": 34, "N:N": 16,
            })
    return buf.getvalue()


def summarize_results(results: list[dict[str, Any]]) -> dict[str, Any]:
    """Create a compact deterministic summary for GPT responses."""
    answer_counts = Counter(r.get("Respuesta", "Unknown") for r in results)
    status_counts = Counter(r.get("Status", "Unknown") for r in results)
    high_subjectivity = [
        r for r in results if str(r.get("Subjetividad", "")).strip() == "Alta"
    ]
    by_subjectivity: dict[str, dict[str, int]] = {}
    for result in results:
        subj = str(result.get("Subjetividad", "Unknown"))
        ans = str(result.get("Respuesta", "Unknown"))
        by_subjectivity.setdefault(subj, {})
        by_subjectivity[subj][ans] = by_subjectivity[subj].get(ans, 0) + 1

    stability_values = []
    unstable_results = []
    repeat_error_count = 0
    criteria_with_repeat_errors = []
    for result in results:
        try:
            stability = float(result.get("Estabilidad (%)"))
        except (TypeError, ValueError):
            continue
        stability_values.append(stability)
        if stability < STABILITY_THRESHOLD_PCT:
            unstable_results.append(result)
        try:
            repeat_errors = int(result.get("Corridas con error", 0) or 0)
        except (TypeError, ValueError):
            repeat_errors = 0
        repeat_error_count += repeat_errors
        if repeat_errors:
            criteria_with_repeat_errors.append(result.get("ID"))

    priority_ids = sorted(
        {str(r.get("ID")) for r in unstable_results}
        | {str(r.get("ID")) for r in results if str(r.get("Respuesta", "")).strip() == "No"},
        key=_id_sort_key,
    )

    return {
        "total_criteria": len(results),
        "answers": dict(answer_counts),
        "statuses": dict(status_counts),
        "by_subjectivity": by_subjectivity,
        "high_subjectivity_count": len(high_subjectivity),
        "high_subjectivity_ids": [r.get("ID") for r in high_subjectivity],
        "stability_repeats": STABILITY_REPEATS,
        "stable_threshold_pct": STABILITY_THRESHOLD_PCT,
        "average_stability_pct": (
            round(sum(stability_values) / len(stability_values), 1)
            if stability_values
            else None
        ),
        "stable_count": len(results) - len(unstable_results),
        "unstable_count": len(unstable_results),
        "unstable_ids": [r.get("ID") for r in unstable_results],
        "unstable_drift": {
            str(r.get("ID")): r.get("Deriva principal (si inestable)", "")
            for r in unstable_results
        },
        "repeat_error_count": repeat_error_count,
        "criteria_with_repeat_errors": criteria_with_repeat_errors,
        # Hoja "Revisión prioritaria": inestables ∪ "No". Revisión obligatoria.
        "priority_review_count": len(priority_ids),
        "priority_review_ids": priority_ids,
    }


def _demo() -> None:
    """Self-check for the reasoning renderer. Run: python tab1_v3_core.py"""
    rubrica = (
        "TESTS:\n"
        "T1: ¿Distingue tipo de enfoque? (sí/no)\n"
        "T2: ¿Articula cómo cuestiona normas de poder? (sí/no)\n"
        "T3: ¿Acciones dedicadas a transformar relaciones? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3"
    )
    result = {
        "Respuesta": "Partial",
        "N corridas": 10,
        "Conteo modal": 6,
        DRIFT_SOURCE_COLUMN: "No",
        "Razonamiento": (
            "T1: verdadero — la sección 3.2 los distingue.\n"
            "T2: falso — no se explica el mecanismo.\n"
            "T3: verdadero — actividad 2.4 con presupuesto.\n"
            "DECISIÓN: T1 ∧ T2 ∧ T3 = falso → Parcial"
        ),
    }
    out = render_reasoning(result, rubrica)
    # el lector nunca ve una etiqueta suelta...
    assert "T1: verdadero" not in out
    # ...pero cada chequeo aparece con su enunciado y su marca
    assert "✓ ¿Distingue tipo de enfoque?" in out
    assert "✗ ¿Articula cómo cuestiona normas de poder?" in out
    # abre por el motivo, y nombra lo que falta
    assert out.startswith("POR QUÉ PARTIAL")
    assert "Se cumplen 2 de 3" in out
    assert "Falta: Articula cómo cuestiona normas de poder." in out
    # la regla formal sigue disponible para auditoría
    assert "REGLA · se requieren los 3 chequeos  (T1 ∧ T2 ∧ T3)" in out
    # estabilidad renderizada desde los campos del resultado
    assert "ESTABILIDAD · 6 de 10 corridas coincidieron. Resultado alternativo: No." in out

    # formato inesperado -> se devuelve intacto, nunca se mutila
    suelto = "El documento no articula un enfoque transformador."
    assert render_reasoning({"Respuesta": "No", "Razonamiento": suelto}, rubrica) == suelto
    # rúbrica ilegible -> también intacto
    assert render_reasoning(result, "sin estructura") == result["Razonamiento"]

    _demo_priority()
    _demo_language()
    print("tab1_v3_core._demo OK")


def _demo_language() -> None:
    """Self-check for the output-language switch."""
    rubrica = (
        "TESTS:\n"
        "T1: ¿Distingue tipo de enfoque? (sí/no)\n"
        "T2: ¿Articula el mecanismo? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2"
    )
    corridas = [
        {"repeat": i, "Respuesta": "No", "Status": "OK", "Evidencia": "q",
         "Razonamiento": "T1: verdadero — ok.\nT2: falso — no.\n"
                         "DECISIÓN: T1 ∧ T2 = falso → No"}
        for i in range(6)
    ] + [
        {"repeat": i, "Respuesta": "Partial", "Status": "OK", "Evidencia": "q",
         "Razonamiento": "T1: falso — x"} for i in range(6, 10)
    ]
    row = pd.Series({"Rúbrica — Sí": rubrica})

    def build(lang):
        agg = aggregate_repeated_criterion_results(
            corridas, STABILITY_THRESHOLD_PCT, lang
        )
        agg.update({"ID": "1.5.6", "Subsección": "1.5", "Criterio": "C",
                    "Subjetividad": "Alta"})
        return [_with_readable_reasoning(agg, row, lang)]

    es, en = build("es"), build("en")

    # el veredicto y la estabilidad NO dependen del idioma: sólo la presentación
    for key in ("Respuesta", "Estabilidad (%)", "Conteo modal", "N corridas"):
        assert es[0][key] == en[0][key], key
    # el resumen sólo puede diferir en texto de presentación (unstable_drift)
    s_es, s_en = summarize_results(es), summarize_results(en)
    assert {k: v for k, v in s_es.items() if k != "unstable_drift"} == \
           {k: v for k, v in s_en.items() if k != "unstable_drift"}

    r_es = results_to_dataframe(es, "es").iloc[0]
    r_en = results_to_dataframe(en, "en").iloc[0]
    assert r_es["Razonamiento"].startswith("POR QUÉ NO")
    assert r_en["Razonamiento"].startswith("WHY NO")
    assert "VERIFICACIÓN" in r_es["Razonamiento"]
    assert "VERIFICATION" in r_en["Razonamiento"]
    assert "corridas coincidieron" in r_es["Razonamiento"]
    assert "runs agreed" in r_en["Razonamiento"]
    assert r_es["Revisión humana recomendada"].startswith("Sí - ")
    assert r_en["Revisión humana recomendada"].startswith("Yes - ")
    assert r_es["Estable (>=80%)"] == "No" and r_en["Estable (>=80%)"] == "No"

    assert priority_to_dataframe(es, "es").iloc[0]["Motivo de prioridad"] == (
        "Inestable (60% < 80%) + Resultado No")
    assert priority_to_dataframe(en, "en").iloc[0]["Motivo de prioridad"] == (
        "Unstable (60% < 80%) + Verdict No")

    # el andamiaje que parsea el renderizador se mantiene en español
    p_en, p_es = system_prompt("en"), system_prompt("es")
    assert "verdadero" in p_en and "DECISIÓN" in p_en
    assert p_es == V3_SYSTEM_PROMPT
    # ...pero la orden de idioma NO puede quedar duplicada ni contradecirse:
    # dejar "Idioma: español." y añadir "escribe en inglés" hacía que el
    # modelo devolviera español con cabeceras inglesas.
    assert _ES_LANG_DIRECTIVE not in p_en, "el prompt en inglés se contradice"
    assert "IDIOMA DE SALIDA: INGLÉS" in p_en
    assert _ES_LANG_DIRECTIVE in p_es
    # el recordatorio final sólo aparece en inglés
    fila = pd.Series({"ID": "1.1.1", "Criterio a evaluar": "C", "Rúbrica — Sí": "T1: x"})
    u_en = build_user_prompt(fila, "texto", "en")
    u_es = build_user_prompt(fila, "texto", "es")
    assert "RECORDATORIO DE IDIOMA" in u_en
    assert "RECORDATORIO DE IDIOMA" not in u_es
    assert u_es == build_user_prompt(fila, "texto")  # es por defecto, sin cambios
    # idioma desconocido -> español, nunca un fallo
    assert _lang("fr") == "es" and _lang(None) == "es" and _lang("EN") == "en"

    import openpyxl
    wb_en = openpyxl.load_workbook(io.BytesIO(results_to_xlsx_bytes(en, None, "en")))
    assert wb_en.sheetnames[0] == "Appraisal Result"
    assert "Priority Review" in wb_en.sheetnames
    assert [c.value for c in wb_en["Appraisal Result"][1]][:3] == [
        "ID", "Subsection", "Criterion"]
    wb_es = openpyxl.load_workbook(io.BytesIO(results_to_xlsx_bytes(es, None, "es")))
    assert wb_es.sheetnames[0] == SHEET_RESULT
    assert [c.value for c in wb_es[SHEET_RESULT][1]][:3] == [
        "ID", "Subsección", "Criterio"]


def _demo_priority() -> None:
    """Self-check for the mandatory-review sheet."""
    rows = [
        # estable y conforme -> fuera
        {"ID": "1.1.1", "Respuesta": "Yes", "Estabilidad (%)": 100.0, "Subjetividad": "Baja"},
        # inestable aunque el veredicto sea favorable -> entra
        {"ID": "1.2.1", "Respuesta": "Yes", "Estabilidad (%)": 60.0, "Subjetividad": "Baja"},
        # incumplimiento estable -> entra
        {"ID": "1.3.1", "Respuesta": "No", "Estabilidad (%)": 100.0, "Subjetividad": "Baja"},
        # los dos motivos -> entra y encabeza
        {"ID": "1.4.1", "Respuesta": "No", "Estabilidad (%)": 50.0, "Subjetividad": "Alta"},
        # subjetividad alta sola no basta: queda a criterio del usuario
        {"ID": "1.5.1", "Respuesta": "Partial", "Estabilidad (%)": 90.0, "Subjetividad": "Alta"},
    ]
    df = priority_to_dataframe(rows)
    assert list(df["ID"]) == ["1.4.1", "1.2.1", "1.3.1"], list(df["ID"])
    assert df.iloc[0]["Motivo de prioridad"] == "Inestable (50% < 80%) + Resultado No"
    assert df.iloc[1]["Motivo de prioridad"] == "Inestable (60% < 80%)"
    assert df.iloc[2]["Motivo de prioridad"] == "Resultado No"
    # sin resultados prioritarios la hoja no se escribe
    assert priority_to_dataframe([rows[0]]).empty
    assert priority_to_dataframe([]).empty
    # el resumen y la hoja no pueden divergir
    resumen = summarize_results(rows)
    assert resumen["priority_review_count"] == 3
    assert sorted(resumen["priority_review_ids"]) == sorted(df["ID"])
    # el libro gana la hoja, y la omite cuando no hace falta
    import openpyxl
    book = openpyxl.load_workbook(io.BytesIO(results_to_xlsx_bytes(rows)))
    assert SHEET_PRIORITY in book.sheetnames, book.sheetnames
    assert book[SHEET_PRIORITY].max_row == 4  # cabecera + 3
    # sin pendientes la hoja SIGUE estando, y lo dice: su ausencia no puede
    # confundirse con una versión vieja del servidor
    limpio = openpyxl.load_workbook(io.BytesIO(results_to_xlsx_bytes([rows[0]])))
    assert SHEET_PRIORITY in limpio.sheetnames
    hoja = limpio[SHEET_PRIORITY]
    assert hoja.max_row == 2
    assert "Ningún criterio exige revisión obligatoria" in hoja.cell(2, 4).value


if __name__ == "__main__":
    _demo()
