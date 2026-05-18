"""Build a detailed per-criterion rubric for Tab 1 in xlsx form.

Output columns:
- ID, Sección, Subsección, Pregunta orientadora (CONTEXTO — no evaluar),
  Criterio a evaluar, Tipo, Aplicabilidad, Aspectos transversales,
  Elementos a verificar, Rúbrica Sí, Rúbrica Parcial, Rúbrica No,
  Rúbrica No aplica, Ejemplo evidencia Sí [LLENAR],
  Ejemplo evidencia Parcial [LLENAR], Ejemplo evidencia No [LLENAR], Notas

Logic:
- The head question (1.1, 1.2, ...) is *contexto*. Evaluation always targets
  the criterion (1.1.1, 1.1.2, ...).
- Transversal-tagged criteria require DEDICATED elements (not FRAMING) to
  score Sí or Parcial — same logic as the Tab 1 prompt today.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

PLACEHOLDER_SI = "[LLENAR — cita textual de propuesta real que SÍ cumple este criterio]"
PLACEHOLDER_PA = "[LLENAR — cita textual de propuesta que cumple PARCIALMENTE este criterio]"
PLACEHOLDER_NO = "[LLENAR — cita textual de propuesta que NO cumple, o nota de ausencia]"

# Transversal A/B gate language reused for transversal-tagged rows
TRANSV_GATE = (
    "Aplicar el filtro DEDICADO vs MARCO. Solo cuentan citas DEDICADAS al sujeto "
    "(sub-objetivo, indicador desagregado, actividad cuyo propósito principal es el "
    "sujeto, partida presupuestaria, o meta cuantificable). NO cuentan menciones de "
    "MARCO (objetivo general que enumera grupos, listas de partes interesadas con "
    "≥3 grupos, lenguaje de inclusión genérico, «entre otros», «incluyendo X, Y, Z»)."
)

# Each entry mirrors the source row order (0..75)
RUBRICA = [
    # ---------------- SECCIÓN 1 — PERTINENCIA ----------------
    {
        "id": "1.1.1",
        "seccion": "1. Pertinencia",
        "subseccion": "1.1",
        "head": "¿Está claramente definido el objetivo general del programa/proyecto?",
        "criterio": "La intervención especifica su contribución a las prioridades a largo plazo de la OIT, los resultados y los resultados del DWCP, incluidos los CPO pertinentes.",
        "tipo": "Calibración",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Cita explícita de prioridades de P&B / resultados OIT.\n"
            "B. Cita explícita de resultados DWCP del país.\n"
            "C. Identifica al menos un CPO pertinente por código o nombre.\n"
            "D. Explica CÓMO la intervención contribuye a A/B/C (no solo lista)."
        ),
        "si": "La propuesta identifica explícitamente al menos un resultado de P&B, al menos un resultado de DWCP y al menos un CPO pertinente, Y articula la cadena causal por la cual la intervención contribuye a esos elementos. La contribución debe estar redactada en términos sustantivos, no decorativos.",
        "parcial": "Identifica 1–2 de los elementos A/B/C pero omite el resto, O menciona los tres pero sin articular cómo la intervención contribuye (lista, no contribución). También parcial si la contribución está implícita pero no explicitada.",
        "no": "No menciona P&B, DWCP ni CPO; o los menciona solo como antecedente (p. ej. «alineado con los marcos de la OIT») sin especificidad. También No si los menciona pero la cadena causal es evidentemente incoherente con la intervención propuesta.",
        "na": "",
        "notas": "Verificar también el anexo de marco lógico — los CPO suelen aparecer ahí.",
    },
    {
        "id": "1.1.2",
        "seccion": "1. Pertinencia",
        "subseccion": "1.1",
        "head": "¿Está claramente definido el objetivo general del programa/proyecto?",
        "criterio": "El nivel de ambición y visibilidad respecto a la inclusión de la discapacidad está a la par con la etiqueta de inclusión de discapacidad de los CPO correspondientes.",
        "tipo": "Calibración transversal",
        "aplicabilidad": "Siempre",
        "transv": "Discapacidad",
        "elementos": (
            "A. Identifica la etiqueta/marcador de discapacidad de los CPO referenciados.\n"
            "B. La propuesta presenta acciones DEDICADAS a discapacidad (no solo mención).\n"
            "C. El nivel (visibilidad y ambición) coincide con la etiqueta del CPO."
        ),
        "si": f"La propuesta nombra la etiqueta/marcador de discapacidad del CPO Y demuestra acciones DEDICADAS a discapacidad cuyo nivel es coherente con esa etiqueta (mayor ambición ↔ etiqueta principal; menor ambición ↔ etiqueta significativa). {TRANSV_GATE}",
        "parcial": "Reconoce el CPO/etiqueta pero la propuesta solo incluye menciones de MARCO sobre discapacidad (lista de grupos vulnerables, «entre otros»), sin elementos DEDICADOS, O incluye 1–2 elementos DEDICADOS pero por debajo de lo que sugiere la etiqueta del CPO.",
        "no": "No identifica la etiqueta de discapacidad del CPO, o la propuesta no incluye ningún elemento DEDICADO a discapacidad pese a que la etiqueta del CPO lo exigiría.",
        "na": "Solo si los CPO referenciados explícitamente no llevan etiqueta de discapacidad — debe documentarse.",
        "notas": "Cruzar con la columna del CPO. Una propuesta «inclusiva en lenguaje» pero sin elementos DEDICADOS NO satisface este criterio.",
    },
    {
        "id": "1.1.3",
        "seccion": "1. Pertinencia",
        "subseccion": "1.1",
        "head": "¿Está claramente definido el objetivo general del programa/proyecto?",
        "criterio": "El nivel de ambición y visibilidad respecto a la igualdad de género está a la par con el marcador de igualdad de género de los CPO correspondientes.",
        "tipo": "Calibración transversal",
        "aplicabilidad": "Siempre",
        "transv": "Género",
        "elementos": (
            "A. Identifica el marcador de género del CPO (0/1/2/3 o equivalente).\n"
            "B. La propuesta presenta acciones DEDICADAS a género.\n"
            "C. El nivel coincide con el marcador del CPO."
        ),
        "si": f"Identifica el marcador de género del CPO Y la propuesta tiene elementos DEDICADOS a género (sub-objetivo, indicador desagregado por sexo, actividad cuyo propósito principal es género, partida presupuestaria de género, o meta cuantificable de género) cuyo nivel coincide con el marcador. {TRANSV_GATE}",
        "parcial": "Reconoce el marcador pero solo incluye menciones de MARCO («mujeres y hombres», «sensible al género»), O incluye 1–2 elementos DEDICADOS pero por debajo del nivel del marcador.",
        "no": "No identifica el marcador, o el marcador es ≥2 (transformador) y la propuesta no incluye ningún elemento DEDICADO a género más allá de lenguaje inclusivo.",
        "na": "",
        "notas": "OIT distingue: marcador 0 = no relevante; 1 = limitado; 2 = significativo; 3 = principal/transformador. La rúbrica más estricta aplica a 2 y 3.",
    },
    {
        "id": "1.2.1",
        "seccion": "1. Pertinencia",
        "subseccion": "1.2",
        "head": "¿Identifica el proyecto los ODS, marcos de cooperación de la ONU y marcos nacionales pertinentes?",
        "criterio": "La propuesta explica cómo el proyecto encaja en el marco más amplio de la ayuda al desarrollo del país, incluidos los planes nacionales de desarrollo.",
        "tipo": "Lista de verificación",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Identifica el plan nacional de desarrollo vigente (por nombre).\n"
            "B. Identifica el marco de cooperación de la ONU (UNSDCF/equivalente).\n"
            "C. Articula el encaje del proyecto en A y B."
        ),
        "si": "Identifica explícitamente el plan nacional de desarrollo y el UNSDCF (u homólogo) vigentes Y articula cómo el proyecto encaja en ambos.",
        "parcial": "Identifica solo uno de los dos marcos, O identifica ambos pero el «encaje» queda en mención decorativa sin articulación causal.",
        "no": "No menciona el plan nacional de desarrollo ni el UNSDCF; o menciona la categoría genérica («marcos nacionales») sin nombrarlos.",
        "na": "",
        "notas": "Validar nombre exacto del plan nacional contra fuente oficial del país.",
    },
    {
        "id": "1.2.2",
        "seccion": "1. Pertinencia",
        "subseccion": "1.2",
        "head": "¿Identifica el proyecto los ODS, marcos de cooperación de la ONU y marcos nacionales pertinentes?",
        "criterio": "La propuesta es coherente con el marco de cooperación de las Naciones Unidas y contribuye a las áreas específicas de las que es responsable la OIT.",
        "tipo": "Lista de verificación",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Identifica los outcomes/áreas del UNSDCF a los que contribuye.\n"
            "B. Identifica las áreas en las que la OIT es agencia responsable o convocante.\n"
            "C. Articula la contribución del proyecto a A y B."
        ),
        "si": "La propuesta nombra los outcomes específicos del UNSDCF, identifica las áreas donde la OIT es responsable, Y muestra la contribución del proyecto a esos outcomes/áreas con suficiente detalle.",
        "parcial": "Solo coherencia general («alineado con el UNSDCF») sin nombrar outcomes; o nombra outcomes pero no demuestra el vínculo OIT-responsabilidad-proyecto.",
        "no": "No referencia el UNSDCF de forma específica.",
        "na": "",
        "notas": "",
    },
    {
        "id": "1.2.3",
        "seccion": "1. Pertinencia",
        "subseccion": "1.2",
        "head": "¿Identifica el proyecto los ODS, marcos de cooperación de la ONU y marcos nacionales pertinentes?",
        "criterio": "La propuesta identifica los indicadores pertinentes de los ODS coherentes con los indicadores de los ODS vinculados a los CPO pertinentes.",
        "tipo": "Lista de verificación",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Identifica indicadores de ODS por código (p. ej. 8.5.2, 1.3.1).\n"
            "B. Esos indicadores corresponden a los ODS que los CPO ya vinculan.\n"
            "C. Los indicadores ODS aparecen en el marco lógico o en el plan de M&E."
        ),
        "si": "Identifica indicadores ODS por código (no solo nombres de ODS), coherentes con los ODS de los CPO, Y los integra en el marco lógico o plan de M&E.",
        "parcial": "Menciona ODS pero sin códigos de indicador, o menciona indicadores pero no están integrados en el marco lógico.",
        "no": "No referencia indicadores de ODS, o solo cita ODS de forma decorativa («contribuye a los ODS 5 y 8»).",
        "na": "",
        "notas": "Verificar contra https://unstats.un.org/sdgs/indicators/indicators-list/",
    },
    {
        "id": "1.3.1",
        "seccion": "1. Pertinencia",
        "subseccion": "1.3",
        "head": "¿Es la intervención propuesta pertinente para abordar el problema?",
        "criterio": "Se ha formulado una descripción del problema mediante un análisis de la situación (estudio de referencia u otras pruebas).",
        "tipo": "Calidad narrativa",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Existe una sección o anexo dedicado al análisis de situación.\n"
            "B. El análisis cita fuentes (estudios, datos oficiales, evaluaciones, consultas).\n"
            "C. El problema está delimitado en alcance, magnitud y población afectada."
        ),
        "si": "El análisis del problema es sustantivo: cita fuentes verificables, cuantifica magnitud cuando aplica, y delimita alcance temporal/geográfico/poblacional.",
        "parcial": "Hay descripción del problema pero faltan fuentes o cuantificación; o las fuentes son genéricas («informes recientes muestran...») sin referencias específicas.",
        "no": "El problema solo se enuncia en la introducción sin análisis ni evidencia.",
        "na": "",
        "notas": "",
    },
    {
        "id": "1.3.2",
        "seccion": "1. Pertinencia",
        "subseccion": "1.3",
        "head": "¿Es la intervención propuesta pertinente para abordar el problema?",
        "criterio": "Se explican las causas del problema.",
        "tipo": "Calidad narrativa",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Se distingue entre causas inmediatas, subyacentes y estructurales (o esquema equivalente).\n"
            "B. Las causas están vinculadas a evidencia o consulta.\n"
            "C. La estrategia del proyecto se relaciona con causas, no solo con síntomas."
        ),
        "si": "El análisis identifica causas en al menos dos niveles (p. ej. árbol de problemas) Y la estrategia del proyecto ataca causas, no solo síntomas.",
        "parcial": "Lista causas sin estructura causal clara, o se queda en causas inmediatas sin tocar lo estructural.",
        "no": "No hay análisis causal; el problema y la solución se presentan sin explicar por qué la solución resuelve el problema.",
        "na": "",
        "notas": "",
    },
    {
        "id": "1.3.3",
        "seccion": "1. Pertinencia",
        "subseccion": "1.3",
        "head": "¿Es la intervención propuesta pertinente para abordar el problema?",
        "criterio": "Se ha identificado el grupo de población/los beneficiarios finales afectados, con análisis de roles de género, división del trabajo, oportunidades y limitaciones de mujeres y hombres.",
        "tipo": "Lista transversal",
        "aplicabilidad": "Siempre",
        "transv": "Género",
        "elementos": (
            "A. Identifica al grupo de población afectado por nombre y magnitud.\n"
            "B. Análisis explícito de roles de género en ese contexto.\n"
            "C. Análisis de división sexual del trabajo.\n"
            "D. Análisis de oportunidades y limitaciones diferenciadas por sexo.\n"
            "E. Datos desagregados por sexo cuando estén disponibles."
        ),
        "si": f"Cumple A y al menos 3 de B/C/D/E con sustento en datos o consulta. {TRANSV_GATE}",
        "parcial": "Cumple A y 1–2 de B/C/D/E, o el análisis de género existe pero es genérico (no específico del contexto del proyecto).",
        "no": "Identifica beneficiarios sin análisis de género, o solo usa lenguaje inclusivo («mujeres y hombres») sin análisis sustantivo.",
        "na": "",
        "notas": "Análisis de género no es lo mismo que mencionar mujeres. Es análisis de relaciones de poder y división del trabajo en el contexto específico.",
    },
    {
        "id": "1.3.4",
        "seccion": "1. Pertinencia",
        "subseccion": "1.3",
        "head": "¿Es la intervención propuesta pertinente para abordar el problema?",
        "criterio": "Se han identificado las partes interesadas relacionadas con el problema y la solución, junto con limitaciones e intereses. El análisis es sensible al género e incluye, cuando aplica, organizaciones de mujeres y de personas con discapacidad.",
        "tipo": "Lista transversal",
        "aplicabilidad": "Siempre",
        "transv": "Género; Discapacidad",
        "elementos": (
            "A. Mapeo de partes interesadas (institucionales, sociales, beneficiarios).\n"
            "B. Limitaciones e intereses de cada actor relevante.\n"
            "C. Inclusión explícita de organizaciones de mujeres.\n"
            "D. Inclusión explícita de organizaciones de personas con discapacidad cuando el proyecto las afecta.\n"
            "E. Análisis de poder/influencia entre actores."
        ),
        "si": f"Mapeo claro de A+B+E con C y, cuando aplica, D. {TRANSV_GATE}",
        "parcial": "Mapeo presente pero superficial en B o E, O ausente C/D pese a relevancia.",
        "no": "Solo lista de actores sin análisis de intereses/limitaciones.",
        "na": "",
        "notas": "",
    },
    {
        "id": "1.3.5",
        "seccion": "1. Pertinencia",
        "subseccion": "1.3",
        "head": "¿Es la intervención propuesta pertinente para abordar el problema?",
        "criterio": "El proyecto incluye una consulta significativa con todas las partes interesadas afectadas, especialmente grupos marginados y excluidos. Garantiza participación equitativa y evita efectos discriminatorios sobre personas con discapacidad, comunidades indígenas, mujeres y otros grupos.",
        "tipo": "Lista transversal",
        "aplicabilidad": "Siempre",
        "transv": "Género; Discapacidad; Pueblos indígenas",
        "elementos": (
            "A. Plan de consulta documentado con metodología.\n"
            "B. Inclusión específica de grupos marginados (no solo lista).\n"
            "C. Medidas para participación equitativa (accesibilidad, traducción, horarios).\n"
            "D. Análisis de potenciales efectos discriminatorios y mitigación.\n"
            "E. Consulta libre, previa e informada para pueblos indígenas cuando aplica (CLPI)."
        ),
        "si": f"Cumple A, B y C; cumple D y E cuando aplican. {TRANSV_GATE}",
        "parcial": "Plan de consulta existe pero los grupos marginados aparecen solo en lista, sin medidas operativas (C) ni análisis de discriminación (D).",
        "no": "Consulta solo con autoridades o socios principales, sin mecanismos para incluir a grupos marginados.",
        "na": "E aplica solo si el proyecto afecta a pueblos indígenas — debe documentarse.",
        "notas": "CLPI = Consentimiento Libre, Previo e Informado (Convenio 169 OIT).",
    },
    {
        "id": "1.4.1",
        "seccion": "1. Pertinencia",
        "subseccion": "1.4",
        "head": "¿Hace hincapié la propuesta en la ventaja comparativa de la OIT?",
        "criterio": "La propuesta destaca la identidad normativa y tripartita de la OIT, su capacidad y su valor añadido para llevar a cabo el proyecto.",
        "tipo": "Calidad narrativa",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Referencia explícita a la identidad normativa OIT (convenios, recomendaciones).\n"
            "B. Referencia explícita al tripartismo (constituyentes: gobierno, empleadores, trabajadores).\n"
            "C. Argumento de valor añadido específico al proyecto (no genérico)."
        ),
        "si": "Cumple A, B y C con argumento específico al proyecto (no plantilla institucional).",
        "parcial": "Cumple A y B pero el argumento de valor añadido es genérico.",
        "no": "Solo describe a la OIT como organización; no argumenta su ventaja comparativa para este proyecto.",
        "na": "",
        "notas": "",
    },
    {
        "id": "1.4.2",
        "seccion": "1. Pertinencia",
        "subseccion": "1.4",
        "head": "¿Hace hincapié la propuesta en la ventaja comparativa de la OIT?",
        "criterio": "Se destaca la presencia de la OIT en el país, así como la cartera de proyectos pasados y en curso.",
        "tipo": "Binario presencia",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Descripción de presencia OIT en país (oficina, equipo).\n"
            "B. Lista o resumen de proyectos pasados/curso pertinentes."
        ),
        "si": "Describe presencia y lista al menos 2 proyectos pasados/en curso pertinentes con breve resumen.",
        "parcial": "Menciona presencia OIT pero solo enumera proyectos sin pertinencia explícita; o enumera proyectos sin describir presencia.",
        "no": "No menciona ni presencia ni cartera.",
        "na": "",
        "notas": "",
    },
    {
        "id": "1.4.3",
        "seccion": "1. Pertinencia",
        "subseccion": "1.4",
        "head": "¿Hace hincapié la propuesta en la ventaja comparativa de la OIT?",
        "criterio": "La propuesta presenta experiencias previas y lecciones aprendidas de evaluaciones pertinentes y explica cómo han influido en la estrategia.",
        "tipo": "Lista de verificación",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Cita evaluaciones específicas (proyecto, programa-país, estrategia).\n"
            "B. Extrae lecciones aprendidas concretas.\n"
            "C. Vincula esas lecciones con decisiones de diseño del proyecto actual."
        ),
        "si": "Cumple A, B y C con vinculación visible (la lección modificó algo en el diseño).",
        "parcial": "Cita evaluaciones y lecciones pero no las vincula con el diseño, O las vincula de forma genérica («incorporamos lecciones»).",
        "no": "No cita evaluaciones, o solo dice «se han considerado lecciones aprendidas» sin especificar cuáles.",
        "na": "",
        "notas": "Fuente: www.ilo.org/ievaldiscovery",
    },
    {
        "id": "1.5.1",
        "seccion": "1. Pertinencia",
        "subseccion": "1.5",
        "head": "¿Refleja la propuesta las áreas transversales de la OIT?",
        "criterio": "El proyecto ha tenido en cuenta las necesidades y preocupaciones de las personas con discapacidad.",
        "tipo": "Transversal",
        "aplicabilidad": "Siempre",
        "transv": "Discapacidad",
        "elementos": (
            "A. Sub-objetivo, resultado o producto cuyo título nombra discapacidad.\n"
            "B. Indicador desagregado por discapacidad o que la mide específicamente.\n"
            "C. Actividad cuyo propósito principal es discapacidad.\n"
            "D. Partida presupuestaria para discapacidad (accesibilidad, intérprete, especialista).\n"
            "E. Meta cuantificable relativa a discapacidad."
        ),
        "si": f"Al menos 3 de A/B/C/D/E presentes con sustento. {TRANSV_GATE}",
        "parcial": f"1–2 de A/B/C/D/E presentes; o todos los elementos son menciones de MARCO. {TRANSV_GATE}",
        "no": "Solo menciones de MARCO; ningún elemento DEDICADO.",
        "na": "Documentar si el proyecto justifica explícitamente que discapacidad no aplica.",
        "notas": "Guía: «Guía práctica» n.º 18 — Inclusión de personas con discapacidad.",
    },
    {
        "id": "1.5.2",
        "seccion": "1. Pertinencia",
        "subseccion": "1.5",
        "head": "¿Refleja la propuesta las áreas transversales de la OIT?",
        "criterio": "La propuesta integra los convenios y recomendaciones pertinentes de la OIT en la estrategia, indicadores y objetivos del proyecto, promoviendo ratificación, aplicación y conocimiento. Se incluyen referencias a NORMLEX u otras fuentes jurídicas.",
        "tipo": "Lista transversal",
        "aplicabilidad": "Siempre",
        "transv": "Normas Internacionales del Trabajo (NIT)",
        "elementos": (
            "A. Cita convenios/recomendaciones por número (p. ej. C111, C190).\n"
            "B. Esos instrumentos están integrados en la estrategia (no solo listados).\n"
            "C. Indicadores u objetivos vinculan a esos instrumentos.\n"
            "D. Acciones de promoción de ratificación/aplicación/conocimiento.\n"
            "E. Referencias a NORMLEX u otras fuentes jurídicas."
        ),
        "si": "Cumple A, B y al menos 2 de C/D/E.",
        "parcial": "Cita convenios pero quedan en antecedentes (no en estrategia/indicadores), O integra estrategia pero sin referencias jurídicas verificables.",
        "no": "No cita convenios por número o solo menciona «las normas de la OIT» genéricamente.",
        "na": "",
        "notas": "NORMLEX: https://normlex.ilo.org/",
    },
    {
        "id": "1.5.3",
        "seccion": "1. Pertinencia",
        "subseccion": "1.5",
        "head": "¿Refleja la propuesta las áreas transversales de la OIT?",
        "criterio": "Se hace referencia a los informes de los organismos de supervisión específicos del país y a las normas internacionales del trabajo pertinentes en la justificación y la estrategia de ejecución.",
        "tipo": "Transversal",
        "aplicabilidad": "Siempre",
        "transv": "Normas Internacionales del Trabajo (NIT); Mecanismos de supervisión",
        "elementos": (
            "A. Cita observaciones/solicitudes directas del CEACR sobre el país.\n"
            "B. Cita conclusiones del Comité de Aplicación de Normas o del Comité de Libertad Sindical cuando aplica.\n"
            "C. Esas observaciones informan la justificación o la estrategia."
        ),
        "si": "Cita observaciones específicas y muestra cómo informan la justificación o estrategia.",
        "parcial": "Cita observaciones pero sin vínculo con la estrategia, O menciona «informes de supervisión» sin especificar cuáles.",
        "no": "No referencia observaciones de organismos de supervisión.",
        "na": "Documentar si el proyecto no se vincula a NIT.",
        "notas": "CEACR = Comisión de Expertos en Aplicación de Convenios y Recomendaciones.",
    },
    {
        "id": "1.5.4",
        "seccion": "1. Pertinencia",
        "subseccion": "1.5",
        "head": "¿Refleja la propuesta las áreas transversales de la OIT?",
        "criterio": "La propuesta garantiza el cumplimiento de las normas laborales de la OIT (salarios justos, condiciones seguras, derechos fundamentales). Si se contrata trabajadores directa o por terceros, se aplican medidas de protección y mecanismos de queja accesibles.",
        "tipo": "Lista condicional transversal",
        "aplicabilidad": "Condicional: aplica si el proyecto contrata trabajadores directa o indirectamente",
        "transv": "Normas Internacionales del Trabajo (NIT); Trabajo decente",
        "elementos": (
            "A. Compromiso explícito de cumplimiento NIT en el proyecto.\n"
            "B. Salarios justos especificados (referencia legal o sectorial).\n"
            "C. Condiciones de SST aplicables al personal del proyecto y contratistas.\n"
            "D. Mecanismo de quejas accesible para trabajadores.\n"
            "E. Cláusulas de cumplimiento en contratos con terceros."
        ),
        "si": "Cumple A, B, C, D y E cuando hay terceros.",
        "parcial": "Cumple A y 2–3 elementos; o D existe pero no es accesible (canal único, idioma único).",
        "no": "Solo afirma «se cumplirán las normas laborales» sin operacionalización.",
        "na": "Si el proyecto no contrata directa ni indirectamente, documentar por qué no aplica.",
        "notas": "",
    },
    {
        "id": "1.5.5",
        "seccion": "1. Pertinencia",
        "subseccion": "1.5",
        "head": "¿Refleja la propuesta las áreas transversales de la OIT?",
        "criterio": "La propuesta da prioridad a la sostenibilidad ambiental, integrando medidas de seguridad y protección, evaluando impactos en biodiversidad y comunidades, y promoviendo prácticas ambientalmente sostenibles.",
        "tipo": "Transversal",
        "aplicabilidad": "Siempre",
        "transv": "Medio ambiente; Sostenibilidad",
        "elementos": (
            "A. Análisis de impactos ambientales potenciales del proyecto.\n"
            "B. Medidas de mitigación de impactos identificados.\n"
            "C. Consideración de biodiversidad y comunidades afectadas.\n"
            "D. Prácticas sostenibles en ejecución (materiales, energía, residuos).\n"
            "E. Si hay infraestructura/construcción: medidas específicas de seguridad ambiental."
        ),
        "si": f"Cumple A, B y al menos 2 de C/D/E. {TRANSV_GATE}",
        "parcial": "Cumple A pero B/C/D son genéricos; o el proyecto involucra infraestructura pero E está ausente.",
        "no": "Solo menciones decorativas («el proyecto es sostenible»); ningún análisis ni medida específica.",
        "na": "Si el proyecto es 100% normativo/asesoría sin actividades físicas, documentar pero igual evaluar A.",
        "notas": "",
    },
    {
        "id": "1.5.6",
        "seccion": "1. Pertinencia",
        "subseccion": "1.5",
        "head": "¿Refleja la propuesta las áreas transversales de la OIT?",
        "criterio": "Siempre que es posible, la propuesta promueve y destaca el uso de un enfoque transformador en materia de género.",
        "tipo": "Transversal",
        "aplicabilidad": "Siempre",
        "transv": "Género",
        "elementos": (
            "A. Distingue enfoque sensible / responsivo / transformador.\n"
            "B. Articula explícitamente cómo el proyecto cuestiona normas o relaciones de poder de género.\n"
            "C. Acciones DEDICADAS a transformar relaciones (no solo a incluir mujeres).\n"
            "D. Indicadores miden cambios en relaciones, no solo participación numérica.\n"
            "E. Cita guía técnica (p. ej. «Guía práctica» n.º 15)."
        ),
        "si": f"Cumple B y C, idealmente con D. {TRANSV_GATE}",
        "parcial": "Enfoque responsivo (incluye mujeres) pero no transformador; o lo declara transformador pero las acciones siguen siendo participación numérica.",
        "no": "No menciona enfoque de género, o solo afirma «transversaliza género» sin contenido.",
        "na": "Si el contexto no permite enfoque transformador, debe justificarse explícitamente.",
        "notas": "Guía: «Guía práctica» n.º 15 — Incorporación de la perspectiva de género.",
    },
    # ---------------- SECCIÓN 2 — VALIDEZ DEL DISEÑO ----------------
    {
        "id": "2.1.1",
        "seccion": "2. Validez del diseño",
        "subseccion": "2.1",
        "head": "¿Han participado los socios tripartitos y otras partes interesadas clave en el diseño?",
        "criterio": "La propuesta describe el proceso de participación de constituyentes y otras partes interesadas en el diseño, y hace referencia específica a su papel en seguimiento y ejecución.",
        "tipo": "Lista de verificación",
        "aplicabilidad": "Siempre",
        "transv": "Tripartismo",
        "elementos": (
            "A. Describe el proceso de consulta/participación durante el diseño (qué, cuándo, con quién).\n"
            "B. Lista los constituyentes consultados (gobierno, empleadores, trabajadores).\n"
            "C. Define el rol de las partes en el seguimiento.\n"
            "D. Define el rol de las partes en la ejecución."
        ),
        "si": "Cumple A, B, C y D con detalle suficiente para ser auditable.",
        "parcial": "Cumple A y B pero C/D son enunciados generales («participarán en el seguimiento»).",
        "no": "Solo lista actores sin describir el proceso ni los roles.",
        "na": "",
        "notas": "Tripartismo OIT: gobierno + organizaciones de empleadores + organizaciones de trabajadores.",
    },
    {
        "id": "2.1.2",
        "seccion": "2. Validez del diseño",
        "subseccion": "2.1",
        "head": "¿Han participado los socios tripartitos y otras partes interesadas clave en el diseño?",
        "criterio": "Las partes interesadas clave se identifican claramente y se presentan en relación con su vínculo con los beneficiarios finales y su papel respecto al problema y la solución propuesta.",
        "tipo": "Binario calidad",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Identificación nominal de partes interesadas clave.\n"
            "B. Vínculo de cada parte con los beneficiarios finales.\n"
            "C. Rol respecto al problema (causa, afectado, mitigador).\n"
            "D. Rol respecto a la solución (ejecutor, validador, beneficiario)."
        ),
        "si": "Cumple A, B, C y D para todas las partes clave (no solo dos o tres).",
        "parcial": "Identifica partes pero los roles son genéricos o solo se cubren para algunas.",
        "no": "Lista de partes sin análisis de rol/vínculo.",
        "na": "",
        "notas": "",
    },
    {
        "id": "2.1.3",
        "seccion": "2. Validez del diseño",
        "subseccion": "2.1",
        "head": "¿Han participado los socios tripartitos y otras partes interesadas clave en el diseño?",
        "criterio": "Se ha consultado a especialistas de ACT/EMP y ACTRAV.",
        "tipo": "Binario presencia",
        "aplicabilidad": "Siempre",
        "transv": "Tripartismo",
        "elementos": (
            "A. Mención explícita de consulta con ACT/EMP.\n"
            "B. Mención explícita de consulta con ACTRAV.\n"
            "C. Resultado de la consulta integrado en el diseño."
        ),
        "si": "Cumple A, B y muestra integración (C).",
        "parcial": "Solo menciona consulta a uno de los dos, o menciona ambos pero sin evidencia de integración.",
        "no": "No menciona consulta a ACT/EMP ni ACTRAV.",
        "na": "",
        "notas": "ACT/EMP = Oficina de Actividades para los Empleadores. ACTRAV = Oficina de Actividades para los Trabajadores.",
    },
    {
        "id": "2.2.1",
        "seccion": "2. Validez del diseño",
        "subseccion": "2.2",
        "head": "¿Existe apropiación del proyecto por parte de los grupos destinatarios?",
        "criterio": "La propuesta describe consultas y acuerdos con socios y su disposición a contribuir a los resultados (intención de reforma, tiempo de personal para capacitación, aplicación de cambios, etc.).",
        "tipo": "Lista de verificación",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Describe consultas realizadas (cuándo, con quién, sobre qué).\n"
            "B. Lista compromisos concretos asumidos por socios.\n"
            "C. Esos compromisos son operativos (tiempo, recursos, decisiones) — no solo declarativos."
        ),
        "si": "Cumple A, B y C con compromisos operativos verificables.",
        "parcial": "Cumple A y B pero los compromisos quedan en declaración de interés sin operativizar.",
        "no": "Solo afirmaciones de apropiación sin evidencia de consulta o compromiso.",
        "na": "",
        "notas": "",
    },
    {
        "id": "2.2.2",
        "seccion": "2. Validez del diseño",
        "subseccion": "2.2",
        "head": "¿Existe apropiación del proyecto por parte de los grupos destinatarios?",
        "criterio": "Cada socio ha aceptado los objetivos y metas del proyecto (según corresponda), el marco de desempeño y las obligaciones y responsabilidades correspondientes.",
        "tipo": "Binario presencia",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Constancia de aceptación de objetivos por cada socio clave.\n"
            "B. Aceptación del marco de desempeño / indicadores.\n"
            "C. Aceptación de obligaciones y responsabilidades."
        ),
        "si": "Hay constancia (carta de intención, MoU, acta de reunión, sección dedicada) que cubre A, B y C para los socios principales.",
        "parcial": "Constancia parcial (cubre solo a algunos socios o solo algunos elementos).",
        "no": "Afirmación general sin constancia documental.",
        "na": "",
        "notas": "",
    },
    {
        "id": "2.2.3",
        "seccion": "2. Validez del diseño",
        "subseccion": "2.2",
        "head": "¿Existe apropiación del proyecto por parte de los grupos destinatarios?",
        "criterio": "La propuesta explica la estrategia mediante la cual se fomentará y garantizará de forma continua la apropiación por parte de las partes interesadas.",
        "tipo": "Calidad narrativa",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Estrategia explícita de apropiación continua (no solo eventos puntuales).\n"
            "B. Mecanismos de gobernanza compartida (comités, grupos de dirección).\n"
            "C. Vinculación con la estrategia de sostenibilidad."
        ),
        "si": "Cumple A, B y C con estrategia diferenciada por tipo de actor.",
        "parcial": "Estrategia presente pero genérica o uniforme para todos los actores.",
        "no": "No describe estrategia de apropiación continua.",
        "na": "",
        "notas": "",
    },
    {
        "id": "2.3.1",
        "seccion": "2. Validez del diseño",
        "subseccion": "2.3",
        "head": "¿Establece la propuesta un enfoque claro y holístico para el desarrollo de capacidades?",
        "criterio": "Se ha realizado o se prevé realizar una evaluación de la capacidad de los grupos destinatarios y sus conclusiones se mencionan explícitamente y se integran en el diseño.",
        "tipo": "Binario presencia con calidad",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Evaluación de capacidades realizada O prevista con metodología.\n"
            "B. Conclusiones explícitas (capacidades existentes vs brechas).\n"
            "C. Esas conclusiones informan el diseño del proyecto."
        ),
        "si": "Cumple A, B y C con vinculación visible.",
        "parcial": "Cumple A y B pero la integración con el diseño no es explícita.",
        "no": "Menciona «se evaluarán las capacidades» sin metodología ni conclusiones.",
        "na": "",
        "notas": "",
    },
    {
        "id": "2.3.2",
        "seccion": "2. Validez del diseño",
        "subseccion": "2.3",
        "head": "¿Establece la propuesta un enfoque claro y holístico para el desarrollo de capacidades?",
        "criterio": "La propuesta establece su enfoque del desarrollo de capacidades, abordando capacidades individuales (técnicas), organizativas y entorno propicio (político).",
        "tipo": "Lista de verificación",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Capacidades individuales (técnicas).\n"
            "B. Capacidades organizativas (sistemas, procesos).\n"
            "C. Entorno propicio (marco legal, políticas, financiación)."
        ),
        "si": "Cumple A, B y C con acciones específicas en cada nivel.",
        "parcial": "Cubre solo 1–2 niveles, o los tres pero con énfasis casi exclusivo en uno (típicamente solo capacitación individual).",
        "no": "Reduce «desarrollo de capacidades» a talleres y capacitaciones sin abordar lo organizativo o el entorno.",
        "na": "",
        "notas": "Guía: «Guía práctica» n.º 12 — Desarrollo de capacidades.",
    },
    {
        "id": "2.4.1",
        "seccion": "2. Validez del diseño",
        "subseccion": "2.4",
        "head": "¿Podrán los beneficiarios garantizar la sostenibilidad de los resultados?",
        "criterio": "De la propuesta se desprende que los constituyentes y/u otras partes interesadas han demostrado un gran interés en contribuir a la obtención de resultados tras la finalización del proyecto.",
        "tipo": "Binario calidad",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Evidencia de interés explícito de constituyentes en post-proyecto.\n"
            "B. Esa evidencia es concreta (carta, MoU, asignación de personal/presupuesto futuro)."
        ),
        "si": "Cumple A y B con evidencia concreta y verificable.",
        "parcial": "Hay declaraciones de interés pero sin operativización.",
        "no": "No hay evidencia de interés post-proyecto.",
        "na": "",
        "notas": "",
    },
    {
        "id": "2.4.2",
        "seccion": "2. Validez del diseño",
        "subseccion": "2.4",
        "head": "¿Podrán los beneficiarios garantizar la sostenibilidad de los resultados?",
        "criterio": "La propuesta incluye planes para garantizar la sostenibilidad o los resultados tras la finalización (formación de formadores, asociaciones con actores locales, etc.).",
        "tipo": "Lista de verificación",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Mecanismo institucional (formación de formadores, integración en sistemas existentes).\n"
            "B. Mecanismo financiero (fuentes de financiación post-proyecto).\n"
            "C. Mecanismo de gobernanza (quién mantiene los resultados).\n"
            "D. Cronograma de transición."
        ),
        "si": "Cumple A, C y al menos uno de B/D.",
        "parcial": "Solo cubre A; financiamiento y gobernanza post-proyecto quedan ambiguos.",
        "no": "Solo afirma «el proyecto será sostenible».",
        "na": "",
        "notas": "",
    },
    {
        "id": "2.4.3",
        "seccion": "2. Validez del diseño",
        "subseccion": "2.4",
        "head": "¿Podrán los beneficiarios garantizar la sostenibilidad de los resultados?",
        "criterio": "En la fase de prueba, la propuesta identifica las condiciones previas necesarias para ampliar el enfoque y presenta el enfoque para demostrar objetivamente el éxito del mismo.",
        "tipo": "Lista condicional",
        "aplicabilidad": "Condicional: aplica si la propuesta es piloto o tiene fase de prueba",
        "transv": "Ninguno",
        "elementos": (
            "A. La propuesta identifica condiciones previas para escalar.\n"
            "B. Define criterios de éxito objetivos.\n"
            "C. Plan de medición/demostración del éxito.\n"
            "D. Plan de transición a fase de escala."
        ),
        "si": "Cumple A, B, C y D.",
        "parcial": "Cumple A y B pero C/D son genéricos.",
        "no": "Identifica como piloto pero no define cómo se evaluará la pertinencia de escalar.",
        "na": "Documentar si el proyecto no es piloto / no tiene fase de prueba.",
        "notas": "",
    },
    {
        "id": "2.4.4",
        "seccion": "2. Validez del diseño",
        "subseccion": "2.4",
        "head": "¿Podrán los beneficiarios garantizar la sostenibilidad de los resultados?",
        "criterio": "La propuesta hace plausible que los recursos necesarios (humanos y financieros) estarán disponibles una vez finalizado el proyecto.",
        "tipo": "Calidad narrativa",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Identifica la fuente de recursos humanos post-proyecto.\n"
            "B. Identifica la fuente de recursos financieros post-proyecto.\n"
            "C. Argumenta plausibilidad (no solo declarativo)."
        ),
        "si": "Cumple A, B y C con sustento (presupuesto sectorial existente, compromiso institucional, etc.).",
        "parcial": "Cubre solo A o solo B; o cubre ambos pero la plausibilidad es débil.",
        "no": "No menciona disponibilidad de recursos post-proyecto.",
        "na": "",
        "notas": "",
    },
    {
        "id": "2.4.5",
        "seccion": "2. Validez del diseño",
        "subseccion": "2.4",
        "head": "¿Podrán los beneficiarios garantizar la sostenibilidad de los resultados?",
        "criterio": "El proyecto incluye una estrategia de salida que establece planes para la transferencia de responsabilidades y desarrollo de capacidades para que socios nacionales mantengan los resultados.",
        "tipo": "Lista de verificación",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Estrategia de salida explícita.\n"
            "B. Plan de transferencia de responsabilidades (qué, a quién, cuándo).\n"
            "C. Acciones de desarrollo de capacidades específicas para la transferencia.\n"
            "D. Cronograma de salida."
        ),
        "si": "Cumple A, B, C y D.",
        "parcial": "Cumple A y B pero falta C o D.",
        "no": "No incluye estrategia de salida o solo enuncia «se realizará una salida ordenada».",
        "na": "",
        "notas": "",
    },
    # ---------------- SECCIÓN 3 — MARCO DE RESULTADOS ----------------
    {
        "id": "3.1.1",
        "seccion": "3. Marco de resultados / R&M",
        "subseccion": "3.1",
        "head": "¿Se expresa claramente la teoría del cambio?",
        "criterio": "Las relaciones causales, mecanismos de cambio y supuestos entre productos, resultados e impacto se explican de forma clara y convincente. El peso recae en el logro de resultados (no en la realización de actividades).",
        "tipo": "Calidad narrativa",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Cadena causal productos → resultados → impacto explicitada.\n"
            "B. Mecanismos de cambio identificados (no solo flechas).\n"
            "C. Supuestos críticos enunciados.\n"
            "D. Énfasis discursivo en resultados (no en lista de actividades)."
        ),
        "si": "Cumple A, B, C y D con narrativa coherente.",
        "parcial": "Tiene cadena causal y supuestos, pero el énfasis sigue siendo actividades; o cadena clara pero mecanismos implícitos.",
        "no": "No hay teoría del cambio explícita; la lógica del proyecto se infiere de la lista de actividades.",
        "na": "",
        "notas": "",
    },
    {
        "id": "3.1.2",
        "seccion": "3. Marco de resultados / R&M",
        "subseccion": "3.1",
        "head": "¿Se expresa claramente la teoría del cambio?",
        "criterio": "Los productos parecen suficientes para el probable logro de los resultados. La contribución esperada de los resultados al impacto también es clara.",
        "tipo": "Calidad narrativa",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Cobertura: los productos cubren los resultados sin saltos lógicos.\n"
            "B. Suficiencia: cada resultado tiene productos que justifican alcanzarlo.\n"
            "C. Vínculo resultados → impacto razonado."
        ),
        "si": "Cumple A, B y C — el revisor puede trazar cada resultado a productos que lo sostienen y a impacto que lo continúa.",
        "parcial": "Hay saltos en uno o dos resultados (no es claro cómo los productos llevan al resultado), o el impacto queda demasiado lejos.",
        "no": "Saltos generalizados; productos y resultados parecen desacoplados.",
        "na": "",
        "notas": "",
    },
    {
        "id": "3.1.3",
        "seccion": "3. Marco de resultados / R&M",
        "subseccion": "3.1",
        "head": "¿Se expresa claramente la teoría del cambio?",
        "criterio": "De la teoría del cambio se desprende qué actores se beneficiarán del desarrollo de capacidades y cómo se espera que conduzca a resultados (mejor desempeño, cambio de comportamiento, etc.).",
        "tipo": "Calidad narrativa",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Actores específicos del desarrollo de capacidades nombrados.\n"
            "B. Tipo de cambio esperado (desempeño, comportamiento, prácticas).\n"
            "C. Mecanismo por el cual la capacidad se traduce en cambio."
        ),
        "si": "Cumple A, B y C con especificidad.",
        "parcial": "Cumple A y B pero el mecanismo (C) queda implícito.",
        "no": "Solo afirma «se desarrollarán capacidades» sin especificar actores ni cambios esperados.",
        "na": "",
        "notas": "",
    },
    {
        "id": "3.1.4",
        "seccion": "3. Marco de resultados / R&M",
        "subseccion": "3.1",
        "head": "¿Se expresa claramente la teoría del cambio?",
        "criterio": "Se identifican los supuestos clave, incluidos los resultados de intervenciones pasadas, en curso o previstas (incluidas las de otros actores).",
        "tipo": "Lista de verificación",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Supuestos enunciados explícitamente.\n"
            "B. Supuestos vinculados a niveles de la cadena (productos / resultados / impacto).\n"
            "C. Referencia a intervenciones pasadas o paralelas (OIT u otros actores)."
        ),
        "si": "Cumple A, B y C.",
        "parcial": "Supuestos genéricos («apoyo político continuo») sin vinculación a niveles, o sin C.",
        "no": "No hay supuestos enunciados, o solo aparecen como riesgos.",
        "na": "",
        "notas": "Distinguir supuesto (lo que tiene que ser cierto) de riesgo (lo que puede salir mal).",
    },
    {
        "id": "3.1.5",
        "seccion": "3. Marco de resultados / R&M",
        "subseccion": "3.1",
        "head": "¿Se expresa claramente la teoría del cambio?",
        "criterio": "Las experiencias anteriores recopiladas a partir de evaluaciones se citan como prueba para respaldar la relación causal prevista entre productos, resultados e impacto.",
        "tipo": "Lista de verificación",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Cita evaluaciones específicas (proyecto, programa, estudio).\n"
            "B. Esas evaluaciones respaldan vínculos causales del marco lógico.\n"
            "C. Las lecciones se aplican explícitamente al diseño actual."
        ),
        "si": "Cumple A, B y C con vinculación visible.",
        "parcial": "Cita evaluaciones pero el vínculo con la cadena causal es genérico.",
        "no": "No cita evaluaciones, o solo dice «se ha aprendido de experiencias pasadas».",
        "na": "",
        "notas": "Solapa con 1.4.3 — pero aquí el foco es respaldar relaciones causales del ToC.",
    },
    {
        "id": "3.2.1",
        "seccion": "3. Marco de resultados / R&M",
        "subseccion": "3.2",
        "head": "¿Las formulaciones de resultados reflejan situaciones concretas?",
        "criterio": "Se utiliza un «lenguaje de cambio» en lugar de «lenguaje de acción» («a través de esto y aquello», «haciendo esto y aquello»).",
        "tipo": "Binario",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Las formulaciones de resultados expresan estados/situaciones (X está fortalecido, Y tiene acceso).\n"
            "B. Ausencia de fórmulas de actividad («mediante», «a través de», «realizando»)."
        ),
        "si": "Todos los resultados (o casi todos) usan lenguaje de cambio.",
        "parcial": "Mezcla — algunos resultados en lenguaje de cambio y otros en lenguaje de acción.",
        "no": "Predomina el lenguaje de acción en las formulaciones de resultados.",
        "na": "",
        "notas": "Manual de gestión basada en resultados del GNUD.",
    },
    {
        "id": "3.2.2",
        "seccion": "3. Marco de resultados / R&M",
        "subseccion": "3.2",
        "head": "¿Las formulaciones de resultados reflejan situaciones concretas?",
        "criterio": "Los resultados (objetivos inmediatos) establecen la situación final, las condiciones de observación y el estándar a cumplir. Son SMART (específicos, medibles, alcanzables, pertinentes, con plazo).",
        "tipo": "Lista de verificación SMART",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "S — Específico (qué exactamente).\n"
            "M — Medible (con qué indicador y umbral).\n"
            "A — Alcanzable (factible con los recursos).\n"
            "R — Relevante (vinculado al problema y la ToC).\n"
            "T — Temporal (con plazo definido)."
        ),
        "si": "Todos los resultados cumplen los 5 atributos SMART.",
        "parcial": "Resultados cumplen 3–4 atributos; típicamente faltan M (no hay umbral) o T (sin plazo).",
        "no": "Resultados son aspiraciones (cumplen ≤2 atributos).",
        "na": "",
        "notas": "",
    },
    {
        "id": "3.2.3",
        "seccion": "3. Marco de resultados / R&M",
        "subseccion": "3.2",
        "head": "¿Las formulaciones de resultados reflejan situaciones concretas?",
        "criterio": "Las formulaciones de los productos se refieren al resultado de las actividades, no a las actividades en sí mismas.",
        "tipo": "Binario",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Productos formulados como salidas («Documento X publicado», «N personas certificadas»).\n"
            "B. Ausencia de formulaciones de actividad («Realizar talleres», «Capacitar a...»)."
        ),
        "si": "Todos (o casi todos) los productos formulados como salidas.",
        "parcial": "Mezcla productos y actividades.",
        "no": "Productos formulados como actividades.",
        "na": "",
        "notas": "",
    },
    {
        "id": "3.3.1",
        "seccion": "3. Marco de resultados / R&M",
        "subseccion": "3.3",
        "head": "¿Se incorporaron igualdad de género e inclusión de discapacidad en el marco lógico?",
        "criterio": "La igualdad de género se refleja en el nivel de los resultados y/o productos.",
        "tipo": "Transversal",
        "aplicabilidad": "Siempre",
        "transv": "Género",
        "elementos": (
            "A. Resultado o producto cuyo título nombra género/igualdad/mujeres.\n"
            "B. Indicador desagregado por sexo o específico de género.\n"
            "C. Meta cuantificable por sexo o de género."
        ),
        "si": f"Cumple A y al menos B o C en el marco lógico. {TRANSV_GATE}",
        "parcial": f"Solo aparece en supuestos o lenguaje transversal del marco; sin elemento DEDICADO. {TRANSV_GATE}",
        "no": "Marco lógico no menciona género en resultados, productos ni indicadores.",
        "na": "Solo si el marcador de género del CPO es 0 — debe documentarse.",
        "notas": "",
    },
    {
        "id": "3.3.2",
        "seccion": "3. Marco de resultados / R&M",
        "subseccion": "3.3",
        "head": "¿Se incorporaron igualdad de género e inclusión de discapacidad en el marco lógico?",
        "criterio": "La inclusión de las personas con discapacidad se refleja en el nivel de los resultados y/o productos.",
        "tipo": "Transversal",
        "aplicabilidad": "Siempre",
        "transv": "Discapacidad",
        "elementos": (
            "A. Resultado o producto que nombra discapacidad.\n"
            "B. Indicador específico de discapacidad o desagregado.\n"
            "C. Meta cuantificable de discapacidad."
        ),
        "si": f"Cumple A y al menos B o C. {TRANSV_GATE}",
        "parcial": f"Mención solo a nivel de supuestos o transversal. {TRANSV_GATE}",
        "no": "Marco lógico no incluye discapacidad.",
        "na": "Solo si la etiqueta de discapacidad del CPO no aplica — debe documentarse.",
        "notas": "",
    },
    {
        "id": "3.4.1",
        "seccion": "3. Marco de resultados / R&M",
        "subseccion": "3.4",
        "head": "¿Se han evaluado los riesgos y propuesto mitigaciones?",
        "criterio": "La propuesta contiene un resumen narrativo/análisis de los riesgos.",
        "tipo": "Binario presencia",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Existe una sección o anexo de análisis de riesgos.\n"
            "B. Los riesgos están categorizados (estratégico, operacional, fiduciario, contextual)."
        ),
        "si": "Cumple A y B.",
        "parcial": "Cumple A pero los riesgos están listados sin categorización.",
        "no": "No hay análisis de riesgos.",
        "na": "",
        "notas": "Guía: «Guía práctica» n.º 1 — Identificación y gestión de riesgos.",
    },
    {
        "id": "3.4.2",
        "seccion": "3. Marco de resultados / R&M",
        "subseccion": "3.4",
        "head": "¿Se han evaluado los riesgos y propuesto mitigaciones?",
        "criterio": "La propuesta evalúa y mitiga los riesgos de explotación y abuso sexuales (EAS), incluyendo afluencia de trabajadores, dinámicas de poder y exposición a entornos inseguros. Hay mecanismos de prevención.",
        "tipo": "Transversal",
        "aplicabilidad": "Siempre",
        "transv": "EAS (Explotación y abuso sexuales); Género",
        "elementos": (
            "A. Riesgos específicos de EAS identificados.\n"
            "B. Mecanismos de prevención (códigos de conducta, capacitación obligatoria).\n"
            "C. Mecanismo de denuncia accesible y confidencial.\n"
            "D. Protocolo de respuesta y atención a víctimas.\n"
            "E. Cláusulas EAS en contratos con terceros."
        ),
        "si": f"Cumple A, B, C y D; cumple E si hay terceros. {TRANSV_GATE}",
        "parcial": "Cumple A y B pero C/D son genéricos o ausentes.",
        "no": "Solo enuncia «cero tolerancia ante EAS» sin operacionalización.",
        "na": "",
        "notas": "Política OIT sobre EAS es estándar mínimo, no opcional.",
    },
    {
        "id": "3.4.3",
        "seccion": "3. Marco de resultados / R&M",
        "subseccion": "3.4",
        "head": "¿Se han evaluado los riesgos y propuesto mitigaciones?",
        "criterio": "La propuesta identifica y mitiga riesgos para la salud y seguridad de la comunidad (contaminación, tráfico, materiales peligrosos). Si hay personal de seguridad, hay medidas para prevenir abusos.",
        "tipo": "Lista condicional transversal",
        "aplicabilidad": "Condicional: aplica si hay infraestructura, presencia comunitaria, o personal de seguridad",
        "transv": "Comunidad; Medio ambiente; SST",
        "elementos": (
            "A. Identificación de riesgos comunitarios (no solo del personal del proyecto).\n"
            "B. Medidas para riesgos físicos (tráfico, materiales peligrosos, contaminación).\n"
            "C. Si hay personal de seguridad: protocolo de uso de la fuerza y prevención de abusos.\n"
            "D. Mecanismo de queja comunitario."
        ),
        "si": "Cumple A, B y D; cumple C cuando aplica.",
        "parcial": "Cubre riesgos del personal del proyecto pero no comunitarios; o cubre A y B sin D.",
        "no": "No identifica riesgos comunitarios.",
        "na": "Si el proyecto es 100% remoto/normativo sin presencia física, documentar.",
        "notas": "",
    },
    {
        "id": "3.4.4",
        "seccion": "3. Marco de resultados / R&M",
        "subseccion": "3.4",
        "head": "¿Se han evaluado los riesgos y propuesto mitigaciones?",
        "criterio": "Se mencionan/evalúan los riesgos de incumplimiento relacionados con normas laborales y observaciones de los organismos de supervisión.",
        "tipo": "Transversal",
        "aplicabilidad": "Siempre",
        "transv": "Normas Internacionales del Trabajo (NIT); Mecanismos de supervisión",
        "elementos": (
            "A. Identifica riesgos de incumplimiento de NIT en el contexto del proyecto.\n"
            "B. Vincula con observaciones del CEACR u otros organismos de supervisión.\n"
            "C. Plan de mitigación específico."
        ),
        "si": "Cumple A, B y C.",
        "parcial": "Cumple A pero sin B; o cita observaciones sin vincularlas a riesgos del proyecto.",
        "no": "No considera riesgos de cumplimiento NIT.",
        "na": "",
        "notas": "",
    },
    {
        "id": "3.4.5",
        "seccion": "3. Marco de resultados / R&M",
        "subseccion": "3.4",
        "head": "¿Se han evaluado los riesgos y propuesto mitigaciones?",
        "criterio": "Para proyectos con presupuesto superior a 1 millón de USD, se ha rellenado la última versión del registro de riesgos y se ha adjuntado a la propuesta.",
        "tipo": "Condicional binario",
        "aplicabilidad": "Condicional: aplica si presupuesto > 1.000.000 USD",
        "transv": "Ninguno",
        "elementos": (
            "A. Registro de riesgos en formato OIT vigente.\n"
            "B. Adjunto a la propuesta (anexo, no solo referencia)."
        ),
        "si": "Cumple A y B.",
        "parcial": "Hay registro pero no es la versión vigente, o no está adjunto pero se referencia.",
        "no": "Presupuesto > 1M USD y no hay registro de riesgos.",
        "na": "Documentar si presupuesto ≤ 1M USD.",
        "notas": "",
    },
    {
        "id": "3.4.6",
        "seccion": "3. Marco de resultados / R&M",
        "subseccion": "3.4",
        "head": "¿Se han evaluado los riesgos y propuesto mitigaciones?",
        "criterio": "Los riesgos identificados son relevantes y suficientemente específicos para ser significativos.",
        "tipo": "Calidad narrativa",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Especificidad: cada riesgo es contextual al proyecto, no plantilla genérica.\n"
            "B. Probabilidad e impacto valorados.\n"
            "C. Riesgos materiales para el éxito del proyecto."
        ),
        "si": "Cumple A, B y C — los riesgos son del proyecto, no transcripciones de plantilla.",
        "parcial": "Algunos riesgos específicos, otros genéricos; o falta valoración de probabilidad/impacto.",
        "no": "Riesgos son plantilla institucional sin contextualización.",
        "na": "",
        "notas": "",
    },
    {
        "id": "3.4.7",
        "seccion": "3. Marco de resultados / R&M",
        "subseccion": "3.4",
        "head": "¿Se han evaluado los riesgos y propuesto mitigaciones?",
        "criterio": "Se proponen medidas pertinentes para abordar y supervisar los riesgos de nivel medio y alto.",
        "tipo": "Lista de verificación",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Cada riesgo medio/alto tiene medida de mitigación.\n"
            "B. Cada medida tiene responsable.\n"
            "C. Cada medida tiene método de monitoreo."
        ),
        "si": "Cumple A, B y C para todos los riesgos medio/alto.",
        "parcial": "Cumple A pero falta B o C en varios.",
        "no": "Riesgos medio/alto sin medidas, o medidas genéricas sin responsable ni monitoreo.",
        "na": "",
        "notas": "",
    },
    {
        "id": "3.5.1",
        "seccion": "3. Marco de resultados / R&M",
        "subseccion": "3.5",
        "head": "¿Se han establecido mecanismos adecuados de seguimiento y evaluación?",
        "criterio": "La sección de S&E describe el sistema previsto para recopilar información de manera precisa y oportuna, justificando los recursos. Incluye una revisión de evaluabilidad (obligatoria para presupuestos altos).",
        "tipo": "Lista de verificación",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Sistema de recopilación descrito (qué datos, con qué frecuencia, por quién).\n"
            "B. Justificación de recursos asignados a S&E.\n"
            "C. Revisión de evaluabilidad cuando aplica (proyecto > umbral).\n"
            "D. Plan de aprendizaje del proyecto."
        ),
        "si": "Cumple A, B; cumple C cuando aplica; cumple D.",
        "parcial": "Cubre A pero la justificación de recursos (B) o el plan de aprendizaje (D) faltan.",
        "no": "S&E reducido a «se realizará seguimiento».",
        "na": "Documentar criterio de umbral para C.",
        "notas": "",
    },
    {
        "id": "3.5.2",
        "seccion": "3. Marco de resultados / R&M",
        "subseccion": "3.5",
        "head": "¿Se han establecido mecanismos adecuados de seguimiento y evaluación?",
        "criterio": "Se describe el ciclo de retroalimentación de información sobre los avances en las decisiones de gestión, identificando claramente las necesidades de información para informes de desempeño. Se identifica al personal responsable.",
        "tipo": "Lista de verificación",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Ciclo de retroalimentación definido (datos → decisiones).\n"
            "B. Necesidades de información para informes especificadas.\n"
            "C. Personal responsable nombrado o por puesto.\n"
            "D. Frecuencia de los ciclos de revisión."
        ),
        "si": "Cumple A, B, C y D.",
        "parcial": "Cubre A y C pero B/D quedan implícitos.",
        "no": "No describe el ciclo de retroalimentación; los datos se recopilan sin uso definido.",
        "na": "",
        "notas": "",
    },
    {
        "id": "3.5.3",
        "seccion": "3. Marco de resultados / R&M",
        "subseccion": "3.5",
        "head": "¿Se han establecido mecanismos adecuados de seguimiento y evaluación?",
        "criterio": "El plan de S&E se ha elaborado con necesidades de información bien identificadas. Los métodos de recopilación y análisis son técnicamente adecuados. Especifica funciones y responsabilidades.",
        "tipo": "Lista de verificación",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Plan de S&E como documento o sección discreta.\n"
            "B. Métodos de recopilación nombrados (encuesta, entrevista, registros administrativos).\n"
            "C. Métodos de análisis especificados.\n"
            "D. Roles y responsabilidades por método/indicador."
        ),
        "si": "Cumple A, B, C y D con coherencia entre métodos y necesidades.",
        "parcial": "Cumple A y B, pero análisis y roles son genéricos.",
        "no": "Plan de S&E ausente o reducido a una tabla de indicadores sin métodos.",
        "na": "",
        "notas": "",
    },
    {
        "id": "3.5.4",
        "seccion": "3. Marco de resultados / R&M",
        "subseccion": "3.5",
        "head": "¿Se han establecido mecanismos adecuados de seguimiento y evaluación?",
        "criterio": "El plan de S&E incluye indicadores SMART y sensibles al género, líneas de base, objetivos e hitos (que también permiten inclusión de personas con discapacidad y/u otros grupos vulnerables). Los indicadores miden los resultados previstos.",
        "tipo": "Lista transversal SMART",
        "aplicabilidad": "Siempre",
        "transv": "Género; Discapacidad",
        "elementos": (
            "A. Indicadores SMART (los 5 atributos).\n"
            "B. Sensibles al género (desagregados por sexo o específicos de género).\n"
            "C. Permiten inclusión de discapacidad (desagregados o específicos).\n"
            "D. Línea de base para cada indicador.\n"
            "E. Meta y hitos definidos.\n"
            "F. Cobertura de los resultados (no solo productos)."
        ),
        "si": f"Cumple A, B, C, D, E y F. {TRANSV_GATE}",
        "parcial": "Cumple A, D, E pero B/C son genéricos (lenguaje inclusivo sin desagregación real).",
        "no": "Indicadores no son SMART, o no hay línea de base, o ignoran género/discapacidad.",
        "na": "",
        "notas": "Este es un criterio compuesto crítico — vale la pena reportar el desglose A–F en el output del LLM.",
    },
    {
        "id": "3.5.5",
        "seccion": "3. Marco de resultados / R&M",
        "subseccion": "3.5",
        "head": "¿Se han establecido mecanismos adecuados de seguimiento y evaluación?",
        "criterio": "El presupuesto de S&E se ajusta a las directrices de la OIT. El presupuesto de evaluación figura en partida separada y es adecuado al tamaño/duración (≈2% del presupuesto total).",
        "tipo": "Binario calibración",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Partida presupuestaria de evaluación separada.\n"
            "B. Monto ≥ ~2% del presupuesto total (o justificación si difiere).\n"
            "C. Adecuación a tamaño y duración."
        ),
        "si": "Cumple A, B y C.",
        "parcial": "Cumple A pero el monto es menor sin justificación; o no hay partida separada pero se contempla un monto identificable.",
        "no": "No hay partida ni monto identificable de evaluación.",
        "na": "",
        "notas": "Verificar el % exacto contra directrices de la OIT vigentes.",
    },
    {
        "id": "3.6.1",
        "seccion": "3. Marco de resultados / R&M",
        "subseccion": "3.6",
        "head": "¿Se pueden lograr todos los resultados dentro del plazo y presupuesto?",
        "criterio": "Si hay período inicial, la propuesta presenta actividades y resultados de ese período. El plazo es suficiente, considerando capacidades de socios, complejidad y dinámicas de cambio.",
        "tipo": "Lista condicional",
        "aplicabilidad": "Condicional: si hay fase inicial / inception",
        "transv": "Ninguno",
        "elementos": (
            "A. Si hay inception: actividades y resultados del período inicial explicitados.\n"
            "B. Plazo total justificado contra complejidad y capacidades.\n"
            "C. Cronograma con hitos por trimestre o semestre."
        ),
        "si": "Cumple A (cuando aplica), B y C.",
        "parcial": "Plazo enunciado pero no justificado, o cronograma a alto nivel sin hitos.",
        "no": "Plazo se afirma sin sustento; cronograma ausente.",
        "na": "Documentar si no hay fase inicial.",
        "notas": "",
    },
    {
        "id": "3.6.2",
        "seccion": "3. Marco de resultados / R&M",
        "subseccion": "3.6",
        "head": "¿Se pueden lograr todos los resultados dentro del plazo y presupuesto?",
        "criterio": "Hay pruebas claras de que el presupuesto se basa en un análisis de cada actividad e incluye todos los recursos necesarios.",
        "tipo": "Binario calidad",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Presupuesto desglosado por actividad o producto.\n"
            "B. Coherencia entre actividades del marco lógico y partidas presupuestarias.\n"
            "C. Costos unitarios o cálculos visibles."
        ),
        "si": "Cumple A, B y C — el revisor puede trazar cada actividad a una partida.",
        "parcial": "Presupuesto agregado por categoría (personal/operativo) sin trazabilidad a actividades.",
        "no": "Presupuesto solo a nivel de totales por categoría amplia.",
        "na": "",
        "notas": "",
    },
    {
        "id": "3.6.3",
        "seccion": "3. Marco de resultados / R&M",
        "subseccion": "3.6",
        "head": "¿Se pueden lograr todos los resultados dentro del plazo y presupuesto?",
        "criterio": "Cuando aplica, el presupuesto incluye disposiciones para conocimientos especializados específicos (género, discapacidad) y medios que promueven la inclusión (idiomas locales, lengua de señas, accesibilidad).",
        "tipo": "Lista transversal",
        "aplicabilidad": "Siempre",
        "transv": "Género; Discapacidad; Pueblos indígenas",
        "elementos": (
            "A. Partida para especialista de género cuando aplica.\n"
            "B. Partida para especialista de discapacidad cuando aplica.\n"
            "C. Partida para accesibilidad de información (lenguaje claro, formatos alternativos).\n"
            "D. Partida para interpretación a idiomas locales o lengua de señas cuando aplica."
        ),
        "si": f"Cumple los elementos aplicables al proyecto, con partidas específicas y montos. {TRANSV_GATE}",
        "parcial": "Reconoce la necesidad pero queda como «si se requiere» sin partida.",
        "no": "Sin partidas específicas; se asume que las consideraciones transversales no tienen costo.",
        "na": "Documentar elemento por elemento si no aplica.",
        "notas": "",
    },
    {
        "id": "3.7.1",
        "seccion": "3. Marco de resultados / R&M",
        "subseccion": "3.7",
        "head": "¿Existe evidencia de consideración de rentabilidad/relación calidad-precio?",
        "criterio": "En la propuesta se hace referencia a consideraciones de rentabilidad.",
        "tipo": "Binario presencia",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Mención explícita de rentabilidad o «value for money».\n"
            "B. Vinculada al diseño (no solo a la administración)."
        ),
        "si": "Cumple A y B.",
        "parcial": "Mención superficial («se buscará la rentabilidad») sin vínculo con decisiones.",
        "no": "No menciona rentabilidad.",
        "na": "",
        "notas": "",
    },
    {
        "id": "3.7.2",
        "seccion": "3. Marco de resultados / R&M",
        "subseccion": "3.7",
        "head": "¿Existe evidencia de consideración de rentabilidad/relación calidad-precio?",
        "criterio": "La propuesta muestra que se ha encontrado el equilibrio más adecuado entre el gasto/ahorro de costos y el aumento/disminución de los beneficios.",
        "tipo": "Calidad narrativa",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Análisis de alternativas de diseño con costo/beneficio.\n"
            "B. Justificación de la opción elegida.\n"
            "C. Cuando aplica: comparación con benchmarks (proyectos similares)."
        ),
        "si": "Cumple A, B y C.",
        "parcial": "Justifica el diseño pero sin comparar alternativas.",
        "no": "No analiza relación costo/beneficio.",
        "na": "",
        "notas": "",
    },
    # ---------------- SECCIÓN 4 — IMPLEMENTACIÓN ----------------
    {
        "id": "4.1.1",
        "seccion": "4. Implementación",
        "subseccion": "4.1",
        "head": "¿Son claras y suficientes las disposiciones de gestión?",
        "criterio": "Proyectos con resultados regionales/nacionales son gestionados por la oficina de terreno más cercana a los beneficiarios. Proyectos con gestión centralizada en sede justifican la decisión por eficacia, rentabilidad, capacidades, etc.",
        "tipo": "Binario calibración",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Identifica la oficina/unidad gestora.\n"
            "B. Si gestión es regional/nacional: oficina de terreno cercana a los beneficiarios.\n"
            "C. Si gestión centralizada: justificación explícita por criterios de eficacia."
        ),
        "si": "Cumple A y B (descentralizado) o A y C (centralizado con justificación).",
        "parcial": "Identifica la oficina pero la decisión de centralización no se justifica.",
        "no": "No identifica oficina gestora, o centralización sin justificación.",
        "na": "",
        "notas": "",
    },
    {
        "id": "4.1.2",
        "seccion": "4. Implementación",
        "subseccion": "4.1",
        "head": "¿Son claras y suficientes las disposiciones de gestión?",
        "criterio": "Las funciones y responsabilidades del personal del proyecto están claramente definidas. Existe línea clara de rendición de cuentas y supervisión, y se identifica al funcionario responsable de la OIT.",
        "tipo": "Lista de verificación",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Organigrama o descripción de roles del personal.\n"
            "B. Línea de rendición de cuentas explícita (a quién reporta cada quién).\n"
            "C. Funcionario OIT responsable identificado por puesto o nombre."
        ),
        "si": "Cumple A, B y C.",
        "parcial": "Cumple A pero la rendición de cuentas o el funcionario responsable no se identifican.",
        "no": "Solo enuncia «se contratará personal según se requiera».",
        "na": "",
        "notas": "",
    },
    {
        "id": "4.1.3",
        "seccion": "4. Implementación",
        "subseccion": "4.1",
        "head": "¿Son claras y suficientes las disposiciones de gestión?",
        "criterio": "El apoyo de las unidades técnicas o administrativas de la sede y/o equipos de trabajo directos se ha previsto en el presupuesto.",
        "tipo": "Binario presencia",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Identifica unidades técnicas o administrativas que apoyarán.\n"
            "B. Apoyo previsto en el presupuesto (partida o asignación)."
        ),
        "si": "Cumple A y B.",
        "parcial": "Identifica unidades pero el presupuesto no contempla el apoyo.",
        "no": "Asume apoyo de sede sin previsión presupuestaria ni identificación.",
        "na": "",
        "notas": "",
    },
    {
        "id": "4.1.4",
        "seccion": "4. Implementación",
        "subseccion": "4.1",
        "head": "¿Son claras y suficientes las disposiciones de gestión?",
        "criterio": "Las disposiciones de ejecución (dotación de personal, adquisiciones, sistemas financieros y autoridades) están claramente definidas.",
        "tipo": "Lista de verificación",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Plan de dotación de personal.\n"
            "B. Procedimientos de adquisiciones (estándares OIT o ajustes).\n"
            "C. Sistemas financieros y de reporte.\n"
            "D. Niveles de autoridad y aprobación."
        ),
        "si": "Cumple A, B, C y D.",
        "parcial": "Cubre 2–3 elementos; típicamente B o D quedan implícitos.",
        "no": "Solo enuncia «se aplicarán los procedimientos OIT».",
        "na": "",
        "notas": "",
    },
    {
        "id": "4.1.5",
        "seccion": "4. Implementación",
        "subseccion": "4.1",
        "head": "¿Son claras y suficientes las disposiciones de gestión?",
        "criterio": "Si hay contratistas, subcontratistas o proveedores externos, la propuesta menciona que se adhieren a principios de trabajo decente y empleo justo. Hay mecanismos de queja y seguimiento del cumplimiento.",
        "tipo": "Lista condicional transversal",
        "aplicabilidad": "Condicional: aplica si hay contratistas/subcontratistas/proveedores externos",
        "transv": "Trabajo decente; Normas Internacionales del Trabajo (NIT)",
        "elementos": (
            "A. Cláusulas de trabajo decente en contratos con terceros.\n"
            "B. Cláusulas de empleo justo (salario, jornada, libertad sindical).\n"
            "C. Mecanismo de queja accesible para trabajadores de terceros.\n"
            "D. Seguimiento del cumplimiento por la OIT."
        ),
        "si": "Cumple A, B, C y D.",
        "parcial": "Cubre A y B pero el mecanismo de queja o el seguimiento son débiles.",
        "no": "Solo afirma que «los terceros cumplirán las normas» sin operacionalización.",
        "na": "Documentar si no hay terceros.",
        "notas": "",
    },
    {
        "id": "4.2.1",
        "seccion": "4. Implementación",
        "subseccion": "4.2",
        "head": "¿Son claras las disposiciones de ejecución?",
        "criterio": "Las responsabilidades y funciones de las instituciones y socios participantes están claramente definidas. Hay procedimientos de ejecución. La propuesta indica por qué se han seleccionado determinadas instituciones.",
        "tipo": "Lista de verificación",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Roles de cada institución/socio explicitados.\n"
            "B. Procedimientos de ejecución (cómo trabajan juntos).\n"
            "C. Justificación de selección de socios (por qué estos y no otros)."
        ),
        "si": "Cumple A, B y C.",
        "parcial": "Cubre A y B pero la selección no se justifica (por qué este socio).",
        "no": "Solo lista socios sin roles ni justificación.",
        "na": "",
        "notas": "",
    },
    {
        "id": "4.2.2",
        "seccion": "4. Implementación",
        "subseccion": "4.2",
        "head": "¿Son claras las disposiciones de ejecución?",
        "criterio": "La propuesta indica si los socios participantes han aceptado las funciones y responsabilidades.",
        "tipo": "Binario presencia",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Constancia explícita de aceptación por cada socio.\n"
            "B. Documento de respaldo (carta, MoU, acta)."
        ),
        "si": "Cumple A y B (con respaldo documental referenciado).",
        "parcial": "Afirma aceptación sin respaldo verificable.",
        "no": "No menciona aceptación.",
        "na": "",
        "notas": "Solapa con 2.2.2 — aquí el foco es socios ejecutores.",
    },
    {
        "id": "4.2.3",
        "seccion": "4. Implementación",
        "subseccion": "4.2",
        "head": "¿Son claras las disposiciones de ejecución?",
        "criterio": "Se ha realizado o se realizará una evaluación de la capacidad organizativa de los socios ejecutores y se hace referencia a ella en la propuesta.",
        "tipo": "Binario presencia",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Evaluación de capacidad organizativa realizada o planificada.\n"
            "B. Referencia explícita en la propuesta (no solo en anexo no citado)."
        ),
        "si": "Cumple A y B.",
        "parcial": "Mención de la evaluación sin metodología ni resultados.",
        "no": "No menciona evaluación de capacidad organizativa.",
        "na": "",
        "notas": "",
    },
    {
        "id": "4.2.4",
        "seccion": "4. Implementación",
        "subseccion": "4.2",
        "head": "¿Son claras las disposiciones de ejecución?",
        "criterio": "Si es necesario, se incluye un plan de desarrollo de capacidad de los socios ejecutores, incluida la capacidad de respuesta en materia de género.",
        "tipo": "Lista condicional transversal",
        "aplicabilidad": "Condicional: aplica si la evaluación de capacidad organizativa identifica brechas",
        "transv": "Género",
        "elementos": (
            "A. Plan de desarrollo de capacidad para los socios ejecutores.\n"
            "B. Componentes operativos (capacitación, mentoría, acompañamiento).\n"
            "C. Componente de capacidad de respuesta en género."
        ),
        "si": f"Cumple A, B y C. {TRANSV_GATE}",
        "parcial": "Cumple A y B pero género queda en menciones de marco.",
        "no": "Identifica brechas pero no propone plan; o plan sin género.",
        "na": "Documentar si la evaluación no identificó brechas.",
        "notas": "",
    },
    {
        "id": "4.2.5",
        "seccion": "4. Implementación",
        "subseccion": "4.2",
        "head": "¿Son claras las disposiciones de ejecución?",
        "criterio": "La propuesta presenta pruebas suficientes de que los socios ejecutores garantizarán la ejecución eficaz del proyecto.",
        "tipo": "Calidad narrativa",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Antecedentes de los socios (proyectos similares, capacidad demostrada).\n"
            "B. Argumento de plausibilidad (no solo «son competentes»).\n"
            "C. Mitigación si la capacidad demostrada es limitada."
        ),
        "si": "Cumple A, B y C cuando aplica C.",
        "parcial": "Antecedentes presentes pero argumento débil; o antecedentes ausentes pero socios reconocidos.",
        "no": "Solo afirma que los socios ejecutarán bien.",
        "na": "",
        "notas": "",
    },
    {
        "id": "4.3.1",
        "seccion": "4. Implementación",
        "subseccion": "4.3",
        "head": "¿Hay estrategia de comunicación continua de avances con partes interesadas?",
        "criterio": "El proyecto describe cómo se comunicarán datos y resultados (incluidos resultados de evaluación) a partes interesadas y ha asignado recursos adecuados.",
        "tipo": "Lista de verificación",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Plan de comunicación con partes interesadas (qué, cuándo, cómo).\n"
            "B. Productos de comunicación esperados (reportes, briefings, eventos).\n"
            "C. Recursos humanos asignados.\n"
            "D. Recursos financieros asignados."
        ),
        "si": "Cumple A, B, C y D.",
        "parcial": "Cubre A y B pero recursos no asignados explícitamente.",
        "no": "Comunicación reducida a «se informará a las partes interesadas».",
        "na": "",
        "notas": "",
    },
    {
        "id": "4.4.1",
        "seccion": "4. Implementación",
        "subseccion": "4.4",
        "head": "¿Hay estrategia de comunicación presupuestada?",
        "criterio": "La propuesta cuenta con estrategia de comunicación pública (público general, perspectiva de interés humano) o información sobre cómo y cuándo se desarrollará.",
        "tipo": "Lista de verificación",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Estrategia de comunicación pública (no solo técnica).\n"
            "B. Cronograma de desarrollo si la estrategia no está aún elaborada.\n"
            "C. Productos esperados (notas de prensa, redes, multimedia)."
        ),
        "si": "Cumple A y C, o cumple B con cronograma claro.",
        "parcial": "Anuncia estrategia sin cronograma ni productos.",
        "no": "No incluye estrategia de comunicación pública.",
        "na": "",
        "notas": "",
    },
    {
        "id": "4.4.2",
        "seccion": "4. Implementación",
        "subseccion": "4.4",
        "head": "¿Hay estrategia de comunicación presupuestada?",
        "criterio": "Se asignan recursos adecuados (humanos y financieros) para desarrollo e implementación de la estrategia. Los productos son accesibles para partes interesadas clave (formatos alternativos para personas con discapacidad, idiomas locales).",
        "tipo": "Lista transversal",
        "aplicabilidad": "Siempre",
        "transv": "Discapacidad; Pueblos indígenas; Accesibilidad",
        "elementos": (
            "A. Recursos humanos asignados a comunicación.\n"
            "B. Recursos financieros asignados.\n"
            "C. Formatos accesibles para personas con discapacidad.\n"
            "D. Productos en idiomas locales / lenguaje claro / lengua de señas cuando aplica."
        ),
        "si": f"Cumple A, B, C y D cuando aplica. {TRANSV_GATE}",
        "parcial": "Cumple A y B pero accesibilidad/idiomas en lenguaje declarativo sin partidas.",
        "no": "Sin recursos asignados o sin consideración de accesibilidad.",
        "na": "",
        "notas": "",
    },
    {
        "id": "4.4.3",
        "seccion": "4. Implementación",
        "subseccion": "4.4",
        "head": "¿Hay estrategia de comunicación presupuestada?",
        "criterio": "Si el proyecto supera 5 millones de USD y se financia mediante PPP o por la Comisión Europea, se ha completado plantilla de estrategia de comunicación con DCOMM y oficina nacional.",
        "tipo": "Condicional binario",
        "aplicabilidad": "Condicional: presupuesto > 5.000.000 USD Y financiamiento PPP o CE",
        "transv": "Ninguno",
        "elementos": (
            "A. Plantilla DCOMM completada.\n"
            "B. Coordinación con la oficina nacional correspondiente."
        ),
        "si": "Cumple A y B (cuando aplica).",
        "parcial": "Cumple A pero sin coordinación con oficina nacional.",
        "no": "Aplican las dos condiciones (>5M y PPP/CE) pero la plantilla no está completada.",
        "na": "Documentar si presupuesto ≤ 5M USD o financiamiento no es PPP/CE.",
        "notas": "Contacto: comms-help@ilo.org",
    },
    # ---------------- SECCIÓN 5 — PRESENTACIÓN ----------------
    {
        "id": "5.1.1",
        "seccion": "5. Presentación",
        "subseccion": "5.1",
        "head": "¿La propuesta sigue la plantilla?",
        "criterio": "Se han completado todas las secciones y se han ordenado de acuerdo con los materiales de orientación.",
        "tipo": "Binario formal",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Todas las secciones de la plantilla OIT presentes.\n"
            "B. Orden conforme a la guía.\n"
            "C. Anexos requeridos incluidos."
        ),
        "si": "Cumple A, B y C.",
        "parcial": "Falta 1–2 secciones, o orden alterado pero contenido completo.",
        "no": "Faltan varias secciones o anexos clave.",
        "na": "",
        "notas": "",
    },
    {
        "id": "5.2.1",
        "seccion": "5. Presentación",
        "subseccion": "5.2",
        "head": "¿La propuesta está redactada de forma clara y concisa?",
        "criterio": "La calidad de redacción es alta; no hay detalles superfluos; las ideas centrales están redactadas de forma clara y directa.",
        "tipo": "Calidad narrativa",
        "aplicabilidad": "Siempre",
        "transv": "Ninguno",
        "elementos": (
            "A. Ideas centrales identificables sin re-lectura.\n"
            "B. Ausencia de jerga innecesaria o repetición.\n"
            "C. Estructura visible (subtítulos, tablas, viñetas cuando ayudan)."
        ),
        "si": "Cumple A, B y C — un revisor puede entender el proyecto en una lectura.",
        "parcial": "Texto mayormente claro pero con secciones densas o repetitivas.",
        "no": "Texto difícil de seguir; ideas centrales sepultadas.",
        "na": "",
        "notas": "",
    },
]


def build_xlsx(rows, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rúbrica Tab 1"

    # Note row at the top
    note = (
        "INSTRUCCIONES DE USO — Cada fila evalúa UN criterio (1.1.1, 1.1.2, ...). "
        "La «Pregunta orientadora» (1.1, 1.2, ...) es contexto y NO se evalúa. "
        "Para criterios marcados como transversales, aplicar el filtro DEDICADO vs MARCO. "
        "Los placeholders [LLENAR] deben completarse con citas textuales de propuestas reales "
        "(una vez calibrado, el sistema usará estos ejemplos como anclas)."
    )
    ws["A1"] = note
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=17)
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="center")
    ws["A1"].font = Font(italic=True, color="555555")
    ws["A1"].fill = PatternFill("solid", fgColor="FFF8DC")
    ws.row_dimensions[1].height = 60

    # Header row
    headers = [
        "ID",
        "Sección",
        "Subsección",
        "Pregunta orientadora (CONTEXTO — no evaluar)",
        "Criterio a evaluar",
        "Tipo de criterio",
        "Aplicabilidad",
        "Aspectos transversales",
        "Elementos a verificar",
        "Rúbrica — Sí",
        "Rúbrica — Parcial",
        "Rúbrica — No",
        "Rúbrica — No aplica",
        "Ejemplo evidencia — Sí [LLENAR]",
        "Ejemplo evidencia — Parcial [LLENAR]",
        "Ejemplo evidencia — No [LLENAR]",
        "Notas",
    ]
    header_fill = PatternFill("solid", fgColor="002F6C")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(wrap_text=True, vertical="center", horizontal="center")
    border = Border(
        left=Side(style="thin", color="888888"),
        right=Side(style="thin", color="888888"),
        top=Side(style="thin", color="888888"),
        bottom=Side(style="thin", color="888888"),
    )
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border
    ws.row_dimensions[2].height = 48

    # Data
    cell_align = Alignment(wrap_text=True, vertical="top", horizontal="left")
    seccion_fill = {
        "1. Pertinencia": PatternFill("solid", fgColor="EAF3FB"),
        "2. Validez del diseño": PatternFill("solid", fgColor="EFF7EE"),
        "3. Marco de resultados / R&M": PatternFill("solid", fgColor="FFF6E6"),
        "4. Implementación": PatternFill("solid", fgColor="FBEEEE"),
        "5. Presentación": PatternFill("solid", fgColor="F0EAF7"),
    }
    transv_font = Font(color="A0522D", italic=True)
    placeholder_font = Font(color="888888", italic=True)

    for row_idx, row in enumerate(rows, start=3):
        values = [
            row["id"],
            row["seccion"],
            row["subseccion"],
            row["head"],
            row["criterio"],
            row["tipo"],
            row["aplicabilidad"],
            row["transv"],
            row["elementos"],
            row["si"],
            row["parcial"],
            row["no"],
            row["na"],
            PLACEHOLDER_SI,
            PLACEHOLDER_PA,
            PLACEHOLDER_NO,
            row["notas"],
        ]
        fill = seccion_fill.get(row["seccion"])
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = cell_align
            cell.border = border
            if fill:
                cell.fill = fill
            if col_idx == 8 and val and val != "Ninguno":  # transv column
                cell.font = transv_font
            if col_idx in (14, 15, 16):  # placeholders
                cell.font = placeholder_font
        # tall row to fit content
        ws.row_dimensions[row_idx].height = 220

    # Column widths
    widths = [
        8,    # ID
        18,   # Sección
        12,   # Subsección
        38,   # Pregunta orientadora
        50,   # Criterio
        20,   # Tipo
        22,   # Aplicabilidad
        22,   # Transversales
        38,   # Elementos
        50,   # Rúbrica Sí
        50,   # Rúbrica Parcial
        50,   # Rúbrica No
        30,   # Rúbrica N/A
        32,   # Ej Sí
        32,   # Ej Parcial
        32,   # Ej No
        30,   # Notas
    ]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header
    ws.freeze_panes = "F3"

    wb.save(output_path)


if __name__ == "__main__":
    out = "./Rubrica_Tab1_Detallada.xlsx"
    build_xlsx(RUBRICA, out)
    print(f"Wrote {out} with {len(RUBRICA)} criteria")
