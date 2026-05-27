"""Apply the 30 client review comments to Rubrica_Tab1_Detallada_Full.xlsx.

Output: Rubrica_Tab1_Detallada_Full_v2.xlsx with:
  - Targeted cell edits resolving each comment
  - Highlighted (orange) cells where content was changed
  - New "Cambios v2 (revisión cliente)" column documenting each edit
  - Operational notes appended to Notas column for items that need app/ADMIN follow-up
  - Threaded comments stripped (resolved in-line)
"""

import shutil
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = "/Users/ageidv/ilo/deploy_3/Rubrica_Tab1_Detallada_Full.xlsx"
DST = "/Users/ageidv/ilo/deploy_3/Rubrica_Tab1_Detallada_Full_v2.xlsx"

# Column letters (data starts at row 3; row 1 = instructions, row 2 = header)
COL = {
    "ID": 1, "Sec": 2, "Sub": 3, "Head": 4, "Crit": 5, "Tipo": 6,
    "Apli": 7, "Transv": 8, "Elem": 9, "Si": 10, "Par": 11, "No": 12,
    "Na": 13, "EjSi": 14, "EjPa": 15, "EjNo": 16, "Notas": 17,
}
CAMBIOS_COL = 18  # new column

# Map ID -> row
ID_TO_ROW = {}

# Each edit: (id, target_col_key, action, new_value, change_summary)
# action: "replace" (overwrite cell) or "append_note" (append to Notas)
EDITS = [
    # 1.1.2 — question about whether element C is included; keep but flag pending
    ("1.1.2", "Notas", "append_note",
     "Pendiente confirmación cliente (J4): clarificar si la coincidencia con la etiqueta del CPO (elemento C) es requisito firme o aspiracional.",
     "Anotada pregunta del revisor sobre elemento C."),

    # 1.2.1 — "y" should be "o" between plan nacional and UNSDCF
    ("1.2.1", "Si", "replace",
     "Identifica explícitamente el plan nacional de desarrollo O el UNSDCF (u homólogo) vigentes, Y articula cómo el proyecto encaja en al menos uno de ellos.",
     "Cambiado «y» → «o» entre plan nacional y UNSDCF (basta con uno)."),

    # 1.3.3 — B/C/D similares: consolidar en un solo elemento de análisis de género
    ("1.3.3", "Elem", "replace",
     "A. Identifica al grupo de población afectado por nombre y magnitud.\n"
     "B. Análisis de género del contexto (cubre al menos uno de: roles, división sexual del trabajo, oportunidades y limitaciones diferenciadas por sexo).\n"
     "C. Datos desagregados por sexo cuando estén disponibles.",
     "Consolidados los elementos B/C/D previos en un único elemento B (análisis de género del contexto)."),
    ("1.3.3", "Si", "replace",
     "Cumple A, B (análisis de género del contexto) y C (datos desagregados cuando estén disponibles). "
     "Aplicar el filtro DEDICADO vs MARCO.",
     "Sí ajustado a la nueva estructura A/B/C."),
    ("1.3.3", "Par", "replace",
     "Cumple A y B pero B es genérico (no específico del contexto del proyecto), O falta C donde sí hay datos disponibles.",
     "Parcial ajustado a la nueva estructura."),

    # 1.4.2 — too hard to judge pertinencia from a list; remove that requirement
    ("1.4.2", "Par", "replace",
     "Menciona presencia OIT y enumera proyectos pasados/en curso, sin necesidad de juicio explícito de pertinencia.",
     "Removido el juicio de «pertinencia» del Parcial (queda solo en Sí)."),

    # 1.4.3 — B podría ser opcional
    ("1.4.3", "Elem", "replace",
     "A. Cita evaluaciones específicas (proyecto, programa-país, estrategia).\n"
     "B. (Opcional) Extrae lecciones aprendidas concretas.\n"
     "C. Vincula esas lecciones con decisiones de diseño del proyecto actual.",
     "B marcado como opcional."),
    ("1.4.3", "Si", "replace",
     "Cumple A y C con vinculación visible (la lección modificó algo en el diseño). B (extracción de lecciones concretas) eleva la calificación pero no es requisito.",
     "Sí no requiere B."),

    # 1.5.1 — D y E opcionales, dependen de naturaleza del proyecto
    ("1.5.1", "Elem", "replace",
     "A. Sub-objetivo, resultado o producto cuyo título nombra discapacidad.\n"
     "B. Indicador desagregado por discapacidad o que la mide específicamente.\n"
     "C. Actividad cuyo propósito principal es discapacidad.\n"
     "D. (Opcional según naturaleza del proyecto) Partida presupuestaria para discapacidad.\n"
     "E. (Opcional según naturaleza del proyecto) Meta cuantificable relativa a discapacidad.",
     "D y E marcados como opcionales según la naturaleza del proyecto."),
    ("1.5.1", "Si", "replace",
     "Cumple al menos 2 de A/B/C. D y E elevan la calificación pero no son obligatorios para Sí cuando la naturaleza del proyecto no los requiere. "
     "Aplicar el filtro DEDICADO vs MARCO.",
     "Sí ajustado a 2/3 de los elementos obligatorios."),

    # 1.5.3 — el "Sí" debe estar más formulado en Elementos a verificar
    ("1.5.3", "Elem", "replace",
     "A. Cita observaciones/solicitudes directas del CEACR sobre el país.\n"
     "B. Cita conclusiones del Comité de Aplicación de Normas o del Comité de Libertad Sindical cuando aplica.\n"
     "C. Esas observaciones informan la justificación o la estrategia del proyecto (vínculo explícito, no decorativo).",
     "Elementos reformulados para soportar la lógica del Sí (cliente pidió mover lógica acá)."),
    ("1.5.3", "Si", "replace",
     "Cumple al menos un elemento de supervisión (A o B) Y el vínculo C con la justificación/estrategia del proyecto.",
     "Sí reformulado: basta A o B + vínculo C."),

    # 1.5.4 — elementos demasiado exigentes
    ("1.5.4", "Si", "replace",
     "Cumple A (compromiso explícito NIT) más al menos 2 de B/C/D; E aplica solo cuando hay terceros y, cuando aplica, es obligatorio.",
     "Sí relajado: A + 2 de B/C/D (antes exigía A+B+C+D+E)."),

    # 1.5.5 — profundidad del análisis ambiental según naturaleza
    ("1.5.5", "Notas", "append_note",
     "Pendiente cliente (I21): definir profundidad mínima del análisis de impacto ambiental según naturaleza del proyecto (normativo vs infraestructura vs mixto).",
     "Anotada consulta pendiente sobre profundidad ambiental."),

    # 1.5.6 — añadir elemento A al Sí
    ("1.5.6", "Si", "replace",
     "Cumple A (distingue entre enfoque sensible / responsivo / transformador), B (articula cómo el proyecto cuestiona normas o relaciones de poder) y C (acciones DEDICADAS a transformar relaciones), idealmente con D. "
     "Aplicar el filtro DEDICADO vs MARCO.",
     "Añadido elemento A al Sí (estaba ausente)."),

    # 2.2.1 — C es exigencia muy alta
    ("2.2.1", "Elem", "replace",
     "A. Describe consultas realizadas (cuándo, con quién, sobre qué).\n"
     "B. Lista compromisos concretos asumidos por socios.\n"
     "C. (Opcional) Esos compromisos son operativos (tiempo, recursos, decisiones) — no solo declarativos.",
     "C marcado como opcional."),
    ("2.2.1", "Si", "replace",
     "Cumple A y B. C (compromisos operativos verificables) eleva la calificación pero no es requisito.",
     "Sí no requiere C."),

    # 2.2.2 — verificar exige cartas adjuntas
    ("2.2.2", "Notas", "append_note",
     "Operativo (I27): la verificación robusta de aceptación requiere cartas de intención / MoU / actas adjuntas. La app debe permitir cargar anexos y referenciarlos.",
     "Anotada necesidad operativa de adjuntar cartas."),

    # 2.4.1 — analizar cartas de compromiso
    ("2.4.1", "Notas", "append_note",
     "Operativo (I31): la evidencia robusta requiere cartas de compromiso/acuerdo adjuntas — la app debe soportar carga y análisis de anexos.",
     "Anotada necesidad de cartas adjuntas."),

    # 2.4.2 — quizá basta plan de sostenibilidad
    ("2.4.2", "Si", "replace",
     "La inclusión de un plan de sostenibilidad EXPLÍCITO es suficiente para Sí. Cumplir múltiples elementos (institucional, financiero, gobernanza, cronograma) eleva la calificación pero no es requisito.",
     "Sí relajado: plan explícito basta (antes exigía A + C + B/D)."),

    # 2.4.3 — la fase de prueba debe estar definida en el documento
    ("2.4.3", "Apli", "replace",
     "Condicional: aplica si la propuesta es piloto o tiene fase de prueba — la fase debe estar EXPLÍCITAMENTE definida en el documento analizado.",
     "Aplicabilidad clarificada: definición explícita requerida."),

    # 2.4.5 — insertos en plan de sostenibilidad; multi-archivo
    ("2.4.5", "Notas", "append_note",
     "Operativo (I35): los planes/actividades pueden estar insertados en un plan de sostenibilidad anexo y no en el PRODOC. La app debe permitir cargar múltiples archivos por propuesta.",
     "Anotada necesidad de soporte multi-archivo."),

    # 3.1.2 — al menos uno (subjetividad)
    ("3.1.2", "Si", "replace",
     "Cumple al menos uno de A (cobertura), B (suficiencia) o C (vínculo resultados→impacto). La verificación tiene componente subjetivo; basta con que el revisor identifique trazabilidad clara en una de las tres dimensiones.",
     "Sí bajado a 1/3 elementos (cliente reconoció subjetividad)."),

    # 3.3.1 — A depende de naturaleza
    ("3.3.1", "Si", "replace",
     "Si la inclusión de género es EXPLÍCITA en el proyecto: cumple A (resultado/producto que nombra género) Y al menos B o C. "
     "Si la inclusión es implícita: basta con B (indicador desagregado por sexo) o C (meta cuantificable de género). "
     "Aplicar el filtro DEDICADO vs MARCO.",
     "Sí condicional según si la inclusión es explícita o implícita."),

    # 3.3.2 — mismo razonamiento para discapacidad
    ("3.3.2", "Si", "replace",
     "Si la inclusión de discapacidad es EXPLÍCITA en el proyecto: cumple A (resultado/producto que nombra discapacidad) Y al menos B o C. "
     "Si la inclusión es implícita: basta con B (indicador) o C (meta). "
     "Aplicar el filtro DEDICADO vs MARCO.",
     "Sí condicional según si la inclusión es explícita o implícita."),

    # 3.4.4 — B no es necesario
    ("3.4.4", "Si", "replace",
     "Cumple A (identifica riesgos de incumplimiento NIT) y C (plan de mitigación específico). B (vínculo con observaciones del CEACR) eleva la calificación pero no es obligatorio si el análisis de cumplimiento NIT es sustantivo.",
     "B (vínculo CEACR) ya no es obligatorio para Sí."),

    # 3.4.5 — operativo: Streamlit debe procesar anexos
    ("3.4.5", "Notas", "append_note",
     "Operativo (J50): Streamlit debe identificar y procesar el registro de riesgos adjunto. La evaluación correcta depende de que la app lea el anexo, no solo el cuerpo del PRODOC.",
     "Anotada necesidad operativa de procesamiento de anexos."),

    # 3.5.1 — D opcional
    ("3.5.1", "Elem", "replace",
     "A. Sistema de recopilación descrito (qué datos, con qué frecuencia, por quién).\n"
     "B. Justificación de recursos asignados a S&E.\n"
     "C. Revisión de evaluabilidad cuando aplica (proyecto > umbral).\n"
     "D. (Opcional) Plan de aprendizaje del proyecto.",
     "D marcado como opcional."),
    ("3.5.1", "Si", "replace",
     "Cumple A y B; cumple C cuando aplica. D (plan de aprendizaje) es opcional.",
     "Sí no requiere D."),

    # 3.5.4 — donante puede imponer lineamientos
    ("3.5.4", "Notas", "append_note",
     "Pendiente cliente (G56/J56): cuando el donante impone lineamientos de M&E (p. ej. UE, BM), esos lineamientos prevalecen sobre la rúbrica. Revisar elementos A–F con este criterio adicional.",
     "Anotada prevalencia de requisitos del donante sobre la rúbrica."),

    # 3.5.5 — quitar C
    ("3.5.5", "Elem", "replace",
     "A. Partida presupuestaria de evaluación separada.\n"
     "B. Monto ≥ ~2% del presupuesto total (o justificación si difiere).",
     "Elemento C eliminado por sugerencia del cliente."),
    ("3.5.5", "Si", "replace",
     "Cumple A y B.",
     "Sí ajustado a A+B."),
    ("3.5.5", "Par", "replace",
     "Cumple A pero el monto es menor al 2% sin justificación; o no hay partida separada pero se contempla un monto identificable.",
     "Parcial ajustado (sin referencia a C)."),

    # 3.6.1 — muy complejo para evaluación AI
    ("3.6.1", "Notas", "append_note",
     "Cliente (I58) reportó alta complejidad para evaluación automatizada. Marcar como criterio de revisión humana prioritaria — la calificación del modelo debe leerse como hipótesis, no veredicto.",
     "Marcado como criterio de alta incertidumbre para IA."),

    # 3.6.2 — C demasiado requerimiento
    ("3.6.2", "Elem", "replace",
     "A. Presupuesto desglosado por actividad o producto.\n"
     "B. Coherencia entre actividades del marco lógico y partidas presupuestarias.\n"
     "C. (Opcional) Costos unitarios o cálculos visibles.",
     "C marcado como opcional."),
    ("3.6.2", "Si", "replace",
     "Cumple A y B. C (costos unitarios visibles) eleva la calificación pero no es requisito.",
     "Sí no requiere C."),

    # 4.1.3 — IRIS / backstopping
    ("4.1.3", "Notas", "append_note",
     "Pendiente verificación (I65): confirmar si IRIS expone el presupuesto de backstopping de forma identificable. Si no, la evaluación de este criterio queda parcialmente fuera del alcance automatizable.",
     "Anotada incertidumbre sobre identificación de backstopping en IRIS."),

    # 4.1.4 — consultar ADMIN
    ("4.1.4", "Notas", "append_note",
     "Pendiente consulta ADMIN (I66): determinar qué elementos (dotación, adquisiciones, sistemas financieros, autoridades) son obligatorios y cuáles son opcionales antes de cerrar la rúbrica.",
     "Bloqueado a consulta con ADMIN."),

    # 4.1.5 — confirmar con ADMIN
    ("4.1.5", "Notas", "append_note",
     "Pendiente consulta ADMIN (I67): confirmar si las cláusulas de trabajo decente (A) y empleo justo (B) son elementos estándar en los contratos OIT con terceros.",
     "Bloqueado a consulta con ADMIN sobre cláusulas estándar."),
]


def main():
    shutil.copy(SRC, DST)
    wb = load_workbook(DST)
    ws = wb["Rúbrica Tab 1"]

    # Build ID → row map
    for r in range(3, ws.max_row + 1):
        v = ws.cell(row=r, column=COL["ID"]).value
        if v:
            ID_TO_ROW[str(v)] = r

    # Strip all threaded comments (resolved in-line)
    stripped = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.comment:
                cell.comment = None
                stripped += 1
    print(f"Comentarios hilvanados eliminados: {stripped}")

    # Highlight + edit
    edit_fill = PatternFill("solid", fgColor="FFD89B")  # warm orange
    note_fill = PatternFill("solid", fgColor="FFF2CC")  # soft yellow for Notas
    cambios_by_row: dict[int, list[str]] = {}

    for crit_id, col_key, action, new_val, summary in EDITS:
        row = ID_TO_ROW.get(crit_id)
        if row is None:
            print(f"  ⚠ ID {crit_id} no encontrado — edición omitida")
            continue
        col = COL[col_key]
        cell = ws.cell(row=row, column=col)

        if action == "replace":
            cell.value = new_val
            cell.fill = edit_fill
        elif action == "append_note":
            current = (cell.value or "").strip()
            sep = "\n\n" if current else ""
            cell.value = f"{current}{sep}▸ {new_val}"
            cell.fill = note_fill

        cambios_by_row.setdefault(row, []).append(f"• [{col_key}] {summary}")

    # New "Cambios v2" column at column 18 (R)
    header_cell = ws.cell(row=2, column=CAMBIOS_COL)
    header_cell.value = "Cambios v2 (revisión cliente)"
    header_cell.font = Font(bold=True, color="FFFFFF", size=11)
    header_cell.fill = PatternFill("solid", fgColor="C0504D")
    header_cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    border = Border(
        left=Side(style="thin", color="888888"),
        right=Side(style="thin", color="888888"),
        top=Side(style="thin", color="888888"),
        bottom=Side(style="thin", color="888888"),
    )
    header_cell.border = border

    # Set width for new column
    ws.column_dimensions[get_column_letter(CAMBIOS_COL)].width = 50

    # Write change summaries
    summary_align = Alignment(wrap_text=True, vertical="top", horizontal="left")
    for row, changes in cambios_by_row.items():
        cell = ws.cell(row=row, column=CAMBIOS_COL, value="\n".join(changes))
        cell.alignment = summary_align
        cell.fill = PatternFill("solid", fgColor="FCE4D6")
        cell.border = border
        cell.font = Font(size=10)

    # Update instructions row (A1) to reflect v2
    ws["A1"] = (
        "RÚBRICA Tab 1 — v2 (revisión cliente sintetizada). "
        "Celdas resaltadas en NARANJA = contenido editado por revisión del cliente. "
        "Celdas resaltadas en AMARILLO = notas operativas/pendientes anexadas. "
        "Columna «Cambios v2» = resumen por fila de qué cambió. "
        "Comentarios hilvanados originales eliminados (resueltos en línea o anotados en Notas). "
        "Los 228 placeholders [LLENAR] siguen sin completar — sigue siendo bloqueante para calibración."
    )

    # Merge A1 across new column count
    try:
        ws.unmerge_cells(start_row=1, start_column=1, end_row=1, end_column=17)
    except Exception:
        pass
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=CAMBIOS_COL)

    wb.save(DST)
    print(f"\nGuardado: {DST}")
    print(f"Filas editadas: {len(cambios_by_row)}")
    print(f"Total ediciones aplicadas: {sum(len(v) for v in cambios_by_row.values())}")


if __name__ == "__main__":
    main()
