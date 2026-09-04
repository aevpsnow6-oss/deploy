"""Write the English display columns into the v3 rubric xlsx.

The Spanish columns stay authoritative: evaluation always reads them. The
«… (EN)» columns are display-only, and any cell left empty falls back to
Spanish at render time. Re-runnable: it overwrites only the (EN) columns.

    python rubrica_en/apply.py
"""
import json
import pathlib
import sys

import openpyxl

HERE = pathlib.Path(__file__).parent
RUBRIC = HERE.parent / "Rubrica_Tab1_Detallada_Full_v3.xlsx"
SHEET = "Rúbrica Tab 1"
HEADER_ROW = 2  # 1-indexed en openpyxl

# fichero json -> columna española que traduce. Los ficheros que faltan se
# omiten sin error, para poder traducir por partes.
SOURCES = {
    "criterios.json": "Criterio a evaluar",
    "tests_si.json": "Rúbrica — Sí",
    "elementos.json": "Elementos a verificar",
    "tests_parcial.json": "Rúbrica — Parcial",
    "tests_no.json": "Rúbrica — No",
    "tests_na.json": "Rúbrica — No aplica",
    "anclas.json": "Anclas verificables (v3)",
}
# éste se indexa por TEXTO, no por ID: la misma pregunta se repite en varias filas
BY_TEXT = {"preguntas_orientadoras.json": "Pregunta orientadora (CONTEXTO — no evaluar)"}


def load(name):
    path = HERE / name
    return json.loads(path.read_text(encoding="utf-8")) if path.exists() else None


def column_index(ws, name):
    for cell in ws[HEADER_ROW]:
        if cell.value == name:
            return cell.column
    return None


def ensure_column(ws, name):
    """Return the index of «name», appending the column if absent."""
    existing = column_index(ws, name)
    if existing:
        return existing
    idx = ws.max_column + 1
    ws.cell(HEADER_ROW, idx).value = name
    return idx


def main():
    wb = openpyxl.load_workbook(RUBRIC)
    ws = wb[SHEET]
    col_id = column_index(ws, "ID")
    rows = {
        str(ws.cell(r, col_id).value).strip(): r
        for r in range(HEADER_ROW + 1, ws.max_row + 1)
        if ws.cell(r, col_id).value is not None
    }

    written = missing = 0
    for fname, es_col in SOURCES.items():
        data = load(fname)
        if data is None:
            print(f"  (omitido, aún sin traducir) {fname}")
            continue
        col = ensure_column(ws, es_col + " (EN)")
        for crit_id, row in rows.items():
            text = data.get(crit_id)
            if text:
                ws.cell(row, col).value = text
                written += 1
            else:
                missing += 1
        print(f"  {fname:24} -> «{es_col} (EN)»  ({len(data)} entradas)")

    for fname, es_col in BY_TEXT.items():
        data = load(fname)
        if data is None:
            print(f"  (omitido) {fname}")
            continue
        src = column_index(ws, es_col)
        col = ensure_column(ws, es_col + " (EN)")
        unknown = set()
        for crit_id, row in rows.items():
            key = str(ws.cell(row, src).value or "").strip()
            if key in data:
                ws.cell(row, col).value = data[key]
                written += 1
            elif key:
                unknown.add(key)
        print(f"  {fname:24} -> «{es_col} (EN)»  ({len(data)} entradas)")
        if unknown:
            print(f"    SIN TRADUCIR: {len(unknown)}")
            for u in sorted(unknown):
                print(f"      {u[:90]}")

    wb.save(RUBRIC)
    print(f"\nceldas escritas: {written}  |  sin traducción (fallback a español): {missing}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
