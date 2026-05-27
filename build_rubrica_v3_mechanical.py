"""Generate v3 of the Tab 1 rubric with maximally mechanical decision rules.

Goal: minimize the need for calibration examples by replacing subjective prose
with atomic tests (T1..Tn) and explicit Boolean decision rules.

Inputs:  Rubrica_Tab1_Detallada_Full_v2.xlsx (preserves client review edits)
Output:  Rubrica_Tab1_Detallada_Full_v3.xlsx with:
  - Sí / Parcial / No rewritten as TEST-list + DECISIÓN rule
  - New column "Anclas verificables" (concrete text patterns / codes / names to look for)
  - New column "Subjetividad residual" (Baja / Media / Alta) — flags where examples
    would still add most value if bootstrapped later
  - Example placeholder columns removed (replaced by tests + anchors)
"""

import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = "/Users/ageidv/ilo/deploy_3/Rubrica_Tab1_Detallada_Full_v2.xlsx"
DST = "/Users/ageidv/ilo/deploy_3/Rubrica_Tab1_Detallada_Full_v3.xlsx"

# Mechanical rubric content per criterion.
# Each entry rewrites Sí / Parcial / No, plus anclas + subjetividad_residual.
# Format convention:
#   Sí/Parcial/No = "TESTS:\n✓ T1: ...\n✓ T2: ...\n\nDECISIÓN: <boolean rule>"

R = {}

R["1.1.1"] = dict(
    si=("TESTS:\n"
        "T1: ¿Cita al menos un resultado/prioridad de P&B con código o nombre? (sí/no)\n"
        "T2: ¿Cita al menos un resultado DWCP con nombre del país? (sí/no)\n"
        "T3: ¿Cita al menos un CPO por código (formato letra+número, p.ej. ABC-101)? (sí/no)\n"
        "T4: ¿El texto vincula la intervención a T1/T2/T3 con verbo de contribución "
        "(«contribuye», «aporta», «alimenta», «se traduce en»)? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3 ∧ T4"),
    par=("DECISIÓN: (T1 ∧ T2 ∧ T3) ∧ ¬T4   (cita todo pero no articula contribución)\n"
         "    O    (exactamente 2 de T1/T2/T3) ∧ T4"),
    no=("DECISIÓN: (≤1 de T1/T2/T3 verdadero)   O   ningún verbo de contribución detectable"),
    anclas="P&B; DWCP; CPO; «contribuye a», «se alinea con», «aporta a»; códigos CPO (e.g. ABC-101)",
    subj="Baja",
)

R["1.1.2"] = dict(
    si=("TESTS:\n"
        "T1: ¿La propuesta nombra la etiqueta/marcador de discapacidad del CPO (texto literal: «etiqueta», «marcador», «label», «principal», «significativa», «limitada»)? (sí/no)\n"
        "T2: ¿Hay ≥1 elemento DEDICADO a discapacidad (sub-objetivo / indicador / actividad / presupuesto / meta)? (sí/no)\n"
        "T3: ¿El número de elementos DEDICADOS coincide con la etiqueta (principal→≥3 DEDICADOS; significativa→≥1; limitada→0–1)? (sí/no)\n\n"
        "FILTRO: solo cuentan DEDICADOS, no MARCO (listas con ≥3 grupos, «entre otros», lenguaje inclusivo).\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3"),
    par=("DECISIÓN: T1 ∧ T2 ∧ ¬T3   (reconoce etiqueta pero el nivel queda por debajo)"),
    no=("DECISIÓN: ¬T1   O   ¬T2 (solo MARCO, ningún DEDICADO)"),
    anclas="«etiqueta de discapacidad»; «marcador de discapacidad»; «discapacidad principal/significativa/limitada»; nombres de CPO con su marcador",
    subj="Media",
)

R["1.1.3"] = dict(
    si=("TESTS:\n"
        "T1: ¿La propuesta nombra el marcador de género del CPO (texto: «marcador 0/1/2/3», «GEM 0–3», «principal», «significativo»)? (sí/no)\n"
        "T2: ¿Hay ≥1 elemento DEDICADO a género (sub-objetivo / indicador desagregado / actividad / partida / meta)? (sí/no)\n"
        "T3: ¿El número de DEDICADOS coincide con el marcador (3→≥3 DEDICADOS; 2→≥2; 1→≥1)? (sí/no)\n\n"
        "FILTRO: DEDICADO vs MARCO igual que 1.1.2.\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3"),
    par=("DECISIÓN: T1 ∧ T2 ∧ ¬T3"),
    no=("DECISIÓN: ¬T1   O   marcador ≥2 ∧ ningún DEDICADO"),
    anclas="«marcador de género»; «GEM»; «principal/significativo/limitado»; código del CPO con marcador",
    subj="Media",
)

R["1.2.1"] = dict(
    si=("TESTS:\n"
        "T1: ¿Nombra el plan nacional de desarrollo del país por título (no «el plan nacional» genérico)? (sí/no)\n"
        "T2: ¿Nombra el UNSDCF / MANUD vigente del país? (sí/no)\n"
        "T3: ¿Articula el encaje del proyecto (verbo + objeto) con al menos uno de los marcos de T1/T2? (sí/no)\n\n"
        "DECISIÓN: (T1 ∨ T2) ∧ T3"),
    par=("DECISIÓN: (T1 ∨ T2) ∧ ¬T3   (nombra marco pero no articula encaje)"),
    no=("DECISIÓN: ¬T1 ∧ ¬T2   (no nombra ningún marco específico)"),
    anclas="«UNSDCF»; «MANUD»; nombre oficial del plan nacional; «Plan Nacional de Desarrollo»; «Estrategia Nacional»",
    subj="Baja",
)

R["1.2.2"] = dict(
    si=("TESTS:\n"
        "T1: ¿Cita outcomes específicos del UNSDCF (por número o título completo)? (sí/no)\n"
        "T2: ¿Identifica las áreas donde la OIT es agencia líder o convocante? (sí/no)\n"
        "T3: ¿Articula la contribución del proyecto a los outcomes de T1? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3"),
    par=("DECISIÓN: T1 ∧ ¬(T2 ∧ T3)"),
    no=("DECISIÓN: ¬T1   (solo coherencia genérica con UNSDCF)"),
    anclas="«UNSDCF Outcome», «Resultado X del UNSDCF», «OIT lidera», «OIT convoca»",
    subj="Baja",
)

R["1.2.3"] = dict(
    si=("TESTS:\n"
        "T1: ¿Cita ≥2 indicadores ODS por código (formato N.N.N como 8.5.2, 1.3.1)? (sí/no)\n"
        "T2: ¿Esos indicadores aparecen en el marco lógico o plan de M&E? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2"),
    par=("DECISIÓN: T1 ∧ ¬T2   (cita códigos pero no se integran al M&E)"),
    no=("DECISIÓN: ¬T1   (solo nombra ODS «5», «8» sin códigos de indicador)"),
    anclas="patrón regex \\d+\\.\\d+\\.\\d+ ; «indicador ODS»; «marcador SDG»",
    subj="Baja",
)

R["1.3.1"] = dict(
    si=("TESTS:\n"
        "T1: ¿Existe sección/anexo etiquetada «análisis de situación» o «descripción del problema»? (sí/no)\n"
        "T2: ¿Cita ≥2 fuentes verificables (estudios, datos oficiales, evaluaciones, censos)? (sí/no)\n"
        "T3: ¿Cuantifica la magnitud (cifra, porcentaje, número de personas afectadas)? (sí/no)\n"
        "T4: ¿Delimita alcance temporal, geográfico o poblacional? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ (T3 ∨ T4)"),
    par=("DECISIÓN: T1 ∧ (T2 ∨ T3 ∨ T4)   pero ¬ (T2 ∧ (T3 ∨ T4))"),
    no=("DECISIÓN: ¬T1   O   T1 ∧ ¬T2 ∧ ¬T3"),
    anclas="«análisis de situación»; «descripción del problema»; citaciones bibliográficas; cifras con %, USD o N=",
    subj="Baja",
)

R["1.3.2"] = dict(
    si=("TESTS:\n"
        "T1: ¿El texto distingue al menos 2 niveles causales (inmediatas, subyacentes, estructurales) o usa árbol de problemas? (sí/no)\n"
        "T2: ¿Las causas se respaldan en evidencia o consulta (cita / referencia)? (sí/no)\n"
        "T3: ¿La estrategia del proyecto se mapea explícitamente a causas (no solo a síntomas)? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T3"),
    par=("DECISIÓN: T1 ∧ ¬T3   O   ¬T1 ∧ T3"),
    no=("DECISIÓN: ¬T1 ∧ ¬T3"),
    anclas="«causas inmediatas/subyacentes/estructurales»; «árbol de problemas»; «raíz del problema»",
    subj="Media",
)

R["1.3.3"] = dict(
    si=("TESTS:\n"
        "T1: ¿Identifica al grupo de población por nombre y magnitud? (sí/no)\n"
        "T2: ¿Hay análisis de género del contexto (cubre al menos uno de: roles, división del trabajo, oportunidades/limitaciones diferenciadas)? (sí/no)\n"
        "T3: ¿Hay datos desagregados por sexo cuando están disponibles? (sí/no)\n\n"
        "FILTRO DEDICADO vs MARCO aplica al análisis de género (T2).\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3"),
    par=("DECISIÓN: T1 ∧ T2 ∧ ¬T3   (existe T2 pero faltan datos desagregados)\n"
         "    O    T1 ∧ T3 ∧ T2-genérico (T2 sin especificidad del contexto)"),
    no=("DECISIÓN: T1 ∧ ¬T2"),
    anclas="«análisis de género»; «roles de género»; «división sexual del trabajo»; datos con desglose «mujeres/hombres» o «M/H»",
    subj="Media",
)

R["1.3.4"] = dict(
    si=("TESTS:\n"
        "T1: ¿Existe mapeo de partes interesadas (tabla/sección dedicada)? (sí/no)\n"
        "T2: ¿Cada parte clave tiene declarados intereses Y limitaciones? (sí/no)\n"
        "T3: ¿Incluye organizaciones de mujeres por nombre? (sí/no)\n"
        "T4: Si el proyecto afecta a personas con discapacidad → ¿incluye OPD (organizaciones de personas con discapacidad)? (sí/no/NA)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3 ∧ (T4 ∨ NA)"),
    par=("DECISIÓN: T1 ∧ T2 ∧ ¬(T3 ∧ T4)"),
    no=("DECISIÓN: ¬T1   O   ¬T2 (solo lista, sin análisis)"),
    anclas="«mapeo de partes interesadas»; «análisis de stakeholders»; nombres de OPD, organizaciones de mujeres",
    subj="Media",
)

R["1.3.5"] = dict(
    si=("TESTS:\n"
        "T1: ¿Hay plan de consulta con metodología documentada? (sí/no)\n"
        "T2: ¿Identifica medidas concretas de accesibilidad/equidad (interpretación, horarios, formatos, idiomas locales)? (sí/no)\n"
        "T3: ¿Hay análisis de potenciales efectos discriminatorios y mitigación? (sí/no)\n"
        "T4: Si afecta a pueblos indígenas → ¿incluye proceso CLPI (Consentimiento Libre, Previo e Informado)? (sí/no/NA)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3 ∧ (T4 ∨ NA)"),
    par=("DECISIÓN: T1 ∧ T2 ∧ ¬T3   O   T1 ∧ ¬T2"),
    no=("DECISIÓN: ¬T1   O   consulta solo con autoridades/socios sin grupos marginados"),
    anclas="«CLPI»; «consentimiento libre, previo e informado»; «interpretación»; «lenguaje accesible»; nombres de grupos marginados",
    subj="Media",
)

R["1.4.1"] = dict(
    si=("TESTS:\n"
        "T1: ¿Cita explícita de identidad normativa OIT (convenios, recomendaciones, normas)? (sí/no)\n"
        "T2: ¿Cita explícita del tripartismo (gobierno + empleadores + trabajadores nombrados)? (sí/no)\n"
        "T3: ¿Argumento de valor añadido específico al proyecto (no plantilla institucional)? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3"),
    par=("DECISIÓN: T1 ∧ T2 ∧ ¬T3"),
    no=("DECISIÓN: ¬T1 ∨ ¬T2"),
    anclas="«normativo», «tripartito», «empleadores y trabajadores», «constituyentes OIT»; convenios citados por número",
    subj="Media",
)

R["1.4.2"] = dict(
    si=("TESTS:\n"
        "T1: ¿Describe presencia OIT en el país (oficina, equipo, antigüedad)? (sí/no)\n"
        "T2: ¿Enumera ≥2 proyectos pasados o en curso con título o código? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2"),
    par=("DECISIÓN: T1 ∨ T2 (cumple solo uno)"),
    no=("DECISIÓN: ¬T1 ∧ ¬T2"),
    anclas="«Oficina de la OIT en»; nombres de proyectos pasados; códigos de proyecto",
    subj="Baja",
)

R["1.4.3"] = dict(
    si=("TESTS:\n"
        "T1: ¿Cita ≥1 evaluación específica (título, año, proyecto)? (sí/no)\n"
        "T2: ¿Vincula al menos una lección con una decisión visible del diseño actual? (sí/no)\n"
        "T3 (opcional): ¿Extrae lecciones concretas en lista o prosa? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2"),
    par=("DECISIÓN: T1 ∧ ¬T2   (cita pero no vincula)"),
    no=("DECISIÓN: ¬T1   O   solo «se han considerado lecciones» sin cita"),
    anclas="«evaluación de», «lecciones aprendidas»; años; URLs ievaldiscovery",
    subj="Media",
)

R["1.5.1"] = dict(
    si=("TESTS (filtro DEDICADO vs MARCO):\n"
        "T1: ¿Sub-objetivo / resultado / producto que NOMBRA discapacidad? (sí/no)\n"
        "T2: ¿Indicador desagregado por discapacidad o específico? (sí/no)\n"
        "T3: ¿Actividad cuyo propósito principal es discapacidad? (sí/no)\n"
        "T4 (opcional): ¿Partida presupuestaria para discapacidad? (sí/no)\n"
        "T5 (opcional): ¿Meta cuantificable relativa a discapacidad? (sí/no)\n\n"
        "DECISIÓN: (#verdaderos de T1/T2/T3) ≥ 2"),
    par=("DECISIÓN: (#verdaderos de T1/T2/T3) = 1   O   solo T4/T5 sin T1/T2/T3"),
    no=("DECISIÓN: (#verdaderos de T1/T2/T3/T4/T5) = 0"),
    anclas="«personas con discapacidad»; «PcD»; «accesibilidad»; «desagregado por discapacidad»",
    subj="Alta",
)

R["1.5.2"] = dict(
    si=("TESTS:\n"
        "T1: ¿Cita convenios/recomendaciones OIT por número (C087, C098, C111, C190…)? (sí/no)\n"
        "T2: ¿Esos instrumentos aparecen en estrategia, indicadores u objetivos (no solo en antecedentes)? (sí/no)\n"
        "T3: ¿Incluye acciones de promoción de ratificación / aplicación / conocimiento? (sí/no)\n"
        "T4: ¿Referencias a NORMLEX u otra fuente jurídica OIT? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ (T3 ∨ T4)"),
    par=("DECISIÓN: T1 ∧ ¬T2   (cita en antecedentes pero no integra a estrategia)"),
    no=("DECISIÓN: ¬T1   (no cita convenios por número)"),
    anclas="«C087», «C098», «C190»; «Convenio núm.»; «NORMLEX»; URL normlex.ilo.org",
    subj="Baja",
)

R["1.5.3"] = dict(
    si=("TESTS:\n"
        "T1: ¿Cita observaciones del CEACR sobre el país? (sí/no)\n"
        "T2: ¿Cita conclusiones del Comité de Aplicación de Normas o Comité de Libertad Sindical? (sí/no)\n"
        "T3: ¿Las observaciones citadas informan justificación o estrategia (vínculo explícito)? (sí/no)\n\n"
        "DECISIÓN: (T1 ∨ T2) ∧ T3"),
    par=("DECISIÓN: (T1 ∨ T2) ∧ ¬T3   (cita observaciones sin vínculo a estrategia)"),
    no=("DECISIÓN: ¬T1 ∧ ¬T2"),
    anclas="«CEACR»; «Comisión de Expertos»; «Comité de Libertad Sindical»; «Comité de Aplicación de Normas»",
    subj="Media",
)

R["1.5.4"] = dict(
    si=("TESTS:\n"
        "T1: ¿Compromiso explícito de cumplimiento NIT en el proyecto? (sí/no)\n"
        "T2: ¿Salarios justos especificados (referencia legal o sectorial)? (sí/no)\n"
        "T3: ¿Condiciones SST aplicables al personal y a contratistas? (sí/no)\n"
        "T4: ¿Mecanismo de quejas accesible (canal + idiomas + confidencialidad)? (sí/no)\n"
        "T5 (cond): si hay terceros → ¿cláusulas de cumplimiento en contratos? (sí/no/NA)\n\n"
        "DECISIÓN: T1 ∧ (#verdaderos de T2/T3/T4) ≥ 2 ∧ (T5 ∨ NA)"),
    par=("DECISIÓN: T1 ∧ exactamente 1 de T2/T3/T4   O   T5 ausente cuando hay terceros"),
    no=("DECISIÓN: ¬T1   O   solo declaración general sin elementos operativos"),
    anclas="«salarios», «SST», «seguridad y salud en el trabajo», «mecanismo de queja», «código de conducta»",
    subj="Media",
)

R["1.5.5"] = dict(
    si=("TESTS:\n"
        "T1: ¿Análisis de impactos ambientales potenciales del proyecto (no genérico)? (sí/no)\n"
        "T2: ¿Medidas de mitigación específicas para impactos identificados? (sí/no)\n"
        "T3: ¿Considera biodiversidad O comunidades afectadas? (sí/no)\n"
        "T4: ¿Prácticas sostenibles en ejecución (materiales / energía / residuos)? (sí/no)\n"
        "T5 (cond): si hay infraestructura/construcción → ¿medidas específicas de seguridad ambiental? (sí/no/NA)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ (#verdaderos de T3/T4) ≥ 1 ∧ (T5 ∨ NA)"),
    par=("DECISIÓN: T1 ∧ ¬T2   O   T1 ∧ T2 ∧ ¬T5 (cuando aplica T5)"),
    no=("DECISIÓN: ¬T1   (solo menciones decorativas)"),
    anclas="«impacto ambiental»; «biodiversidad»; «residuos»; «energía»; «sostenibilidad ambiental»",
    subj="Media",
)

R["1.5.6"] = dict(
    si=("TESTS:\n"
        "T1: ¿Distingue tipo de enfoque (sensible / responsivo / transformador)? (sí/no)\n"
        "T2: ¿Articula cómo el proyecto cuestiona normas o relaciones de poder? (sí/no)\n"
        "T3: ¿Acciones DEDICADAS a transformar relaciones (no solo a incluir mujeres)? (sí/no)\n"
        "T4 (opcional): ¿Indicadores miden cambios en relaciones (no solo participación numérica)? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3"),
    par=("DECISIÓN: T1 ∧ ¬(T2 ∧ T3)   (declara enfoque pero acciones siguen siendo numéricas)"),
    no=("DECISIÓN: ¬T1   O   solo «transversaliza género» sin contenido"),
    anclas="«enfoque transformador», «relaciones de poder», «normas de género», «Guía práctica n.º 15»",
    subj="Alta",
)

# ---------------- Sección 2 ----------------
R["2.1.1"] = dict(
    si=("TESTS:\n"
        "T1: ¿Describe el proceso (qué, cuándo, con quién) de consultas en diseño? (sí/no)\n"
        "T2: ¿Lista los constituyentes consultados (gobierno + empleadores + trabajadores nombrados)? (sí/no)\n"
        "T3: ¿Define rol en seguimiento del proyecto? (sí/no)\n"
        "T4: ¿Define rol en ejecución? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3 ∧ T4"),
    par=("DECISIÓN: T1 ∧ T2 ∧ ¬(T3 ∧ T4)"),
    no=("DECISIÓN: ¬T1 ∨ ¬T2"),
    anclas="nombres de ministerios; centrales sindicales; cámaras empresariales; «en el seguimiento», «en la ejecución»",
    subj="Baja",
)

R["2.1.2"] = dict(
    si=("TESTS:\n"
        "T1: Identificación nominal de partes clave (no genérico). (sí/no)\n"
        "T2: Vínculo de cada parte con beneficiarios finales. (sí/no)\n"
        "T3: Rol respecto al problema (causa/afectado/mitigador). (sí/no)\n"
        "T4: Rol respecto a la solución (ejecutor/validador/beneficiario). (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3 ∧ T4 para TODAS las partes clave"),
    par=("DECISIÓN: T1 ∧ (T2 ∨ T3 ∨ T4) pero no los tres para todas"),
    no=("DECISIÓN: ¬T1   (solo lista sin análisis)"),
    anclas="tabla de stakeholders; columnas «rol», «interés», «vínculo»",
    subj="Media",
)

R["2.1.3"] = dict(
    si=("TESTS:\n"
        "T1: ¿Mención explícita de consulta con ACT/EMP? (sí/no)\n"
        "T2: ¿Mención explícita de consulta con ACTRAV? (sí/no)\n"
        "T3: ¿Evidencia de integración (decisión, ajuste, sección citada)? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3"),
    par=("DECISIÓN: (T1 ∨ T2) ∧ ¬T3   O   T1 ∧ T2 ∧ ¬T3"),
    no=("DECISIÓN: ¬T1 ∧ ¬T2"),
    anclas="«ACT/EMP»; «ACTRAV»; «especialistas consultados»; nombres concretos",
    subj="Baja",
)

R["2.2.1"] = dict(
    si=("TESTS:\n"
        "T1: ¿Describe consultas realizadas (cuándo, con quién, sobre qué)? (sí/no)\n"
        "T2: ¿Lista compromisos concretos asumidos por socios? (sí/no)\n"
        "T3 (opcional): ¿Los compromisos son operativos (tiempo / recursos / decisiones)? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2"),
    par=("DECISIÓN: T1 ∧ ¬T2   O   solo afirmaciones de apropiación sin compromisos"),
    no=("DECISIÓN: ¬T1"),
    anclas="«acta de reunión», «MoU», «carta de intención»; fechas de consulta",
    subj="Media",
)

R["2.2.2"] = dict(
    si=("TESTS:\n"
        "T1: ¿Constancia de aceptación de objetivos por cada socio clave? (sí/no)\n"
        "T2: ¿Aceptación del marco de desempeño / indicadores? (sí/no)\n"
        "T3: ¿Aceptación de obligaciones y responsabilidades? (sí/no)\n"
        "T4: ¿Documento de respaldo referenciado (MoU, acta, carta)? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3 ∧ T4"),
    par=("DECISIÓN: (T1 ∨ T2 ∨ T3) ∧ ¬T4   (afirmación sin respaldo documental)"),
    no=("DECISIÓN: ¬T1 ∧ ¬T2 ∧ ¬T3"),
    anclas="«MoU», «carta de intención», «acuerdo de cooperación»; referencias a anexos",
    subj="Media",
)

R["2.2.3"] = dict(
    si=("TESTS:\n"
        "T1: ¿Estrategia explícita de apropiación continua (no solo eventos puntuales)? (sí/no)\n"
        "T2: ¿Mecanismos de gobernanza compartida (comités, grupos directivos)? (sí/no)\n"
        "T3: ¿Vinculación con la estrategia de sostenibilidad post-proyecto? (sí/no)\n"
        "T4: ¿Estrategia diferenciada por tipo de actor (no uniforme)? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3 ∧ T4"),
    par=("DECISIÓN: T1 ∧ T2 ∧ (T3 ∨ T4) pero no ambos"),
    no=("DECISIÓN: ¬T1"),
    anclas="«comité directivo», «grupo de coordinación», «steering committee»",
    subj="Media",
)

R["2.3.1"] = dict(
    si=("TESTS:\n"
        "T1: ¿Evaluación de capacidades realizada O planificada con metodología nombrada? (sí/no)\n"
        "T2: ¿Conclusiones explícitas (brechas vs capacidades existentes)? (sí/no)\n"
        "T3: ¿Esas conclusiones aparecen integradas al diseño del proyecto? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3"),
    par=("DECISIÓN: T1 ∧ T2 ∧ ¬T3"),
    no=("DECISIÓN: ¬T1   O   solo enunciado vago «se evaluarán las capacidades»"),
    anclas="«evaluación de capacidades», «capacity assessment», «análisis de brechas»",
    subj="Baja",
)

R["2.3.2"] = dict(
    si=("TESTS:\n"
        "T1: ¿Capacidades individuales (técnicas)? (sí/no)\n"
        "T2: ¿Capacidades organizativas (sistemas, procesos)? (sí/no)\n"
        "T3: ¿Entorno propicio (marco legal, políticas, financiación)? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3"),
    par=("DECISIÓN: 2 de T1/T2/T3   O   los 3 pero énfasis casi exclusivo en T1 (capacitación individual)"),
    no=("DECISIÓN: solo T1   (reduce desarrollo de capacidades a talleres)"),
    anclas="«capacidades individuales / organizativas / entorno propicio»; «sistémico»; «político»",
    subj="Media",
)

R["2.4.1"] = dict(
    si=("TESTS:\n"
        "T1: ¿Evidencia explícita de interés de constituyentes en post-proyecto? (sí/no)\n"
        "T2: ¿Esa evidencia es concreta (carta, MoU, asignación de personal/presupuesto futuro)? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2"),
    par=("DECISIÓN: T1 ∧ ¬T2   (declaración sin operativización)"),
    no=("DECISIÓN: ¬T1"),
    anclas="«post-proyecto», «sostenibilidad», cartas de compromiso, asignaciones presupuestarias futuras",
    subj="Media",
)

R["2.4.2"] = dict(
    si=("TESTS:\n"
        "T1: ¿Plan de sostenibilidad EXPLÍCITO (sección, anexo o referencia clara)? (sí/no)\n"
        "T2 (suma calidad): mecanismo institucional. (sí/no)\n"
        "T3 (suma calidad): mecanismo financiero. (sí/no)\n"
        "T4 (suma calidad): mecanismo de gobernanza. (sí/no)\n"
        "T5 (suma calidad): cronograma de transición. (sí/no)\n\n"
        "DECISIÓN: T1   (plan explícito basta; T2–T5 elevan la nota)"),
    par=("DECISIÓN: ¬T1 ∧ alguno de T2/T3/T4/T5   (elementos sueltos sin plan integrado)"),
    no=("DECISIÓN: ¬T1 ∧ ningún elemento operativo"),
    anclas="«plan de sostenibilidad»; «sustainability plan»; «mantenimiento de resultados»",
    subj="Baja",
)

R["2.4.3"] = dict(
    si=("TESTS:\n"
        "T1: ¿El documento define EXPLÍCITAMENTE la fase de prueba/piloto? (sí/no/NA)\n"
        "T2: ¿Identifica condiciones previas necesarias para escalar? (sí/no)\n"
        "T3: ¿Define criterios de éxito objetivos? (sí/no)\n"
        "T4: ¿Plan de medición/demostración del éxito? (sí/no)\n"
        "T5: ¿Plan de transición a fase de escala? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3 ∧ T4 ∧ T5"),
    par=("DECISIÓN: T1 ∧ T2 ∧ T3 ∧ ¬(T4 ∧ T5)"),
    no=("DECISIÓN: T1 ∧ ¬T2 ∧ ¬T3"),
    na="DECISIÓN: ¬T1 (no se define fase de prueba) → N/A explícito",
    anclas="«piloto», «fase de prueba», «proof of concept», «escalamiento»",
    subj="Baja",
)

R["2.4.4"] = dict(
    si=("TESTS:\n"
        "T1: ¿Fuente identificada de recursos humanos post-proyecto? (sí/no)\n"
        "T2: ¿Fuente identificada de recursos financieros post-proyecto? (sí/no)\n"
        "T3: ¿Argumento de plausibilidad con sustento (presupuesto sectorial existente, compromiso institucional)? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3"),
    par=("DECISIÓN: (T1 ∨ T2) ∧ ¬T3"),
    no=("DECISIÓN: ¬T1 ∧ ¬T2"),
    anclas="«recursos post-proyecto», «presupuesto sectorial», «línea presupuestaria»",
    subj="Media",
)

R["2.4.5"] = dict(
    si=("TESTS:\n"
        "T1: ¿Estrategia de salida EXPLÍCITA (sección o anexo)? (sí/no)\n"
        "T2: ¿Plan de transferencia de responsabilidades (qué, a quién, cuándo)? (sí/no)\n"
        "T3: ¿Acciones de desarrollo de capacidades específicas para la transferencia? (sí/no)\n"
        "T4: ¿Cronograma de salida con hitos? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3 ∧ T4"),
    par=("DECISIÓN: T1 ∧ T2 ∧ ¬(T3 ∧ T4)"),
    no=("DECISIÓN: ¬T1"),
    anclas="«estrategia de salida», «exit strategy», «transferencia de responsabilidades»",
    subj="Baja",
)

# ---------------- Sección 3 ----------------
R["3.1.1"] = dict(
    si=("TESTS:\n"
        "T1: ¿Cadena causal productos → resultados → impacto explicitada (no implícita)? (sí/no)\n"
        "T2: ¿Mecanismos de cambio identificados (verbo + sujeto)? (sí/no)\n"
        "T3: ¿Supuestos críticos enunciados como tales (no como riesgos)? (sí/no)\n"
        "T4: Razón texto-resultados / texto-actividades > 1. (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3 ∧ T4"),
    par=("DECISIÓN: T1 ∧ T2 ∧ T3 ∧ ¬T4   (cadena clara pero foco en actividades)"),
    no=("DECISIÓN: ¬T1   (lógica solo deducible de la lista de actividades)"),
    anclas="«teoría del cambio», «ToC», «mecanismo de cambio», «supuesto crítico», «si X entonces Y»",
    subj="Alta",
)

R["3.1.2"] = dict(
    si=("TESTS:\n"
        "T1: Cobertura — productos cubren resultados sin saltos lógicos. (sí/no)\n"
        "T2: Suficiencia — cada resultado tiene productos que justifican alcanzarlo. (sí/no)\n"
        "T3: Vínculo resultados→impacto razonado. (sí/no)\n\n"
        "DECISIÓN: T1 ∨ T2 ∨ T3   (al menos uno; reconocida subjetividad)"),
    par=("DECISIÓN: existe trazabilidad parcial pero hay saltos en ≥1 resultado"),
    no=("DECISIÓN: saltos generalizados; productos y resultados desacoplados"),
    anclas="«marco lógico», «logframe»; verbos que conectan filas",
    subj="Alta",
)

R["3.1.3"] = dict(
    si=("TESTS:\n"
        "T1: ¿Actores del desarrollo de capacidades nombrados (no «los socios»)? (sí/no)\n"
        "T2: ¿Tipo de cambio esperado especificado (desempeño / comportamiento / prácticas)? (sí/no)\n"
        "T3: ¿Mecanismo por el cual capacidad → cambio articulado? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3"),
    par=("DECISIÓN: T1 ∧ T2 ∧ ¬T3"),
    no=("DECISIÓN: solo «se desarrollarán capacidades» sin especificar"),
    anclas="«cambio de comportamiento», «mejor desempeño», «aplicación práctica»",
    subj="Media",
)

R["3.1.4"] = dict(
    si=("TESTS:\n"
        "T1: ¿Supuestos enunciados con verbo en condicional («se asume que…»)? (sí/no)\n"
        "T2: ¿Supuestos vinculados a niveles de cadena (productos/resultados/impacto)? (sí/no)\n"
        "T3: ¿Referencia a intervenciones pasadas o paralelas (OIT u otros)? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3"),
    par=("DECISIÓN: T1 ∧ ¬T2   (supuestos genéricos)"),
    no=("DECISIÓN: ¬T1   O   supuestos confundidos con riesgos"),
    anclas="«supuesto», «se asume», «condición previa», «assumption»",
    subj="Media",
)

R["3.1.5"] = dict(
    si=("TESTS:\n"
        "T1: ¿Cita evaluaciones específicas (título, año)? (sí/no)\n"
        "T2: ¿Esas evaluaciones respaldan vínculos causales del marco lógico? (sí/no)\n"
        "T3: ¿Lecciones se aplican explícitamente al diseño actual? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3"),
    par=("DECISIÓN: T1 ∧ ¬(T2 ∧ T3)"),
    no=("DECISIÓN: ¬T1"),
    anclas="«evaluación», «meta-evaluación», «ievaldiscovery», años de evaluación",
    subj="Media",
)

R["3.2.1"] = dict(
    si=("TESTS:\n"
        "T1: ¿Resultados expresan estados / situaciones («X está fortalecido», «Y tiene acceso»)? (sí/no)\n"
        "T2: Ausencia de fórmulas de actividad («mediante», «a través de», «realizando») en >80% de los resultados. (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2"),
    par=("DECISIÓN: T1 ∧ ¬T2   (mezcla de cambio y acción)"),
    no=("DECISIÓN: ¬T1   (predomina lenguaje de acción)"),
    anclas="«está fortalecido», «tiene acceso», «se ha mejorado»; ausencia de «mediante», «a través de»",
    subj="Baja",
)

R["3.2.2"] = dict(
    si=("TESTS SMART por resultado:\n"
        "S — Específico (sujeto + verbo + objeto concreto). (sí/no)\n"
        "M — Medible (indicador con umbral cuantificable). (sí/no)\n"
        "A — Alcanzable (factible con los recursos del proyecto). (sí/no)\n"
        "R — Relevante (vinculado al problema y la ToC). (sí/no)\n"
        "T — Temporal (con plazo definido o trazable al cronograma). (sí/no)\n\n"
        "DECISIÓN: TODOS los resultados cumplen S+M+A+R+T"),
    par=("DECISIÓN: resultados cumplen 3–4 atributos SMART; típicamente faltan M o T"),
    no=("DECISIÓN: resultados son aspiraciones (≤2 atributos)"),
    anclas="indicadores con cifra/%, plazos «para 2026», «al fin del proyecto»",
    subj="Media",
)

R["3.2.3"] = dict(
    si=("TESTS:\n"
        "T1: ¿Productos formulados como salidas («Documento X publicado», «N personas certificadas»)? (sí/no)\n"
        "T2: Ausencia de formulaciones de actividad («Realizar talleres», «Capacitar a…») en >80% de los productos. (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2"),
    par=("DECISIÓN: T1 ∧ ¬T2   (mezcla productos y actividades)"),
    no=("DECISIÓN: ¬T1"),
    anclas="«publicado», «certificados», «entregado»; ausencia de «realizar», «llevar a cabo»",
    subj="Baja",
)

R["3.3.1"] = dict(
    si=("TESTS (filtro DEDICADO vs MARCO):\n"
        "T1: ¿Resultado o producto cuyo título NOMBRA género/igualdad/mujeres? (sí/no)\n"
        "T2: ¿Indicador desagregado por sexo o específico de género? (sí/no)\n"
        "T3: ¿Meta cuantificable por sexo o de género? (sí/no)\n"
        "T4 (cond): ¿Inclusión de género declarada como EXPLÍCITA en el proyecto? (sí/no)\n\n"
        "DECISIÓN:\n"
        "  si T4=sí → T1 ∧ (T2 ∨ T3)\n"
        "  si T4=no → T2 ∨ T3"),
    par=("DECISIÓN: solo aparece en supuestos o lenguaje transversal del marco; sin elemento DEDICADO"),
    no=("DECISIÓN: marco lógico no menciona género ni en resultados, productos ni indicadores"),
    anclas="«igualdad de género», «desagregado por sexo», «GEM»; títulos de resultados que nombren género",
    subj="Alta",
)

R["3.3.2"] = dict(
    si=("TESTS (filtro DEDICADO vs MARCO):\n"
        "T1: ¿Resultado o producto que NOMBRA discapacidad? (sí/no)\n"
        "T2: ¿Indicador específico de discapacidad o desagregado? (sí/no)\n"
        "T3: ¿Meta cuantificable de discapacidad? (sí/no)\n"
        "T4 (cond): ¿Inclusión de discapacidad declarada como EXPLÍCITA en el proyecto? (sí/no)\n\n"
        "DECISIÓN:\n"
        "  si T4=sí → T1 ∧ (T2 ∨ T3)\n"
        "  si T4=no → T2 ∨ T3"),
    par=("DECISIÓN: mención solo a nivel de supuestos o lenguaje transversal; sin elemento DEDICADO"),
    no=("DECISIÓN: marco lógico no incluye discapacidad"),
    anclas="«inclusión de discapacidad», «desagregado por discapacidad», «PcD»",
    subj="Alta",
)

R["3.4.1"] = dict(
    si=("TESTS:\n"
        "T1: ¿Existe sección o anexo de análisis de riesgos? (sí/no)\n"
        "T2: ¿Los riesgos están categorizados (estratégico / operacional / fiduciario / contextual / reputacional)? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2"),
    par=("DECISIÓN: T1 ∧ ¬T2"),
    no=("DECISIÓN: ¬T1"),
    anclas="«matriz de riesgos», «análisis de riesgos», categorías («operacional», «fiduciario»)",
    subj="Baja",
)

R["3.4.2"] = dict(
    si=("TESTS:\n"
        "T1: ¿Riesgos específicos de EAS identificados? (sí/no)\n"
        "T2: ¿Mecanismos de prevención (código de conducta, capacitación obligatoria)? (sí/no)\n"
        "T3: ¿Mecanismo de denuncia accesible y confidencial? (sí/no)\n"
        "T4: ¿Protocolo de respuesta y atención a víctimas? (sí/no)\n"
        "T5 (cond): si hay terceros → ¿cláusulas EAS en contratos? (sí/no/NA)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3 ∧ T4 ∧ (T5 ∨ NA)"),
    par=("DECISIÓN: T1 ∧ T2 ∧ ¬(T3 ∧ T4)"),
    no=("DECISIÓN: solo enuncia «cero tolerancia» sin operacionalización"),
    anclas="«EAS», «explotación y abuso sexuales», «PSEA», «código de conducta», «mecanismo de denuncia»",
    subj="Media",
)

R["3.4.3"] = dict(
    si=("TESTS:\n"
        "T1: ¿Identifica riesgos comunitarios (no solo del personal del proyecto)? (sí/no)\n"
        "T2: ¿Medidas para riesgos físicos (tráfico / materiales peligrosos / contaminación)? (sí/no)\n"
        "T3 (cond): si hay personal de seguridad → ¿protocolo de uso de la fuerza y prevención de abusos? (sí/no/NA)\n"
        "T4: ¿Mecanismo de queja comunitario? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T4 ∧ (T3 ∨ NA)"),
    par=("DECISIÓN: cubre solo riesgos del personal, no comunitarios; o T1∧T2 ∧ ¬T4"),
    no=("DECISIÓN: ¬T1"),
    anclas="«salud y seguridad de la comunidad», «riesgos comunitarios», «mecanismo de queja comunitario»",
    subj="Media",
)

R["3.4.4"] = dict(
    si=("TESTS:\n"
        "T1: ¿Identifica riesgos de incumplimiento NIT en el contexto del proyecto? (sí/no)\n"
        "T2 (opcional): ¿Vínculo con observaciones del CEACR u organismos de supervisión? (sí/no)\n"
        "T3: ¿Plan de mitigación específico? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T3   (T2 eleva pero no es obligatorio)"),
    par=("DECISIÓN: T1 ∧ ¬T3"),
    no=("DECISIÓN: ¬T1"),
    anclas="«incumplimiento normativo», «riesgos de cumplimiento», «CEACR», convenios por número",
    subj="Media",
)

R["3.4.5"] = dict(
    si=("TESTS:\n"
        "T1: ¿Presupuesto del proyecto > 1 000 000 USD? (sí/no/NA)\n"
        "T2: ¿Registro de riesgos en formato OIT vigente, adjunto? (sí/no)\n\n"
        "DECISIÓN: si T1=sí → T2; si T1=no → N/A"),
    par=("DECISIÓN: T1 ∧ registro presente pero no es versión vigente, o no adjunto pero referenciado"),
    no=("DECISIÓN: T1 ∧ ¬T2"),
    na="DECISIÓN: ¬T1 (presupuesto ≤ 1M USD)",
    anclas="«registro de riesgos», «risk register», presupuesto total en USD",
    subj="Baja",
)

R["3.4.6"] = dict(
    si=("TESTS:\n"
        "T1: ¿Cada riesgo es contextual al proyecto (no plantilla genérica)? (sí/no)\n"
        "T2: ¿Probabilidad e impacto valorados (escala numérica o categórica)? (sí/no)\n"
        "T3: ¿Riesgos son materiales para el éxito del proyecto? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3"),
    par=("DECISIÓN: algunos riesgos específicos, otros genéricos; o falta valoración"),
    no=("DECISIÓN: riesgos son plantilla institucional sin contextualización"),
    anclas="«probabilidad alta/media/baja», «impacto», «escala 1–5»",
    subj="Alta",
)

R["3.4.7"] = dict(
    si=("TESTS, por cada riesgo nivel medio o alto:\n"
        "T1: ¿Medida de mitigación? (sí/no)\n"
        "T2: ¿Responsable de la medida? (sí/no)\n"
        "T3: ¿Método de monitoreo? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3 para TODOS los riesgos medio/alto"),
    par=("DECISIÓN: T1 universal pero T2 o T3 faltan en varios"),
    no=("DECISIÓN: riesgos medio/alto sin medidas o solo medidas genéricas"),
    anclas="columnas «medida», «responsable», «monitoreo» en matriz de riesgos",
    subj="Media",
)

R["3.5.1"] = dict(
    si=("TESTS:\n"
        "T1: ¿Sistema de recopilación descrito (qué datos, frecuencia, responsables)? (sí/no)\n"
        "T2: ¿Justificación de recursos asignados a S&E (vínculo con presupuesto)? (sí/no)\n"
        "T3 (cond): si presupuesto > umbral → ¿revisión de evaluabilidad? (sí/no/NA)\n"
        "T4 (opcional): ¿Plan de aprendizaje del proyecto? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ (T3 ∨ NA)"),
    par=("DECISIÓN: T1 ∧ ¬T2"),
    no=("DECISIÓN: «se realizará seguimiento» sin sistema descrito"),
    anclas="«sistema de S&E», «evaluabilidad», «plan de aprendizaje», «MEL»",
    subj="Baja",
)

R["3.5.2"] = dict(
    si=("TESTS:\n"
        "T1: ¿Ciclo de retroalimentación definido (datos → decisiones)? (sí/no)\n"
        "T2: ¿Necesidades de información para informes especificadas? (sí/no)\n"
        "T3: ¿Personal responsable nombrado o por puesto? (sí/no)\n"
        "T4: ¿Frecuencia de los ciclos de revisión definida? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3 ∧ T4"),
    par=("DECISIÓN: T1 ∧ T3 ∧ ¬(T2 ∧ T4)"),
    no=("DECISIÓN: ¬T1   (datos se recopilan sin uso definido)"),
    anclas="«ciclo de retroalimentación», «revisión trimestral», «informe semestral»",
    subj="Media",
)

R["3.5.3"] = dict(
    si=("TESTS:\n"
        "T1: ¿Plan de S&E como documento o sección discreta? (sí/no)\n"
        "T2: ¿Métodos de recopilación nombrados (encuesta, entrevista, registros admin.)? (sí/no)\n"
        "T3: ¿Métodos de análisis especificados? (sí/no)\n"
        "T4: ¿Roles y responsabilidades por método/indicador? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3 ∧ T4"),
    par=("DECISIÓN: T1 ∧ T2 ∧ ¬(T3 ∧ T4)"),
    no=("DECISIÓN: ¬T1   O   solo tabla de indicadores sin métodos"),
    anclas="«método de recopilación», «análisis cualitativo/cuantitativo», «roles»",
    subj="Baja",
)

R["3.5.4"] = dict(
    si=("TESTS por indicador del plan de M&E:\n"
        "T1: ¿Cumple SMART completo (S+M+A+R+T)? (sí/no)\n"
        "T2: ¿Sensible al género (desagregado por sexo o específico de género)? (sí/no)\n"
        "T3: ¿Permite inclusión de discapacidad (desagregado o específico)? (sí/no)\n"
        "T4: ¿Línea de base definida? (sí/no)\n"
        "T5: ¿Meta e hitos definidos? (sí/no)\n"
        "T6: ¿Cubre resultados (no solo productos)? (sí/no)\n\n"
        "DECISIÓN: TODOS los indicadores satisfacen T1+T2+T3+T4+T5+T6"),
    par=("DECISIÓN: T1 ∧ T4 ∧ T5 ∧ T6 pero T2/T3 son genéricos (inclusivo en lenguaje sin desagregación real)"),
    no=("DECISIÓN: indicadores no SMART, sin línea de base, o ignoran género/discapacidad"),
    anclas="«SMART», «desagregado por sexo / discapacidad», «línea de base», «meta», «hito»",
    subj="Alta",
)

R["3.5.5"] = dict(
    si=("TESTS:\n"
        "T1: ¿Partida presupuestaria de evaluación SEPARADA y nombrada? (sí/no)\n"
        "T2: ¿Monto ≥ ~2% del presupuesto total (o justificación si difiere)? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2"),
    par=("DECISIÓN: T1 ∧ ¬T2   O   monto identificable sin partida separada"),
    no=("DECISIÓN: ¬T1   (no hay partida ni monto identificable)"),
    anclas="«presupuesto de evaluación», «partida evaluación», porcentaje del total",
    subj="Baja",
)

R["3.6.1"] = dict(
    si=("TESTS:\n"
        "T1 (cond): si hay inception → ¿actividades y resultados del período inicial explicitados? (sí/no/NA)\n"
        "T2: ¿Plazo total justificado contra complejidad y capacidades? (sí/no)\n"
        "T3: ¿Cronograma con hitos por trimestre o semestre? (sí/no)\n\n"
        "DECISIÓN: T2 ∧ T3 ∧ (T1 ∨ NA)"),
    par=("DECISIÓN: T3 ∧ ¬T2   (cronograma sin justificación)"),
    no=("DECISIÓN: ¬T3"),
    anclas="«inception», «fase inicial», «cronograma», «Gantt», «hitos trimestrales»",
    subj="Alta",
)

R["3.6.2"] = dict(
    si=("TESTS:\n"
        "T1: ¿Presupuesto desglosado por actividad o producto? (sí/no)\n"
        "T2: ¿Coherencia entre actividades del marco lógico y partidas presupuestarias? (sí/no)\n"
        "T3 (opcional): ¿Costos unitarios o cálculos visibles? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2"),
    par=("DECISIÓN: T1 ∧ ¬T2"),
    no=("DECISIÓN: presupuesto solo a nivel de categorías amplias"),
    anclas="«desglose por actividad», «BoQ», «cost driver», «costo unitario»",
    subj="Media",
)

R["3.6.3"] = dict(
    si=("TESTS (verifica los aplicables al proyecto):\n"
        "T1: ¿Partida para especialista de género cuando aplica? (sí/no/NA)\n"
        "T2: ¿Partida para especialista de discapacidad cuando aplica? (sí/no/NA)\n"
        "T3: ¿Partida para accesibilidad de información (lenguaje claro, formatos alternativos)? (sí/no/NA)\n"
        "T4: ¿Partida para interpretación a idiomas locales o lengua de señas cuando aplica? (sí/no/NA)\n\n"
        "DECISIÓN: TODOS los aplicables (T1–T4) tienen partida específica con monto"),
    par=("DECISIÓN: reconoce necesidad pero queda como «si se requiere» sin partida"),
    no=("DECISIÓN: sin partidas y sin reconocimiento de costos transversales"),
    anclas="«especialista de género/discapacidad», «accesibilidad», «idioma local», «lengua de señas»",
    subj="Media",
)

R["3.7.1"] = dict(
    si=("TESTS:\n"
        "T1: ¿Mención explícita de rentabilidad o «value for money»? (sí/no)\n"
        "T2: ¿Vinculada al diseño del proyecto (no solo a la administración)? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2"),
    par=("DECISIÓN: T1 ∧ ¬T2"),
    no=("DECISIÓN: ¬T1"),
    anclas="«rentabilidad», «value for money», «relación calidad-precio», «eficiencia»",
    subj="Baja",
)

R["3.7.2"] = dict(
    si=("TESTS:\n"
        "T1: ¿Análisis de alternativas de diseño con costo/beneficio? (sí/no)\n"
        "T2: ¿Justificación de la opción elegida? (sí/no)\n"
        "T3 (cond): ¿Comparación con benchmarks (proyectos similares) cuando hay datos? (sí/no/NA)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ (T3 ∨ NA)"),
    par=("DECISIÓN: T2 ∧ ¬T1   (justifica sin comparar alternativas)"),
    no=("DECISIÓN: ¬T1 ∧ ¬T2"),
    anclas="«análisis de alternativas», «benchmark», «coste-beneficio»",
    subj="Alta",
)

# ---------------- Sección 4 ----------------
R["4.1.1"] = dict(
    si=("TESTS:\n"
        "T1: ¿Identifica la oficina/unidad gestora por nombre? (sí/no)\n"
        "T2 (cond): si proyecto regional/nacional → ¿oficina de terreno cercana a los beneficiarios? (sí/no/NA)\n"
        "T3 (cond): si gestión centralizada en sede → ¿justificación explícita por eficacia/rentabilidad/capacidad? (sí/no/NA)\n\n"
        "DECISIÓN: T1 ∧ (T2 ∨ T3)"),
    par=("DECISIÓN: T1 ∧ ¬T3 (centralización sin justificar)"),
    no=("DECISIÓN: ¬T1"),
    anclas="«oficina sobre el terreno», «sede», «descentralizado», «justificación de centralización»",
    subj="Media",
)

R["4.1.2"] = dict(
    si=("TESTS:\n"
        "T1: ¿Organigrama o descripción de roles del personal? (sí/no)\n"
        "T2: ¿Línea de rendición de cuentas explícita (a quién reporta cada quién)? (sí/no)\n"
        "T3: ¿Funcionario OIT responsable identificado por puesto o nombre? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3"),
    par=("DECISIÓN: T1 ∧ ¬(T2 ∧ T3)"),
    no=("DECISIÓN: solo «se contratará personal según se requiera»"),
    anclas="«organigrama», «reporta a», «responsable OIT», nombres/puestos específicos",
    subj="Baja",
)

R["4.1.3"] = dict(
    si=("TESTS:\n"
        "T1: ¿Identifica unidades técnicas o administrativas que apoyarán? (sí/no)\n"
        "T2: ¿Apoyo previsto en el presupuesto (partida o asignación)? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2"),
    par=("DECISIÓN: T1 ∧ ¬T2"),
    no=("DECISIÓN: asume apoyo sin previsión ni identificación"),
    anclas="«backstopping», «unidad técnica», «apoyo administrativo», nombres de unidades (PROGRAM, ENTERPRISES, …)",
    subj="Media",
)

R["4.1.4"] = dict(
    si=("TESTS:\n"
        "T1: ¿Plan de dotación de personal? (sí/no)\n"
        "T2: ¿Procedimientos de adquisiciones (estándares OIT o ajustes)? (sí/no)\n"
        "T3: ¿Sistemas financieros y de reporte? (sí/no)\n"
        "T4: ¿Niveles de autoridad y aprobación? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3 ∧ T4"),
    par=("DECISIÓN: 2–3 de T1/T2/T3/T4"),
    no=("DECISIÓN: solo «se aplicarán los procedimientos OIT»"),
    anclas="«procurement», «adquisiciones», «IRIS», «niveles de autorización»",
    subj="Media",
)

R["4.1.5"] = dict(
    si=("TESTS (si hay contratistas/subcontratistas/proveedores):\n"
        "T1: ¿Cláusulas de trabajo decente en contratos con terceros? (sí/no)\n"
        "T2: ¿Cláusulas de empleo justo (salario, jornada, libertad sindical)? (sí/no)\n"
        "T3: ¿Mecanismo de queja accesible para trabajadores de terceros? (sí/no)\n"
        "T4: ¿Seguimiento del cumplimiento por la OIT? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3 ∧ T4"),
    par=("DECISIÓN: T1 ∧ T2 ∧ ¬(T3 ∧ T4)"),
    no=("DECISIÓN: solo «los terceros cumplirán las normas»"),
    na="DECISIÓN: no hay contratistas — documentar",
    anclas="«contratista», «subcontratista», «cláusula de trabajo decente», «mecanismo de queja»",
    subj="Media",
)

R["4.2.1"] = dict(
    si=("TESTS:\n"
        "T1: ¿Roles de cada institución/socio explicitados? (sí/no)\n"
        "T2: ¿Procedimientos de ejecución (cómo trabajan juntos)? (sí/no)\n"
        "T3: ¿Justificación de selección de socios (por qué estos y no otros)? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3"),
    par=("DECISIÓN: T1 ∧ T2 ∧ ¬T3"),
    no=("DECISIÓN: solo lista socios sin roles ni justificación"),
    anclas="«selección de socios», «criterio de selección», «modus operandi»",
    subj="Media",
)

R["4.2.2"] = dict(
    si=("TESTS:\n"
        "T1: ¿Constancia explícita de aceptación por cada socio? (sí/no)\n"
        "T2: ¿Documento de respaldo (carta, MoU, acta) referenciado? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2"),
    par=("DECISIÓN: T1 ∧ ¬T2"),
    no=("DECISIÓN: ¬T1"),
    anclas="«carta de aceptación», «MoU firmado», «acta de acuerdo»",
    subj="Baja",
)

R["4.2.3"] = dict(
    si=("TESTS:\n"
        "T1: ¿Evaluación de capacidad organizativa realizada o planificada con metodología? (sí/no)\n"
        "T2: ¿Referencia explícita en la propuesta (no solo en anexo no citado)? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2"),
    par=("DECISIÓN: T1 ∧ ¬T2   (mención sin método ni resultados)"),
    no=("DECISIÓN: ¬T1"),
    anclas="«evaluación de capacidad organizativa», «OCA», «institutional assessment»",
    subj="Baja",
)

R["4.2.4"] = dict(
    si=("TESTS (cond: si la evaluación identificó brechas):\n"
        "T1: ¿Plan de desarrollo de capacidad para los socios ejecutores? (sí/no)\n"
        "T2: ¿Componentes operativos (capacitación, mentoría, acompañamiento)? (sí/no)\n"
        "T3: ¿Componente de capacidad de respuesta en género? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3"),
    par=("DECISIÓN: T1 ∧ T2 ∧ ¬T3   (género queda en marco)"),
    no=("DECISIÓN: brechas identificadas pero sin plan"),
    na="DECISIÓN: la evaluación no identificó brechas — documentar",
    anclas="«plan de fortalecimiento institucional», «capacity development plan», «gender-responsive»",
    subj="Media",
)

R["4.2.5"] = dict(
    si=("TESTS:\n"
        "T1: ¿Antecedentes de socios documentados (proyectos similares, capacidad demostrada)? (sí/no)\n"
        "T2: ¿Argumento de plausibilidad con evidencia (no «son competentes»)? (sí/no)\n"
        "T3 (cond): si capacidad demostrada es limitada → ¿mitigación específica? (sí/no/NA)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ (T3 ∨ NA)"),
    par=("DECISIÓN: T1 ∧ ¬T2   O   antecedentes ausentes pero socios reconocidos"),
    no=("DECISIÓN: solo afirma que ejecutarán bien"),
    anclas="«track record», «proyectos previos», «capacidad demostrada»",
    subj="Alta",
)

R["4.3.1"] = dict(
    si=("TESTS:\n"
        "T1: ¿Plan de comunicación con partes interesadas (qué, cuándo, cómo)? (sí/no)\n"
        "T2: ¿Productos de comunicación esperados (reportes, briefings, eventos)? (sí/no)\n"
        "T3: ¿Recursos humanos asignados? (sí/no)\n"
        "T4: ¿Recursos financieros asignados? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3 ∧ T4"),
    par=("DECISIÓN: T1 ∧ T2 ∧ ¬(T3 ∧ T4)"),
    no=("DECISIÓN: «se informará a las partes interesadas» sin operacionalización"),
    anclas="«plan de comunicación», «briefing», «informe trimestral»",
    subj="Baja",
)

R["4.4.1"] = dict(
    si=("TESTS:\n"
        "T1: ¿Estrategia de comunicación pública (no solo técnica)? (sí/no)\n"
        "T2 (alt): ¿Cronograma de desarrollo si la estrategia no está aún elaborada? (sí/no)\n"
        "T3: ¿Productos esperados (notas de prensa, redes, multimedia)? (sí/no)\n\n"
        "DECISIÓN: (T1 ∧ T3)   O   T2 con cronograma claro"),
    par=("DECISIÓN: anuncia estrategia sin cronograma ni productos"),
    no=("DECISIÓN: no incluye estrategia de comunicación pública"),
    anclas="«estrategia de comunicación», «público general», «interés humano», «human interest»",
    subj="Media",
)

R["4.4.2"] = dict(
    si=("TESTS:\n"
        "T1: ¿Recursos humanos asignados a comunicación (personal/consultores)? (sí/no)\n"
        "T2: ¿Recursos financieros asignados (partida)? (sí/no)\n"
        "T3: ¿Formatos accesibles para personas con discapacidad? (sí/no)\n"
        "T4 (cond): ¿Productos en idiomas locales / lenguaje claro / lengua de señas cuando aplica? (sí/no/NA)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3 ∧ (T4 ∨ NA)"),
    par=("DECISIÓN: T1 ∧ T2 ∧ ¬T3   (accesibilidad declarativa sin partida)"),
    no=("DECISIÓN: ¬T1 ∧ ¬T2"),
    anclas="«accesibilidad», «formatos alternativos», «lengua de señas», «idioma local»",
    subj="Media",
)

R["4.4.3"] = dict(
    si=("TESTS:\n"
        "T1: ¿Presupuesto > 5 000 000 USD? (sí/no/NA)\n"
        "T2: ¿Financiamiento por PPP (alianza público-privada) o Comisión Europea? (sí/no/NA)\n"
        "T3: si T1 ∧ T2 → ¿Plantilla DCOMM completada? (sí/no/NA)\n"
        "T4: si T1 ∧ T2 → ¿Coordinación con la oficina nacional correspondiente? (sí/no/NA)\n\n"
        "DECISIÓN: si T1 ∧ T2 → T3 ∧ T4; si ¬(T1 ∧ T2) → N/A"),
    par=("DECISIÓN: T1 ∧ T2 ∧ T3 ∧ ¬T4"),
    no=("DECISIÓN: T1 ∧ T2 ∧ ¬T3 ∧ ¬T4"),
    na="DECISIÓN: ¬T1   O   ¬T2",
    anclas="«PPP», «public-private partnership», «Comisión Europea», «DCOMM», presupuesto en USD",
    subj="Baja",
)

# ---------------- Sección 5 ----------------
R["5.1.1"] = dict(
    si=("TESTS:\n"
        "T1: ¿Todas las secciones de la plantilla OIT presentes? (sí/no)\n"
        "T2: ¿Orden conforme a la guía? (sí/no)\n"
        "T3: ¿Anexos requeridos incluidos? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3"),
    par=("DECISIÓN: 2 de T1/T2/T3"),
    no=("DECISIÓN: faltan secciones o anexos clave"),
    anclas="«Resumen ejecutivo», «Antecedentes», «Marco lógico», «Presupuesto», «Anexos» — comparar con plantilla OIT",
    subj="Baja",
)

R["5.2.1"] = dict(
    si=("TESTS:\n"
        "T1: ¿Ideas centrales identificables sin re-lectura (resumen ejecutivo claro)? (sí/no)\n"
        "T2: ¿Ausencia de jerga innecesaria o repetición evidente? (sí/no)\n"
        "T3: ¿Estructura visible (subtítulos, tablas, viñetas cuando ayudan)? (sí/no)\n\n"
        "DECISIÓN: T1 ∧ T2 ∧ T3"),
    par=("DECISIÓN: texto mayormente claro pero con secciones densas o repetitivas"),
    no=("DECISIÓN: ideas centrales sepultadas; difícil de seguir"),
    anclas="resumen ejecutivo, subtítulos, longitud de párrafos, tablas vs prosa",
    subj="Alta",
)

# ---------- Column constants ----------
COL = {
    "ID": 1, "Sec": 2, "Sub": 3, "Head": 4, "Crit": 5, "Tipo": 6,
    "Apli": 7, "Transv": 8, "Elem": 9, "Si": 10, "Par": 11, "No": 12,
    "Na": 13, "EjSi": 14, "EjPa": 15, "EjNo": 16, "Notas": 17,
    "CambiosV2": 18,
}
# New columns we will add at the end
COL["AnclasV3"] = 19
COL["SubjV3"] = 20


def main():
    shutil.copy(SRC, DST)
    wb = load_workbook(DST)
    ws = wb["Rúbrica Tab 1"]

    # Build ID → row map
    id_to_row = {}
    for r in range(3, ws.max_row + 1):
        v = ws.cell(row=r, column=COL["ID"]).value
        if v:
            id_to_row[str(v)] = r

    # Styling
    border = Border(
        left=Side(style="thin", color="888888"),
        right=Side(style="thin", color="888888"),
        top=Side(style="thin", color="888888"),
        bottom=Side(style="thin", color="888888"),
    )
    tightened_fill = PatternFill("solid", fgColor="D9E8FB")  # soft blue for tightened cells
    anclas_fill_high = PatternFill("solid", fgColor="FCE4D6")    # high subj — example useful
    anclas_fill_med = PatternFill("solid", fgColor="FFF2CC")
    anclas_fill_low = PatternFill("solid", fgColor="E2EFDA")
    cell_align = Alignment(wrap_text=True, vertical="top", horizontal="left")
    mono_font = Font(name="Menlo", size=10)

    # New column headers
    header_fill = PatternFill("solid", fgColor="002F6C")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(wrap_text=True, vertical="center", horizontal="center")
    for col_idx, header in [
        (COL["AnclasV3"], "Anclas verificables (v3)"),
        (COL["SubjV3"], "Subjetividad residual (v3)"),
    ]:
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    # Column widths for new columns
    ws.column_dimensions[get_column_letter(COL["AnclasV3"])].width = 50
    ws.column_dimensions[get_column_letter(COL["SubjV3"])].width = 16

    # Drop the now-irrelevant example columns: replace placeholder content with "N/A (v3)"
    # but keep columns for compatibility with downstream parsers; mark them inactive.
    inactive_fill = PatternFill("solid", fgColor="F2F2F2")
    inactive_font = Font(color="888888", italic=True, size=9)
    for col_key in ("EjSi", "EjPa", "EjNo"):
        for r in range(3, ws.max_row + 1):
            cell = ws.cell(row=r, column=COL[col_key], value="(v3) — ejemplos no requeridos para criterios de baja/media subjetividad; pendiente bootstrap para alta")
            cell.fill = inactive_fill
            cell.font = inactive_font
            cell.alignment = cell_align

    # Apply mechanical rubrics
    applied = 0
    missing = []
    for crit_id, row in id_to_row.items():
        spec = R.get(crit_id)
        if not spec:
            missing.append(crit_id)
            continue
        applied += 1

        # Replace Sí / Parcial / No
        for col_key, content in (("Si", spec.get("si", "")),
                                  ("Par", spec.get("par", "")),
                                  ("No", spec.get("no", ""))):
            cell = ws.cell(row=row, column=COL[col_key], value=content)
            cell.alignment = cell_align
            cell.fill = tightened_fill
            cell.font = mono_font
            cell.border = border

        # N/A only when defined
        if spec.get("na"):
            cell = ws.cell(row=row, column=COL["Na"], value=spec["na"])
            cell.alignment = cell_align
            cell.fill = tightened_fill
            cell.font = mono_font
            cell.border = border

        # Anchors
        cell = ws.cell(row=row, column=COL["AnclasV3"], value=spec.get("anclas", ""))
        cell.alignment = cell_align
        cell.border = border
        cell.font = Font(size=10)

        # Subjetividad
        subj = spec.get("subj", "—")
        cell = ws.cell(row=row, column=COL["SubjV3"], value=subj)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = border
        cell.font = Font(bold=True, size=11)
        if subj == "Alta":
            cell.fill = anclas_fill_high
        elif subj == "Media":
            cell.fill = anclas_fill_med
        else:
            cell.fill = anclas_fill_low

        # Bump row height
        ws.row_dimensions[row].height = 320

    print(f"Aplicado: {applied} criterios. Faltantes: {missing or 'ninguno'}")

    # Update instructions
    instr = (
        "RÚBRICA Tab 1 — v3 (mecanizada). "
        "Cada criterio se evalúa mediante TESTS atómicos (T1, T2, ...) y una DECISIÓN booleana. "
        "Columna «Anclas verificables» = patrones de texto / códigos / nombres concretos a buscar en el documento. "
        "Columna «Subjetividad residual» = Baja/Media/Alta — los criterios «Alta» son los que más se beneficiarían "
        "de ejemplos calibradores si se ejecuta el bootstrap (opción B con repositorio de PRODOCs). "
        "Las columnas de Ejemplo evidencia están desactivadas en v3."
    )
    ws["A1"] = instr
    try:
        ws.unmerge_cells(start_row=1, start_column=1, end_row=1, end_column=18)
    except Exception:
        pass
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=COL["SubjV3"])

    wb.save(DST)
    print(f"Guardado: {DST}")

    # Summary stats
    counts = {"Baja": 0, "Media": 0, "Alta": 0}
    for spec in R.values():
        counts[spec.get("subj", "Media")] = counts.get(spec.get("subj"), 0) + 1
    print("\nSubjetividad residual:")
    for k, v in counts.items():
        print(f"  {k}: {v} criterios")


if __name__ == "__main__":
    main()
